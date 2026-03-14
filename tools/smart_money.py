#!/usr/bin/env python3
"""
Smart Money Graph v4.0 — Self-Expanding OSINT Engine.

4-layer architecture: Identifier Backbone → Data Ingest → Signal Engine → Decision Integration.
Tracks fund holdings (13F), short positions (FCA/AMF/CONSOB/AFM-NL), insider transactions.
Detects actionable signals, discovers overlooked opportunities, integrates with decision pipeline.
Self-expanding: each refresh --expand grows the graph with newly discovered stocks and funds.

v4.0: Self-expanding refresh, OpenFIGI identifier fallback, fund discovery, island pruning,
      CONSOB/AFM-NL download helpers, insider expansion to universe stocks.
v3.0: Signal engine, discovery engine, live ingest, sector overlay, EU regulator expansion.
v2.0: Universal identifier resolution, automated ingest, one-command refresh.

Usage:
  # OSINT ENGINE (v3.0+)
  python3 tools/smart_money.py signals [--ticker T] [--portfolio-only]  # Detect actionable signals
  python3 tools/smart_money.py discover [--source 13f|fca|amf|all] [--min-funds N]  # Anti-bias discovery
  python3 tools/smart_money.py discover-funds [--min-stocks N]          # Find untracked funds (quarterly)
  python3 tools/smart_money.py sector-overlay SECTOR                    # Institutional positioning by sector
  python3 tools/smart_money.py ingest-live --type {holder|short|insider|mention} --fund F --ticker T [--data JSON]
  python3 tools/smart_money.py capture Elliott holds 5.2% LULU       # Quick-capture (v4.1)

  # EU REGULATOR EXPANSION (v3.0)
  python3 tools/smart_money.py ingest-consob [--file PATH]    # CONSOB XLSX → graph (IT shorts)
  python3 tools/smart_money.py ingest-afm-nl [--file PATH]    # AFM NL CSV → graph (NL shorts)

  # ONE-COMMAND OPERATION (v4.0)
  python3 tools/smart_money.py refresh [--full] [--expand] [--skip-download]
  python3 tools/smart_money.py coverage [--portfolio-only]

  # IDENTIFIER RESOLUTION (v4.0)
  python3 tools/smart_money.py resolve [TICKER ...] [--force] [--retry-failed] [--purge-invalid]

  # AUTOMATED INGEST (v4.0)
  python3 tools/smart_money.py ingest-fca [--file PATH]       # FCA XLSX → graph (UK shorts)
  python3 tools/smart_money.py ingest-amf                     # AMF CSV → graph (FR shorts)
  python3 tools/smart_money.py ingest-13f                     # 13F XMLs → graph (US holdings)
  python3 tools/smart_money.py ingest-insider [TICKER ...] [--universe] [--all-enrolled]

  # ENROLLMENT (v2.0)
  python3 tools/smart_money.py enroll [TICKER ...] [--from-universe] [--from-pipeline]

  # DATA ACQUISITION
  python3 tools/smart_money.py stale
  python3 tools/smart_money.py download {fca|amf|consob|afm-nl|shorts|13f|form4|all}
  python3 tools/smart_money.py parse-fca [--file PATH]
  python3 tools/smart_money.py parse-13f
  python3 tools/smart_money.py filter-13f FUND_CIK
  python3 tools/smart_money.py short-interest

  # GRAPH MUTATION
  python3 tools/smart_money.py bulk-update
  python3 tools/smart_money.py add-node TYPE ID [--attr K=V ...]
  python3 tools/smart_money.py add-edge FROM TO RELATION [--attr K=V ...]
  python3 tools/smart_money.py sync-portfolio

  # QUERIES
  python3 tools/smart_money.py report [--portfolio-only]
  python3 tools/smart_money.py stock-profile TICKER
  python3 tools/smart_money.py who-holds TICKER
  python3 tools/smart_money.py crowding [--top N]
  python3 tools/smart_money.py alerts

  # ANALYTICS
  python3 tools/smart_money.py metrics
  python3 tools/smart_money.py communities

  # MAINTENANCE
  python3 tools/smart_money.py snapshot
  python3 tools/smart_money.py visualize [--portfolio-only] [--output FILE]
  python3 tools/smart_money.py stats
  python3 tools/smart_money.py gc  # Prunes old files + island stock nodes
"""

import argparse
import json
import os
import sys
import shutil
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import date, datetime, timedelta

import networkx as nx
import yaml

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(SCRIPT_DIR, "smart_money")
GRAPH_PATH = os.path.join(DATA_DIR, "graph.json")
SOURCES_PATH = os.path.join(DATA_DIR, "sources.yaml")
TRACKED_FUNDS_PATH = os.path.join(DATA_DIR, "tracked_funds.yaml")
CUSIP_MAP_PATH = os.path.join(DATA_DIR, "cusip_map.yaml")
IDENTIFIERS_PATH = os.path.join(DATA_DIR, "identifiers.yaml")
PORTFOLIO_PATH = os.path.join(PROJECT_ROOT, "portfolio", "current.yaml")
SHORTS_DIR = os.path.join(DATA_DIR, "data", "shorts")
HOLDINGS_DIR = os.path.join(DATA_DIR, "data", "holdings")
INSIDERS_DIR = os.path.join(DATA_DIR, "data", "insiders")
SNAPSHOTS_DIR = os.path.join(DATA_DIR, "data", "snapshots")

EDGAR_HEADERS = {
    "User-Agent": "ValueInvestorCLI/1.0 (educational@example.com)",
    "Accept": "application/json",
}


# ---------------------------------------------------------------------------
# Init — ensure directories exist
# ---------------------------------------------------------------------------

def ensure_dirs():
    for d in [DATA_DIR, SHORTS_DIR, HOLDINGS_DIR, INSIDERS_DIR, SNAPSHOTS_DIR]:
        os.makedirs(d, exist_ok=True)


def init_sources():
    """Create sources.yaml if missing."""
    if not os.path.exists(SOURCES_PATH):
        sources = {
            "fca_uk_shorts": {"last_download": None, "cadence_days": 3, "path": None},
            "amf_france_shorts": {"last_download": None, "cadence_days": 3, "path": None},
            "sec_13f": {"last_download": None, "cadence_days": 90, "path": None},
            "sec_form4": {"last_download": None, "cadence_days": 30, "path": None},
        }
        with open(SOURCES_PATH, "w") as f:
            yaml.dump(sources, f, default_flow_style=False)
    with open(SOURCES_PATH) as f:
        return yaml.safe_load(f) or {}


def save_sources(sources):
    with open(SOURCES_PATH, "w") as f:
        yaml.dump(sources, f, default_flow_style=False)


# ---------------------------------------------------------------------------
# Graph I/O
# ---------------------------------------------------------------------------

def load_graph():
    """Load graph from JSON (node_link format). Create empty if missing."""
    if os.path.exists(GRAPH_PATH):
        with open(GRAPH_PATH) as f:
            data = json.load(f)
        return nx.node_link_graph(data, directed=True, multigraph=True)
    return nx.MultiDiGraph()


def save_graph(G):
    """Save graph as node_link JSON."""
    data = nx.node_link_data(G)
    with open(GRAPH_PATH, "w") as f:
        json.dump(data, f, indent=2)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def today_str():
    return date.today().isoformat()


def days_since(date_str):
    """Days since a date string. Returns None if date_str is None."""
    if not date_str:
        return None
    try:
        d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").date()
        return (date.today() - d).days
    except (ValueError, TypeError):
        return None


def staleness_label(days, cadence):
    if days is None:
        return "NEVER"
    if days <= cadence:
        return "FRESH"
    if days <= cadence * 2:
        return "STALE"
    return "VERY_STALE"


def get_portfolio_tickers():
    """Load tickers from portfolio/current.yaml."""
    if not os.path.exists(PORTFOLIO_PATH):
        return []
    with open(PORTFOLIO_PATH) as f:
        pf = yaml.safe_load(f) or {}
    positions = pf.get("positions", [])
    return [p["ticker"] for p in positions if p.get("ticker")]


def get_portfolio_positions():
    """Load full position data from portfolio/current.yaml."""
    if not os.path.exists(PORTFOLIO_PATH):
        return []
    with open(PORTFOLIO_PATH) as f:
        pf = yaml.safe_load(f) or {}
    return pf.get("positions", [])


SUFFIX_TO_COUNTRY = {
    ".L": "UK", ".PA": "FR", ".DE": "DE", ".AS": "NL", ".MI": "IT",
    ".HE": "FI", ".MC": "ES", ".ST": "SE", ".SW": "CH", ".BR": "BE",
    ".CO": "DK", ".OL": "NO", ".LS": "PT", ".VI": "AT",
}


def ticker_country(ticker):
    """Infer country from ticker suffix."""
    for suffix, country in SUFFIX_TO_COUNTRY.items():
        if ticker.endswith(suffix):
            return country
    return "US"


# Valid ISIN prefixes per country (most stocks, some exceptions for dual-listed)
COUNTRY_TO_ISIN_PREFIX = {
    "UK": ["GB", "JE", "GG", "IM"],  # Jersey, Guernsey, Isle of Man listings
    "FR": ["FR"], "DE": ["DE"], "NL": ["NL"], "IT": ["IT"],
    "FI": ["FI"], "ES": ["ES"], "SE": ["SE"], "CH": ["CH"],
    "BE": ["BE"], "DK": ["DK"], "NO": ["NO"], "PT": ["PT"], "AT": ["AT"],
    "US": ["US", "IE", "GB", "NL", "CH", "CA", "BM", "KY", "VG", "PA"],
}


def _is_valid_isin_for_ticker(isin, ticker):
    """Validate ISIN country prefix matches ticker's expected country.
    Returns True if valid or if we can't determine (benefit of the doubt)."""
    if not isin or len(isin) < 2:
        return False
    country = ticker_country(ticker)
    valid_prefixes = COUNTRY_TO_ISIN_PREFIX.get(country)
    if valid_prefixes is None:
        return True  # Unknown country, accept
    return isin[:2] in valid_prefixes


def nodes_by_type(G, node_type):
    """Get all nodes of a given type."""
    return {n: d for n, d in G.nodes(data=True) if d.get("type") == node_type}


def edges_by_relation(G, relation):
    """Get all edges of a given relation type."""
    results = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == relation:
            results.append((u, v, k, d))
    return results


# ---------------------------------------------------------------------------
# Identifier Resolution System
# ---------------------------------------------------------------------------

def load_identifiers():
    """Load identifiers.yaml cache."""
    if os.path.exists(IDENTIFIERS_PATH):
        with open(IDENTIFIERS_PATH) as f:
            data = yaml.safe_load(f) or {}
    else:
        data = {}
    if "identifiers" not in data:
        data["identifiers"] = {}
    if "failed_resolutions" not in data:
        data["failed_resolutions"] = []
    return data


def save_identifiers(data):
    """Save identifiers.yaml cache."""
    with open(IDENTIFIERS_PATH, "w") as f:
        yaml.dump(data, f, default_flow_style=False, sort_keys=False)


# ---------------------------------------------------------------------------
# Fund Alias / Deduplication System
# ---------------------------------------------------------------------------

FUND_ALIASES_PATH = os.path.join(DATA_DIR, "fund_aliases.yaml")

def load_fund_aliases():
    """Load fund_aliases.yaml. Returns dict: alternate_slug -> canonical_slug."""
    if not os.path.exists(FUND_ALIASES_PATH):
        return {}
    with open(FUND_ALIASES_PATH) as f:
        data = yaml.safe_load(f) or {}
    lookup = {}
    for canonical, alternates in data.get("aliases", {}).items():
        for alt in (alternates or []):
            lookup[alt] = canonical
    return lookup


def canonicalize_fund_slug(slug, aliases=None):
    """Return canonical slug if alias exists, otherwise return original."""
    if aliases is None:
        aliases = load_fund_aliases()
    return aliases.get(slug, slug)


def cmd_dedup_funds(args):
    """Scan graph for duplicate fund nodes matching aliases. Merge edges to canonical."""
    aliases = load_fund_aliases()
    if not aliases:
        print("No fund_aliases.yaml or no aliases defined.")
        return

    G = load_graph()
    merged_count = 0
    removed_nodes = []

    for alt_slug, canonical_slug in aliases.items():
        if alt_slug not in G.nodes:
            continue
        if G.nodes[alt_slug].get("type") != "fund":
            continue

        # Ensure canonical node exists
        if canonical_slug not in G.nodes:
            G.add_node(canonical_slug, type="fund",
                       full_name=G.nodes[alt_slug].get("full_name", canonical_slug),
                       fund_type=G.nodes[alt_slug].get("fund_type", "unknown"))

        # Move all edges from alt to canonical
        # Outgoing edges (fund -> stock)
        for _, target, key, data in list(G.edges(alt_slug, data=True, keys=True)):
            G.add_edge(canonical_slug, target, **data)
            merged_count += 1

        # Incoming edges (rare, but handle)
        for source, _, key, data in list(G.in_edges(alt_slug, data=True, keys=True)):
            G.add_edge(source, canonical_slug, **data)
            merged_count += 1

        G.remove_node(alt_slug)
        removed_nodes.append(alt_slug)

    if removed_nodes:
        save_graph(G)
        print(f"Dedup: merged {len(removed_nodes)} duplicate fund nodes, moved {merged_count} edges")
        for node in removed_nodes:
            canonical = aliases[node]
            print(f"  {node} -> {canonical}")
    else:
        print("Dedup: no duplicate fund nodes found matching aliases.")


def _resolve_isin_yfinance(ticker):
    """Try to get ISIN from yfinance. Validates country prefix to reject garbage."""
    try:
        import yfinance as yf
        t = yf.Ticker(ticker)
        isin = getattr(t, 'isin', None)
        if isin and isinstance(isin, str) and len(isin) >= 12 and isin != "-":
            if _is_valid_isin_for_ticker(isin, ticker):
                return isin
    except Exception:
        pass
    return None


def _resolve_isin_from_fca(ticker, issuer_name=None):
    """Try to find ISIN for a UK stock from downloaded FCA XLSX."""
    if not ticker.endswith(".L"):
        return None
    try:
        import openpyxl
    except ImportError:
        return None

    # Find most recent FCA file
    files = sorted([f for f in os.listdir(SHORTS_DIR) if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
    if not files:
        return None

    fpath = os.path.join(SHORTS_DIR, files[0])
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None

    # Find header row
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = str(row).lower()
        if "position holder" in row_str or "issuer" in row_str:
            header_idx = i
            break

    headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
    isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
    issuer_col = next((i for i, h in enumerate(headers) if "issuer" in h or "name of share" in h), None)

    if isin_col is None or issuer_col is None:
        return None

    # Get stock name from graph or portfolio for matching
    G = load_graph()
    stock_name = ""
    if ticker in G.nodes:
        stock_name = G.nodes[ticker].get("name", "").upper()
    if issuer_name:
        stock_name = issuer_name.upper()

    # Try to match by name
    if stock_name:
        # Clean name for matching (remove PLC, LTD, GROUP, etc.)
        clean_name = stock_name.replace(" PLC", "").replace(" LTD", "").replace(" GROUP", "").replace(" LIMITED", "").strip()
        for row in rows[header_idx + 1:]:
            if row[issuer_col] and clean_name.lower() in str(row[issuer_col]).lower():
                isin = str(row[isin_col]).strip()
                if len(isin) >= 12:
                    return isin
    return None


def _resolve_isin_from_amf(ticker, issuer_name=None):
    """Try to find ISIN for a French stock from downloaded AMF CSV."""
    if not ticker.endswith(".PA"):
        return None

    files = sorted([f for f in os.listdir(SHORTS_DIR) if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
    if not files:
        return None

    fpath = os.path.join(SHORTS_DIR, files[0])

    G = load_graph()
    stock_name = ""
    if ticker in G.nodes:
        stock_name = G.nodes[ticker].get("name", "").upper()
    if issuer_name:
        stock_name = issuer_name.upper()

    if not stock_name:
        return None

    # Clean name
    clean_parts = stock_name.replace(" SA", "").replace(" SE", "").replace(" S.A.", "").strip().split()
    search_term = clean_parts[0].lower() if clean_parts else ""
    if not search_term:
        return None

    try:
        with open(fpath, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                if search_term in line.lower():
                    parts = line.strip().split(";")
                    if len(parts) >= 5:
                        isin = parts[4].strip().strip('"')
                        if isin.startswith("FR") and len(isin) >= 12:
                            return isin
    except Exception:
        pass
    return None


def _resolve_cik_from_edgar(ticker):
    """Get CIK from cached EDGAR company_tickers.json or download it."""
    edgar_cache = os.path.join(DATA_DIR, "data", "edgar_tickers.json")

    # Download if not cached or stale (>90 days)
    if not os.path.exists(edgar_cache) or days_since(
            datetime.fromtimestamp(os.path.getmtime(edgar_cache)).strftime("%Y-%m-%d")) > 90:
        url = "https://www.sec.gov/files/company_tickers.json"
        req = urllib.request.Request(url, headers=EDGAR_HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = resp.read()
                os.makedirs(os.path.dirname(edgar_cache), exist_ok=True)
                with open(edgar_cache, "wb") as f:
                    f.write(data)
        except Exception:
            if not os.path.exists(edgar_cache):
                return None

    try:
        with open(edgar_cache) as f:
            tickers_data = json.load(f)
    except Exception:
        return None

    # company_tickers.json: {"0": {"cik_str": 320193, "ticker": "AAPL", "title": "Apple Inc."}, ...}
    for entry in tickers_data.values():
        if entry.get("ticker", "").upper() == ticker.upper():
            return str(entry["cik_str"])
    return None


EXCHANGE_SUFFIX_TO_MIC = {
    ".DE": "XETR",
    ".AS": "XAMS",
    ".MI": "XMIL",
    ".MC": "XMAD",
    ".HE": "XHEL",
    ".ST": "XSTO",
    ".SW": "XSWX",
    ".BR": "XBRU",
    ".L": "XLON",
    ".PA": "XPAR",
    ".CO": "XCSE",
    ".OL": "XOSL",
    ".LS": "XLIS",
    ".VI": "XWBO",
}


def _resolve_isin_openfigi(ticker):
    """Fallback: confirm stock exists on exchange via OpenFIGI.

    NOTE: OpenFIGI does NOT return ISINs. This function confirms the stock
    exists on the expected exchange and returns the compositeFIGI as metadata.
    For actual ISIN resolution, use harvest_isins_from_regulators().
    Free tier: 25 requests/min, batch up to 10.
    """
    url = "https://api.openfigi.com/v3/mapping"

    suffix = ""
    for s in EXCHANGE_SUFFIX_TO_MIC:
        if ticker.endswith(s):
            suffix = s
            break

    if not suffix:
        return None

    mic = EXCHANGE_SUFFIX_TO_MIC[suffix]
    base_ticker = ticker[: -len(suffix)]

    payload = json.dumps([{
        "idType": "TICKER",
        "idValue": base_ticker,
        "exchCode": mic,
        "securityType2": "Common Stock",
    }]).encode("utf-8")

    req = urllib.request.Request(url, data=payload, headers={
        "Content-Type": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            results = json.loads(resp.read())
        if results and isinstance(results, list) and results[0].get("data"):
            for item in results[0]["data"]:
                sec_type = item.get("securityType2", "")
                if sec_type in ("Common Stock", "Depositary Receipt"):
                    return item.get("compositeFIGI")
    except Exception:
        pass
    return None


# _figi_to_isin REMOVED: OpenFIGI does not return ISINs. Dead code since v2.0.
# ISIN resolution now uses: yfinance → FCA/AMF/CONSOB/AFM-NL harvest → manual CUSIP.


def _resolve_isin_openfigi_batch(tickers):
    """Batch resolve ISINs via OpenFIGI. Up to 10 tickers per request.

    Returns dict: ticker -> ISIN (or None).
    Since OpenFIGI doesn't return ISINs directly, this uses the FIGI
    confirmation + country prefix to attempt ISIN construction, and
    falls back to confirming the stock exists on the expected exchange.
    """
    import time
    url = "https://api.openfigi.com/v3/mapping"
    results_map = {}

    # Process in batches of 10
    for i in range(0, len(tickers), 10):
        batch = tickers[i:i + 10]
        jobs = []
        batch_tickers = []

        for ticker in batch:
            suffix = ""
            for s in EXCHANGE_SUFFIX_TO_MIC:
                if ticker.endswith(s):
                    suffix = s
                    break
            if not suffix:
                results_map[ticker] = None
                continue

            mic = EXCHANGE_SUFFIX_TO_MIC[suffix]
            base_ticker = ticker[: -len(suffix)]
            jobs.append({
                "idType": "TICKER",
                "idValue": base_ticker,
                "exchCode": mic,
            })
            batch_tickers.append(ticker)

        if not jobs:
            continue

        payload = json.dumps(jobs).encode("utf-8")
        req = urllib.request.Request(url, data=payload, headers={
            "Content-Type": "application/json",
        })
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                api_results = json.loads(resp.read())
            for j, ticker in enumerate(batch_tickers):
                if j < len(api_results):
                    entry = api_results[j]
                    if entry.get("data"):
                        # Found on exchange — extract any available ISIN-like data
                        for item in entry["data"]:
                            sec_type = item.get("securityType2", "")
                            if sec_type in ("Common Stock", "Depositary Receipt", ""):
                                # OpenFIGI confirms stock exists. Store compositeFIGI as fallback.
                                results_map[ticker] = {
                                    "figi": item.get("compositeFIGI"),
                                    "name": item.get("name", ""),
                                    "ticker_confirmed": item.get("ticker", ""),
                                    "exchange": item.get("exchCode", ""),
                                }
                                break
                        else:
                            results_map[ticker] = None
                    else:
                        results_map[ticker] = None
        except Exception:
            for ticker in batch_tickers:
                results_map[ticker] = None

        if i + 10 < len(tickers):
            time.sleep(2.5)  # Rate limit: 25 req/min

    return results_map


# ---------------------------------------------------------------------------
# ISIN Check Digit + CUSIP→ISIN conversion
# ---------------------------------------------------------------------------

def _isin_check_digit(isin_without_check):
    """Compute ISIN check digit (Luhn mod-10 on alpha-numeric expansion)."""
    digits = []
    for ch in isin_without_check:
        if ch.isdigit():
            digits.append(int(ch))
        elif ch.isalpha():
            val = ord(ch.upper()) - ord('A') + 10
            digits.append(val // 10)
            digits.append(val % 10)
    # Luhn on the digit string
    total = 0
    for i, d in enumerate(reversed(digits)):
        if i % 2 == 0:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return str((10 - (total % 10)) % 10)


def _cusip_to_isin(cusip, country="US"):
    """Convert a 9-digit CUSIP to a 12-char ISIN with check digit."""
    base = country + cusip
    return base + _isin_check_digit(base)


# Known CUSIPs for US stocks where yfinance ISIN resolution failed
MANUAL_CUSIPS = {
    "MA": "57636Q104",
    "FICO": "303308102",
    "MMC": "571748102",
    "WRB": "084423102",
    "MNST": "61174X109",
    "CL": "194162103",
    "HSY": "427866108",
    "LRCX": "512807108",
    "SYK": "863667101",
    "PODD": "45765Y101",
    "TDG": "893641100",
    "HEI": "422806109",
    "HWM": "443052104",
    "CTAS": "177340102",
    "CMCSA": "20030N101",
    "SHW": "824348106",
    "DECK": "243537107",
    "IDXX": "45168D104",
    "COKE": "191098102",
}


# ---------------------------------------------------------------------------
# ISIN Harvester — extract ISINs from downloaded FCA/AMF files
# ---------------------------------------------------------------------------

def _clean_name_variants(name):
    """Generate name variants for matching, ordered by specificity (most specific first)."""
    if not name:
        return []
    name = name.upper().strip()
    import re as _re
    # Strip share descriptions like "ORD 38 6/13P", "ORD EUR0.01", etc.
    name = _re.sub(r'\s+ORD\b.*$', '', name)
    # Suffixes to strip
    suffixes = [" PLC", " LTD", " LIMITED", " GROUP", " HOLDINGS",
                " SA", " SE", " S.A.", " S.E.", " NV", " N.V.",
                " AG", " GMBH", " INC", " CORP", " CO", " CORPORATION"]
    variants = [name]
    stripped = name
    for sfx in suffixes:
        stripped = stripped.replace(sfx, "")
    stripped = stripped.strip()
    if stripped != name:
        variants.append(stripped)
    # For multi-word names: first two words as higher-specificity variant
    words = stripped.split()
    if len(words) >= 2:
        variants.append(" ".join(words[:2]))
    # Single-word after cleaning: include it. Multi-word: first word as lowest priority
    if len(words) == 1:
        variants.append(words[0])
    return variants


def _match_name_to_isin(stock_name, isin_map, country_prefix):
    """Find the best ISIN match for a stock name against an {isin: issuer_name} map.

    Returns (isin, score) or (None, 0). Higher score = better match.
    Score: 100 = exact full-name match, 80 = cleaned-name match,
           60 = two-word match, 40 = substring of 6+ chars.
    """
    variants = _clean_name_variants(stock_name)
    if not variants:
        return None, 0

    best_isin = None
    best_score = 0

    for isin, issuer in isin_map.items():
        if not isin.startswith(country_prefix):
            continue
        issuer_variants = _clean_name_variants(issuer)
        if not issuer_variants:
            continue

        score = 0
        # Exact full name match (highest confidence)
        if variants[0] == issuer_variants[0]:
            score = 100
        # Cleaned name match
        elif len(variants) > 1 and len(issuer_variants) > 1 and variants[1] == issuer_variants[1]:
            score = 80
        # Two-word prefix match
        elif (len(variants) > 2 and len(issuer_variants) > 2
              and variants[2] == issuer_variants[2] and len(variants[2]) >= 6):
            score = 60
        # Substring match — only for long enough cleaned names to avoid "DASSAULT" matching wrong
        elif len(variants) > 1 and len(issuer_variants) > 1:
            our_clean = variants[1]  # cleaned name without suffixes
            their_clean = issuer_variants[1]
            if len(our_clean) >= 6 and our_clean == their_clean:
                score = 75
            elif len(our_clean) >= 8 and our_clean in issuer_variants[0]:
                score = 40

        if score > best_score:
            best_score = score
            best_isin = isin

    return best_isin, best_score


def harvest_isins_from_regulators():
    """Harvest ISINs from FCA XLSX, AMF CSV, AFM-NL CSV, and CONSOB XLSX
    for unresolved tickers.

    These regulatory files contain ISINs for every shorted stock in
    UK/France/Netherlands/Italy — we use them as ISIN databases.
    Returns dict of {ticker: isin} for newly resolved tickers.
    """
    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})
    G = load_graph()

    # Collect unresolved tickers by country
    unresolved_uk = []
    unresolved_fr = []
    unresolved_nl = []
    unresolved_it = []
    for ticker, info in ident.items():
        if info.get("isin"):
            continue
        country = info.get("country", ticker_country(ticker))
        if country == "UK":
            unresolved_uk.append(ticker)
        elif country == "FR":
            unresolved_fr.append(ticker)
        elif country == "NL":
            unresolved_nl.append(ticker)
        elif country == "IT":
            unresolved_it.append(ticker)

    newly_resolved = {}

    # --- FCA XLSX: harvest all ISINs for UK tickers ---
    if unresolved_uk:
        try:
            import openpyxl
        except ImportError:
            print("  openpyxl not installed, skipping FCA harvest")
            unresolved_uk = []

    if unresolved_uk:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
        if files:
            fpath = os.path.join(SHORTS_DIR, files[0])
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            # Find header
            header_idx = 0
            for i, row in enumerate(rows):
                row_str = str(row).lower()
                if "position holder" in row_str or "issuer" in row_str:
                    header_idx = i
                    break

            headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
            isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
            issuer_col = next((i for i, h in enumerate(headers) if "issuer" in h or "name of share" in h), None)

            if isin_col is not None and issuer_col is not None:
                # Build ISIN→issuer_name index (GB ISINs only)
                fca_isin_map = {}  # {isin: issuer_name}
                for row in rows[header_idx + 1:]:
                    if all(v is None for v in row):
                        continue
                    isin = str(row[isin_col]).strip() if row[isin_col] else ""
                    issuer = str(row[issuer_col]).strip() if row[issuer_col] else ""
                    if isin.startswith("GB") and len(isin) >= 12 and issuer:
                        if isin not in fca_isin_map:
                            fca_isin_map[isin] = issuer

                # For each unresolved UK ticker, try to match against FCA issuers
                for ticker in unresolved_uk:
                    stock_name = G.nodes[ticker].get("name", "") if ticker in G.nodes else ""
                    if not stock_name:
                        try:
                            import yfinance as yf
                            t = yf.Ticker(ticker)
                            info_data = t.info or {}
                            stock_name = info_data.get("shortName", "") or info_data.get("longName", "")
                        except Exception:
                            pass

                    if not stock_name:
                        continue

                    matched_isin, score = _match_name_to_isin(stock_name, fca_isin_map, "GB")
                    if matched_isin and score >= 40:
                        newly_resolved[ticker] = matched_isin

    # --- AMF CSV: harvest all ISINs for FR tickers ---
    if unresolved_fr:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
        if files:
            fpath = os.path.join(SHORTS_DIR, files[0])

            # Build ISIN→issuer_name index (FR ISINs only)
            amf_isin_map = {}  # {isin: issuer_name}
            try:
                with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                    f.readline()  # skip header
                    for line in f:
                        parts = line.strip().split(";")
                        if len(parts) >= 5:
                            issuer = parts[2].strip().strip('"')
                            isin = parts[4].strip().strip('"')
                            if isin.startswith("FR") and len(isin) >= 12 and issuer:
                                if isin not in amf_isin_map:
                                    amf_isin_map[isin] = issuer
            except Exception:
                amf_isin_map = {}

            for ticker in unresolved_fr:
                stock_name = G.nodes[ticker].get("name", "") if ticker in G.nodes else ""
                if not stock_name:
                    try:
                        import yfinance as yf
                        t = yf.Ticker(ticker)
                        info_data = t.info or {}
                        stock_name = info_data.get("shortName", "") or info_data.get("longName", "")
                    except Exception:
                        pass

                if not stock_name:
                    continue

                matched_isin, score = _match_name_to_isin(stock_name, amf_isin_map, "FR")
                if matched_isin and score >= 40:
                    newly_resolved[ticker] = matched_isin

    # --- AFM-NL CSV: harvest all ISINs for NL tickers ---
    if unresolved_nl:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("afm_nl_") and f.endswith(".csv")], reverse=True)
        if files:
            fpath = os.path.join(SHORTS_DIR, files[0])

            afm_nl_isin_map = {}  # {isin: issuer_name}
            try:
                with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                    header_line = f.readline()
                    delimiter = ";" if ";" in header_line else ","
                    for line in f:
                        parts = line.strip().split(delimiter)
                        if len(parts) < 4:
                            continue
                        issuer = parts[1].strip().strip('"') if len(parts) > 1 else ""
                        # Find ISIN column (NL prefix, length 12)
                        isin = ""
                        for part in parts:
                            clean = part.strip().strip('"')
                            if len(clean) == 12 and clean[:2].isalpha() and clean[2:].replace(" ", "").isalnum():
                                isin = clean
                                break
                        if isin.startswith("NL") and len(isin) >= 12 and issuer:
                            if isin not in afm_nl_isin_map:
                                afm_nl_isin_map[isin] = issuer
            except Exception:
                afm_nl_isin_map = {}

            for ticker in unresolved_nl:
                stock_name = G.nodes[ticker].get("name", "") if ticker in G.nodes else ""
                if not stock_name:
                    try:
                        import yfinance as yf
                        t = yf.Ticker(ticker)
                        info_data = t.info or {}
                        stock_name = info_data.get("shortName", "") or info_data.get("longName", "")
                    except Exception:
                        pass

                if not stock_name:
                    continue

                matched_isin, score = _match_name_to_isin(stock_name, afm_nl_isin_map, "NL")
                if matched_isin and score >= 40:
                    newly_resolved[ticker] = matched_isin

    # --- CONSOB XLSX: harvest all ISINs for IT tickers ---
    if unresolved_it:
        try:
            import openpyxl
        except ImportError:
            unresolved_it = []

    if unresolved_it:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("consob_italy_") and f.endswith(".xlsx")], reverse=True)
        if files:
            fpath = os.path.join(SHORTS_DIR, files[0])
            import openpyxl
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            header_idx = 0
            for i, row in enumerate(rows):
                row_str = str(row).lower()
                if any(w in row_str for w in ["soggetto", "position holder", "emittente", "issuer"]):
                    header_idx = i
                    break

            headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
            isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
            issuer_col = next((i for i, h in enumerate(headers)
                               if any(w in h for w in ["emittente", "issuer", "name of share"])), None)

            if isin_col is not None and issuer_col is not None:
                consob_isin_map = {}  # {isin: issuer_name}
                for row in rows[header_idx + 1:]:
                    if all(v is None for v in row):
                        continue
                    isin = str(row[isin_col]).strip() if row[isin_col] else ""
                    issuer = str(row[issuer_col]).strip() if row[issuer_col] else ""
                    if isin.startswith("IT") and len(isin) >= 12 and issuer:
                        if isin not in consob_isin_map:
                            consob_isin_map[isin] = issuer

                for ticker in unresolved_it:
                    stock_name = G.nodes[ticker].get("name", "") if ticker in G.nodes else ""
                    if not stock_name:
                        try:
                            import yfinance as yf
                            t = yf.Ticker(ticker)
                            info_data = t.info or {}
                            stock_name = info_data.get("shortName", "") or info_data.get("longName", "")
                        except Exception:
                            pass

                    if not stock_name:
                        continue

                    matched_isin, score = _match_name_to_isin(stock_name, consob_isin_map, "IT")
                    if matched_isin and score >= 40:
                        newly_resolved[ticker] = matched_isin

    return newly_resolved


def cmd_harvest_isins(args):
    """Harvest ISINs from FCA/AMF regulatory files and update identifiers.yaml."""
    print("=== HARVESTING ISINs FROM REGULATORY FILES ===\n")
    identifiers_data = load_identifiers()

    newly_resolved = harvest_isins_from_regulators()

    if not newly_resolved:
        print("No new ISINs resolved from regulatory files.")
        return

    # Apply to identifiers
    for ticker, isin in sorted(newly_resolved.items()):
        existing = identifiers_data["identifiers"].get(ticker, {})
        existing["isin"] = isin
        existing["country"] = existing.get("country", ticker_country(ticker))
        existing["resolved_at"] = today_str()
        existing["source"] = "regulatory_harvest"
        identifiers_data["identifiers"][ticker] = existing

        # Remove from failed_resolutions
        if ticker in identifiers_data["failed_resolutions"]:
            identifiers_data["failed_resolutions"].remove(ticker)

        print(f"  {ticker:12s}  RESOLVED  ISIN={isin}")

    save_identifiers(identifiers_data)
    print(f"\nNewly resolved: {len(newly_resolved)} tickers.")


def resolve_ticker(ticker, identifiers_data, force=False):
    """Resolve identifiers for a single ticker. Returns dict of resolved fields."""
    cache = identifiers_data["identifiers"]
    existing = cache.get(ticker, {})

    # Check if cached and fresh (within 90 days)
    if not force and existing.get("isin") and existing.get("resolved_at"):
        age = days_since(existing["resolved_at"])
        if age is not None and age < 90:
            return existing

    country = ticker_country(ticker)
    result = {"country": country, "resolved_at": today_str()}

    # Step 1: yfinance ISIN
    import time
    isin = _resolve_isin_yfinance(ticker)
    if isin:
        result["isin"] = isin
    time.sleep(0.3)

    # Step 2: FCA fallback for UK
    if not result.get("isin") and country == "UK":
        isin = _resolve_isin_from_fca(ticker)
        if isin:
            result["isin"] = isin

    # Step 3: AMF fallback for FR
    if not result.get("isin") and country == "FR":
        isin = _resolve_isin_from_amf(ticker)
        if isin:
            result["isin"] = isin

    # Step 3b: Manual CUSIP fallback for US stocks where yfinance fails
    if not result.get("isin") and country == "US" and ticker in MANUAL_CUSIPS:
        cusip = MANUAL_CUSIPS[ticker]
        result["cusip"] = cusip
        result["isin"] = _cusip_to_isin(cusip)
        result["source"] = "manual_cusip"

    # Step 3c: Regulatory harvest fallback for EU stocks where yfinance fails
    if not result.get("isin") and country not in ("US",):
        harvested = harvest_isins_from_regulators()
        if ticker in harvested:
            result["isin"] = harvested[ticker]
            result["source"] = "regulatory_harvest"

    # Step 3d: OpenFIGI for exchange confirmation (stores FIGI metadata, not ISIN)
    if not result.get("isin") and country not in ("US",):
        suffix = ""
        for s in EXCHANGE_SUFFIX_TO_MIC:
            if ticker.endswith(s):
                suffix = s
                break
        if suffix:
            figi = _resolve_isin_openfigi(ticker)
            if figi:
                result["figi"] = figi
                result["source"] = "openfigi_confirmed"

    # Step 4: Derive CUSIP from ISIN for US stocks
    if result.get("isin") and country == "US" and not result.get("cusip"):
        result["cusip"] = result["isin"][2:11]

    # Step 5: CIK from EDGAR for US stocks
    if country == "US":
        cik = _resolve_cik_from_edgar(ticker)
        if cik:
            result["cik"] = cik
        time.sleep(0.2)

    # Record
    if result.get("isin"):
        cache[ticker] = result
        # Remove from failed_resolutions if previously failed
        if ticker in identifiers_data["failed_resolutions"]:
            identifiers_data["failed_resolutions"].remove(ticker)
    else:
        # Try to keep any existing data
        if existing.get("isin"):
            existing["resolved_at"] = today_str()
            cache[ticker] = existing
        else:
            # Mark as failed
            if ticker not in identifiers_data["failed_resolutions"]:
                identifiers_data["failed_resolutions"].append(ticker)
            result["isin"] = None
            cache[ticker] = result

    return cache.get(ticker, result)


# ---------------------------------------------------------------------------
# CMD: resolve
# ---------------------------------------------------------------------------

def cmd_resolve(args):
    """Resolve ISIN/CUSIP/CIK for tickers and cache in identifiers.yaml."""
    identifiers_data = load_identifiers()

    if getattr(args, "purge_invalid", False):
        # Scan all entries and purge ISINs that fail country-prefix validation
        cache = identifiers_data["identifiers"]
        purged = []
        for ticker, info in cache.items():
            isin = info.get("isin")
            if isin and not _is_valid_isin_for_ticker(isin, ticker):
                purged.append((ticker, isin))
                info["isin"] = None
                if ticker not in identifiers_data["failed_resolutions"]:
                    identifiers_data["failed_resolutions"].append(ticker)
        save_identifiers(identifiers_data)
        print(f"=== PURGE INVALID ISINs ===\n")
        if purged:
            for ticker, old_isin in purged:
                country = ticker_country(ticker)
                print(f"  {ticker:12s}  PURGED  {old_isin}  (prefix {old_isin[:2]} invalid for {country})")
            print(f"\nPurged {len(purged)} invalid ISINs. Run --retry-failed to re-resolve.")
        else:
            print("  No invalid ISINs found.")
        return

    if getattr(args, "retry_failed", False):
        # Retry only previously failed tickers
        tickers = list(identifiers_data.get("failed_resolutions", []))
        if not tickers:
            print("No failed resolutions to retry.")
            return
        print(f"=== RETRYING FAILED RESOLUTIONS ({len(tickers)} tickers) ===\n")
        # Force re-resolve for these
        args.force = True
    elif args.tickers:
        tickers = args.tickers
    else:
        # Default: all stocks in graph
        G = load_graph()
        tickers = sorted(n for n, d in G.nodes(data=True) if d.get("type") == "stock")

    if not tickers:
        print("No tickers to resolve. Run sync-portfolio first.")
        return

    if not getattr(args, "retry_failed", False):
        print(f"=== RESOLVING IDENTIFIERS ({len(tickers)} tickers) ===\n")
    resolved = 0
    failed = 0

    for ticker in tickers:
        existing = identifiers_data["identifiers"].get(ticker, {})
        if not args.force and existing.get("isin") and existing.get("resolved_at"):
            age = days_since(existing["resolved_at"])
            if age is not None and age < 90:
                isin = existing.get("isin", "")
                cusip = existing.get("cusip", "")
                cik = existing.get("cik", "")
                print(f"  {ticker:12s}  CACHED  ISIN={isin}  CUSIP={cusip or '-'}  CIK={cik or '-'}")
                resolved += 1
                continue

        result = resolve_ticker(ticker, identifiers_data, force=args.force)
        isin = result.get("isin", "")
        cusip = result.get("cusip", "")
        cik = result.get("cik", "")
        status = "OK" if isin else "FAIL"
        print(f"  {ticker:12s}  {status:6s}  ISIN={isin or '?'}  CUSIP={cusip or '-'}  CIK={cik or '-'}")
        if isin:
            resolved += 1
        else:
            failed += 1

    save_identifiers(identifiers_data)
    print(f"\nResolved: {resolved}/{len(tickers)}. Failed: {failed}.")
    if identifiers_data["failed_resolutions"]:
        print(f"Failed tickers: {', '.join(identifiers_data['failed_resolutions'])}")


# ---------------------------------------------------------------------------
# CMD: stale
# ---------------------------------------------------------------------------

def cmd_stale(args):
    sources = init_sources()
    print("=== DATA SOURCE STALENESS ===\n")
    any_stale = False
    for name, info in sorted(sources.items()):
        last = info.get("last_download")
        cadence = info.get("cadence_days", 7)
        days = days_since(last)
        label = staleness_label(days, cadence)
        days_str = f"{days}d ago" if days is not None else "never"
        marker = "" if label == "FRESH" else f" ← {label}"
        print(f"  {name:25s}  last: {str(last or 'never'):12s}  ({days_str})  cadence: {cadence}d{marker}")
        if label != "FRESH":
            any_stale = True
    if not any_stale:
        print("\n  All sources FRESH.")
    print()


# ---------------------------------------------------------------------------
# CMD: stats
# ---------------------------------------------------------------------------

def cmd_stats(args):
    G = load_graph()
    sources = init_sources()
    stocks = nodes_by_type(G, "stock")
    funds = nodes_by_type(G, "fund")
    persons = nodes_by_type(G, "person")
    portfolio_stocks = {n for n, d in stocks.items() if d.get("in_portfolio")}

    print("=== SMART MONEY GRAPH STATS ===\n")
    print(f"  Nodes:  {G.number_of_nodes()} total")
    print(f"    stocks:  {len(stocks)} ({len(portfolio_stocks)} in portfolio)")
    print(f"    funds:   {len(funds)}")
    print(f"    persons: {len(persons)}")
    print(f"  Edges:  {G.number_of_edges()} total")

    # Edge breakdown
    rel_counts = defaultdict(int)
    for u, v, d in G.edges(data=True):
        rel_counts[d.get("relation", "unknown")] += 1
    if rel_counts:
        print("    by relation:")
        for rel, cnt in sorted(rel_counts.items(), key=lambda x: -x[1]):
            print(f"      {rel:20s}  {cnt}")

    print(f"\n  Graph file: {GRAPH_PATH}")
    print(f"  Size: {os.path.getsize(GRAPH_PATH) / 1024:.1f} KB" if os.path.exists(GRAPH_PATH) else "  Size: 0 KB (no file)")

    # Source staleness summary
    print("\n  Sources:")
    for name, info in sorted(sources.items()):
        last = info.get("last_download")
        days = days_since(last)
        label = staleness_label(days, info.get("cadence_days", 7))
        print(f"    {name:25s}  {label}")
    print()


# ---------------------------------------------------------------------------
# CMD: sync-portfolio
# ---------------------------------------------------------------------------

def get_standing_orders():
    """Load standing orders from state/standing_orders.yaml."""
    so_path = os.path.join(PROJECT_ROOT, "state", "standing_orders.yaml")
    if not os.path.exists(so_path):
        return []
    with open(so_path) as f:
        data = yaml.safe_load(f) or {}
    return data.get("standing_orders", [])


def cmd_sync_portfolio(args):
    G = load_graph()
    positions = get_portfolio_positions()
    standing_orders = get_standing_orders()

    # Mark all existing stocks as not in portfolio / not SO first
    for n, d in G.nodes(data=True):
        if d.get("type") == "stock":
            d["in_portfolio"] = False
            d["has_standing_order"] = False
            d["so_category"] = None
            d["so_action"] = None

    added = 0
    updated = 0
    for pos in positions:
        ticker = pos["ticker"]
        if ticker in G.nodes:
            G.nodes[ticker]["in_portfolio"] = True
            G.nodes[ticker]["name"] = pos.get("name", "")
            G.nodes[ticker]["conviction"] = pos.get("conviction", "")
            updated += 1
        else:
            country = ticker_country(ticker)
            G.add_node(ticker, type="stock", name=pos.get("name", ""),
                       sector="", country=country, in_portfolio=True,
                       in_universe=False, conviction=pos.get("conviction", ""))
            added += 1

    # Sync standing orders
    so_added = 0
    so_tickers_seen = set()
    for so in standing_orders:
        ticker = so.get("ticker", "")
        if not ticker or ticker in so_tickers_seen:
            continue
        so_tickers_seen.add(ticker)
        category = so.get("category", "GATED")
        action = so.get("action", "BUY")
        tier = so.get("tier", "")
        trigger = so.get("trigger", "")
        fv = so.get("fair_value", "")

        if ticker in G.nodes:
            G.nodes[ticker]["has_standing_order"] = True
            G.nodes[ticker]["so_category"] = category
            G.nodes[ticker]["so_action"] = action
            G.nodes[ticker]["so_trigger"] = trigger
            G.nodes[ticker]["so_tier"] = tier
            G.nodes[ticker]["so_fair_value"] = fv
        else:
            country = ticker_country(ticker)
            G.add_node(ticker, type="stock", name="",
                       sector="", country=country, in_portfolio=False,
                       in_universe=False, has_standing_order=True,
                       so_category=category, so_action=action,
                       so_trigger=trigger, so_tier=tier, so_fair_value=fv)
            so_added += 1

    save_graph(G)
    total_so = len(so_tickers_seen)
    print(f"Synced portfolio: {added} added, {updated} updated, {len(positions)} positions.")
    print(f"Synced standing orders: {so_added} new, {total_so} total SO tickers ({total_so - so_added} already in graph).")


# ---------------------------------------------------------------------------
# CMD: add-node
# ---------------------------------------------------------------------------

def cmd_add_node(args):
    G = load_graph()
    node_type = args.node_type
    node_id = args.node_id
    attrs = {"type": node_type}
    if args.attr:
        for kv in args.attr:
            if "=" in kv:
                k, v = kv.split("=", 1)
                # Try to parse numbers and booleans
                if v.lower() == "true":
                    v = True
                elif v.lower() == "false":
                    v = False
                else:
                    try:
                        v = float(v) if "." in v else int(v)
                    except ValueError:
                        pass
                attrs[k] = v

    if node_id in G.nodes:
        G.nodes[node_id].update(attrs)
        print(f"Updated node: {node_id} (type={node_type})")
    else:
        G.add_node(node_id, **attrs)
        print(f"Added node: {node_id} (type={node_type})")

    save_graph(G)


# ---------------------------------------------------------------------------
# CMD: add-edge
# ---------------------------------------------------------------------------

def cmd_add_edge(args):
    G = load_graph()
    from_node = args.from_node
    to_node = args.to_node
    relation = args.relation

    edge_attrs = {"relation": relation, "date_added": today_str()}
    if args.attr:
        for kv in args.attr:
            if "=" in kv:
                k, v = kv.split("=", 1)
                if v.lower() == "true":
                    v = True
                elif v.lower() == "false":
                    v = False
                else:
                    try:
                        v = float(v) if "." in v else int(v)
                    except ValueError:
                        pass
                edge_attrs[k] = v

    # Ensure nodes exist
    if from_node not in G.nodes:
        print(f"Warning: creating node {from_node} (no type set)")
        G.add_node(from_node)
    if to_node not in G.nodes:
        print(f"Warning: creating node {to_node} (no type set)")
        G.add_node(to_node)

    # Remove existing edges with same relation between these nodes
    keys_to_remove = []
    for k, d in G[from_node].get(to_node, {}).items():
        if d.get("relation") == relation:
            keys_to_remove.append(k)
    for k in keys_to_remove:
        G.remove_edge(from_node, to_node, key=k)

    G.add_edge(from_node, to_node, **edge_attrs)
    print(f"Added edge: {from_node} --[{relation}]--> {to_node}")
    save_graph(G)


# ---------------------------------------------------------------------------
# CMD: bulk-update
# ---------------------------------------------------------------------------

def cmd_bulk_update(args):
    """Apply batch operations from JSON stdin.

    Expected format: list of operations:
    [
      {"op": "add_node", "type": "fund", "id": "greenlight", "attrs": {"full_name": "..."}},
      {"op": "add_edge", "from": "greenlight", "to": "ADBE", "relation": "holds",
       "attrs": {"shares": 100000, "value_usd": 50000000}},
      {"op": "remove_edge", "from": "X", "to": "Y", "relation": "holds"},
      {"op": "remove_node", "id": "X"}
    ]
    """
    G = load_graph()

    if sys.stdin.isatty():
        print("Error: bulk-update expects JSON on stdin.")
        print('Example: echo \'[{"op":"add_node","type":"fund","id":"test","attrs":{}}]\' | python3 tools/smart_money.py bulk-update')
        sys.exit(1)

    raw = sys.stdin.read().strip()
    if not raw:
        print("Error: empty stdin.")
        sys.exit(1)

    try:
        ops = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}")
        sys.exit(1)

    if not isinstance(ops, list):
        ops = [ops]

    added_nodes = 0
    added_edges = 0
    removed_nodes = 0
    removed_edges = 0

    for op in ops:
        action = op.get("op", "")
        if action == "add_node":
            node_id = op["id"]
            attrs = {"type": op.get("type", "unknown")}
            attrs.update(op.get("attrs", {}))
            if node_id in G.nodes:
                G.nodes[node_id].update(attrs)
            else:
                G.add_node(node_id, **attrs)
                added_nodes += 1

        elif action == "add_edge":
            from_n = op["from"]
            to_n = op["to"]
            relation = op["relation"]
            edge_attrs = {"relation": relation, "date_added": today_str()}
            edge_attrs.update(op.get("attrs", {}))

            # Ensure nodes exist
            if from_n not in G.nodes:
                G.add_node(from_n, type="unknown")
            if to_n not in G.nodes:
                G.add_node(to_n, type="unknown")

            # Replace existing edge with same relation
            keys_to_remove = []
            if G.has_node(from_n) and G.has_node(to_n):
                for k, d in G[from_n].get(to_n, {}).items():
                    if d.get("relation") == relation:
                        keys_to_remove.append(k)
            for k in keys_to_remove:
                G.remove_edge(from_n, to_n, key=k)

            G.add_edge(from_n, to_n, **edge_attrs)
            added_edges += 1

        elif action == "remove_edge":
            from_n = op["from"]
            to_n = op["to"]
            relation = op.get("relation")
            if G.has_node(from_n) and G.has_node(to_n):
                keys_to_remove = []
                for k, d in G[from_n].get(to_n, {}).items():
                    if relation is None or d.get("relation") == relation:
                        keys_to_remove.append(k)
                for k in keys_to_remove:
                    G.remove_edge(from_n, to_n, key=k)
                    removed_edges += 1

        elif action == "remove_node":
            node_id = op["id"]
            if node_id in G.nodes:
                G.remove_node(node_id)
                removed_nodes += 1

    save_graph(G)
    print(f"Bulk update: +{added_nodes} nodes, +{added_edges} edges, -{removed_nodes} nodes, -{removed_edges} edges")


# ---------------------------------------------------------------------------
# CMD: snapshot
# ---------------------------------------------------------------------------

def cmd_snapshot(args):
    if not os.path.exists(GRAPH_PATH):
        print("No graph to snapshot.")
        return
    fname = f"graph_{today_str()}.json"
    dest = os.path.join(SNAPSHOTS_DIR, fname)
    shutil.copy2(GRAPH_PATH, dest)
    # Count existing snapshots
    snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
    print(f"Snapshot saved: {fname} ({len(snaps)} total)")


# ---------------------------------------------------------------------------
# CMD: gc
# ---------------------------------------------------------------------------

def cmd_gc(args):
    """Clean up old snapshots (>30 days), raw data files, and island stock nodes."""
    cutoff = date.today() - timedelta(days=30)
    removed = 0

    # Clean snapshots
    if os.path.exists(SNAPSHOTS_DIR):
        for f in os.listdir(SNAPSHOTS_DIR):
            if f.startswith("graph_") and f.endswith(".json"):
                try:
                    d = datetime.strptime(f[6:16], "%Y-%m-%d").date()
                    if d < cutoff:
                        os.remove(os.path.join(SNAPSHOTS_DIR, f))
                        removed += 1
                except ValueError:
                    pass

    # Clean old raw files (shorts >60 days)
    for subdir in [SHORTS_DIR]:
        if os.path.exists(subdir):
            for f in os.listdir(subdir):
                fpath = os.path.join(subdir, f)
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).date()
                if mtime < cutoff:
                    os.remove(fpath)
                    removed += 1

    print(f"GC files: {removed} files removed.")

    # Prune island stock nodes (0 edges, not in portfolio/universe/standing orders, >30 days old)
    G = load_graph()
    portfolio_tickers = set(get_portfolio_tickers())

    # Load standing order tickers
    so_tickers = set()
    so_path = os.path.join(PROJECT_ROOT, "state", "standing_orders.yaml")
    if os.path.exists(so_path):
        try:
            with open(so_path) as f:
                so_data = yaml.safe_load(f) or {}
            for order in so_data.get("standing_orders", []):
                t = order.get("ticker")
                if t:
                    so_tickers.add(t)
        except Exception:
            pass

    # Load universe tickers
    universe_tickers = set()
    for upath in [os.path.join(PROJECT_ROOT, "state", "quality_universe.yaml"),
                  os.path.join(DATA_DIR, "data", "quality_universe.yaml")]:
        if os.path.exists(upath):
            try:
                with open(upath) as f:
                    udata = yaml.safe_load(f) or {}
                companies = udata.get("quality_universe", {}).get("companies", [])
                for s in companies:
                    t = s.get("ticker")
                    if t:
                        universe_tickers.add(t)
                break
            except Exception:
                continue

    protected = portfolio_tickers | so_tickers | universe_tickers
    pruned = []

    for node, data in list(G.nodes(data=True)):
        if data.get("type") != "stock":
            continue
        if node in protected:
            continue

        # Check edges (in + out)
        in_edges = list(G.in_edges(node))
        out_edges = list(G.out_edges(node))
        if in_edges or out_edges:
            continue  # Has connections, keep

        # Check enrollment age
        enrolled = data.get("enrolled_date") or data.get("date_added")
        if enrolled:
            age = days_since(enrolled)
            if age is not None and age < 30:
                continue  # Too young to prune

        pruned.append(node)

    for node in pruned:
        G.remove_node(node)

    if pruned:
        save_graph(G)
        print(f"GC islands: pruned {len(pruned)} island stock nodes:")
        for node in sorted(pruned)[:20]:
            print(f"  - {node}")
        if len(pruned) > 20:
            print(f"  ... and {len(pruned) - 20} more")
    else:
        print(f"GC islands: no island stock nodes to prune.")


# ---------------------------------------------------------------------------
# CMD: download
# ---------------------------------------------------------------------------

BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def download_file(url, dest, headers=None):
    """Download a file with progress indication."""
    h = dict(BROWSER_HEADERS)
    if headers:
        h.update(headers)
    req = urllib.request.Request(url, headers=h)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
            with open(dest, "wb") as f:
                f.write(data)
            return len(data)
    except urllib.error.HTTPError as e:
        print(f"  HTTP Error {e.code}: {e.reason}")
        return None
    except urllib.error.URLError as e:
        print(f"  URL Error: {e.reason}")
        return None
    except Exception as e:
        print(f"  Error: {e}")
        return None


def cmd_download(args):
    source = args.source
    sources = init_sources()

    if source in ("fca", "all"):
        print("Downloading FCA UK short positions...")
        url = "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx"
        fname = f"fca_uk_{today_str()}.xlsx"
        dest = os.path.join(SHORTS_DIR, fname)
        size = download_file(url, dest)
        if size:
            print(f"  Saved: {dest} ({size / 1024:.0f} KB)")
            sources["fca_uk_shorts"]["last_download"] = today_str()
            sources["fca_uk_shorts"]["path"] = dest
        else:
            print("  Failed to download FCA data.")

    if source in ("amf", "all"):
        print("Downloading AMF France short positions...")
        url = "https://www.data.gouv.fr/api/1/datasets/r/c2539d1c-8531-4937-9cba-3bd8e9786cc5"
        fname = f"amf_france_{today_str()}.csv"
        dest = os.path.join(SHORTS_DIR, fname)
        size = download_file(url, dest)
        if size:
            print(f"  Saved: {dest} ({size / 1024:.0f} KB)")
            sources["amf_france_shorts"]["last_download"] = today_str()
            sources["amf_france_shorts"]["path"] = dest
        else:
            print("  Failed to download AMF data.")

    if source in ("13f", "all"):
        print("Downloading 13F filings for tracked funds...")
        if not os.path.exists(TRACKED_FUNDS_PATH):
            print("  No tracked_funds.yaml found. Create it first.")
        else:
            with open(TRACKED_FUNDS_PATH) as f:
                tracked = yaml.safe_load(f) or {}
            funds = tracked.get("funds", [])
            downloaded = 0
            for fund in funds:
                cik = str(fund.get("cik", ""))
                name = fund.get("name", cik)
                if not cik:
                    continue
                print(f"  Fetching 13F for {name} (CIK {cik})...")
                success = download_13f_for_cik(cik, name)
                if success:
                    downloaded += 1
                import time
                time.sleep(0.5)  # EDGAR rate limit
            if downloaded:
                sources["sec_13f"]["last_download"] = today_str()
            print(f"  Downloaded {downloaded}/{len(funds)} fund filings.")

    if source in ("consob", "shorts", "all"):
        print("Downloading CONSOB Italy short positions...")
        # CONSOB does not offer a direct programmatic download URL.
        # Data must be downloaded manually from the web interface.
        consob_dest = os.path.join(SHORTS_DIR, f"consob_italy_{today_str()}.xlsx")
        print("  CONSOB requires manual download:")
        print("  1. Go to: https://www.consob.it/web/area-pubblica/pnc")
        print("  2. Click 'Scarica' / 'Download' for the current positions file")
        print(f"  3. Save as: {consob_dest}")
        # Update freshness if file already exists from today
        if os.path.exists(consob_dest):
            print(f"  Found today's file: {consob_dest}")
            sources.setdefault("consob_italy_shorts", {"cadence_days": 7})
            sources["consob_italy_shorts"]["last_download"] = today_str()
            sources["consob_italy_shorts"]["path"] = consob_dest

    if source in ("afm-nl", "shorts", "all"):
        print("Downloading AFM Netherlands short positions...")
        afm_nl_url = "https://www.afm.nl/~/profmedia/files/registers/netto-shortposities/netto-shortposities-actueel.csv"
        fname = f"afm_nl_{today_str()}.csv"
        dest = os.path.join(SHORTS_DIR, fname)
        size = download_file(afm_nl_url, dest)
        if size:
            print(f"  Saved: {dest} ({size / 1024:.0f} KB)")
            sources.setdefault("afm_nl_shorts", {"cadence_days": 7})
            sources["afm_nl_shorts"]["last_download"] = today_str()
            sources["afm_nl_shorts"]["path"] = dest
        else:
            print("  Failed to download AFM NL data programmatically.")
            print("  Manual download:")
            print("  1. Go to: https://www.afm.nl/en/sector/registers/meldingenregisters/netto-shortposities-actueel")
            print("  2. Download the CSV file")
            print(f"  3. Save as: {dest}")

    if source == "form4":
        # DEPRECATED: Form 4 downloads files nothing reads. Insider data via yfinance in cmd_ingest_insider.
        print("Form 4 download DEPRECATED. Use 'ingest-insider' instead (uses yfinance).")
        print("Skipping. Run: python3 tools/smart_money.py ingest-insider --universe")

    save_sources(sources)


def download_13f_for_cik(cik, name=""):
    """Download latest 13F information table for a given CIK from EDGAR."""
    padded = cik.zfill(10)
    submissions_url = f"https://data.sec.gov/submissions/CIK{padded}.json"

    req = urllib.request.Request(submissions_url, headers=EDGAR_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
    except Exception as e:
        print(f"    Error fetching submissions: {e}")
        return False

    # Find latest 13F-HR in recent filings
    filings = data.get("filings", {}).get("recent", {})
    forms = filings.get("form", [])
    accessions = filings.get("accessionNumber", [])
    dates = filings.get("filingDate", [])

    filing_idx = None
    for i, form in enumerate(forms):
        if "13F" in form.upper() and "HR" in form.upper():
            filing_idx = i
            break
    # Fallback to any 13F
    if filing_idx is None:
        for i, form in enumerate(forms):
            if "13F" in form.upper():
                filing_idx = i
                break

    if filing_idx is None:
        print(f"    No 13F found for CIK {cik}")
        return False

    accession_raw = accessions[filing_idx]
    accession = accession_raw.replace("-", "")
    filing_date = dates[filing_idx]

    # Fetch filing index to find infotable document
    import time
    time.sleep(0.2)
    index_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession}/index.json"
    req2 = urllib.request.Request(index_url, headers=EDGAR_HEADERS)
    try:
        with urllib.request.urlopen(req2, timeout=15) as resp2:
            index_data = json.loads(resp2.read())
    except Exception as e:
        print(f"    Error fetching filing index: {e}")
        return False

    # Find the infotable XML (usually contains "infotable" in filename)
    infotable_doc = None
    items = index_data.get("directory", {}).get("item", [])
    for item in items:
        fname = item.get("name", "").lower()
        if "infotable" in fname and fname.endswith(".xml"):
            infotable_doc = item["name"]
            break
    # Fallback: look for any XML that's not the primary doc
    if not infotable_doc:
        for item in items:
            fname = item.get("name", "").lower()
            if fname.endswith(".xml") and "primary" not in fname:
                infotable_doc = item["name"]
                break

    if not infotable_doc:
        print(f"    No infotable XML found in filing index for CIK {cik}")
        return False

    doc_url = f"https://www.sec.gov/Archives/edgar/data/{cik}/{accession}/{infotable_doc}"
    slug = name.lower().replace(" ", "-").replace(",", "")[:30] if name else cik
    dest = os.path.join(HOLDINGS_DIR, f"13f_{slug}_{filing_date}.xml")
    time.sleep(0.2)
    size = download_file(doc_url, dest, EDGAR_HEADERS)
    if size:
        print(f"    Saved: {os.path.basename(dest)} ({size / 1024:.0f} KB, filed {filing_date}, table: {infotable_doc})")
        return True
    return False


def download_form4_for_ticker(ticker):
    """DEPRECATED: downloads files nothing reads. Insider data via yfinance in cmd_ingest_insider."""
    # Use EDGAR XBRL company search for insider filings
    search_url = f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker}%22&dateRange=custom&startdt={date.today() - timedelta(days=90)}&enddt={date.today()}&forms=4"

    # Simpler approach: use EDGAR company search API
    search_url = f"https://efts.sec.gov/LATEST/search-index?q={ticker}&forms=4&dateRange=custom&startdt={(date.today() - timedelta(days=90)).isoformat()}&enddt={date.today().isoformat()}"

    # Actually, let's use the full-text search API
    api_url = f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker}%22&forms=4"
    dest = os.path.join(INSIDERS_DIR, f"form4_{ticker}_{today_str()}.json")

    # Use the EDGAR full-text search
    search_api = f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker}%22&forms=4&dateRange=custom&startdt={(date.today() - timedelta(days=90)).isoformat()}&enddt={date.today().isoformat()}"

    req = urllib.request.Request(search_api, headers=EDGAR_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = resp.read()
            with open(dest, "wb") as f:
                f.write(data)
            print(f"    Saved: {os.path.basename(dest)} ({len(data) / 1024:.0f} KB)")
            return True
    except Exception as e:
        print(f"    Error: {e}")
        return False


# ---------------------------------------------------------------------------
# CMD: parse-fca
# ---------------------------------------------------------------------------

def cmd_parse_fca(args):
    """Parse FCA XLSX and output as CSV to stdout for Claude to read."""
    import openpyxl

    # Find the file to parse
    if args.file:
        fpath = args.file
    else:
        # Find most recent FCA file
        files = sorted([f for f in os.listdir(SHORTS_DIR) if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
        if not files:
            print("No FCA XLSX files found. Run: python3 tools/smart_money.py download fca", file=sys.stderr)
            sys.exit(1)
        fpath = os.path.join(SHORTS_DIR, files[0])

    print(f"# Parsing: {os.path.basename(fpath)}", file=sys.stderr)

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty spreadsheet.", file=sys.stderr)
        return

    # Find header row (contains "Position Holder" or "Issuer" etc)
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = str(row).lower()
        if "position holder" in row_str or "issuer" in row_str or "net short" in row_str:
            header_idx = i
            break

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]
    print(",".join(f'"{h}"' for h in headers))

    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            elif isinstance(v, datetime):
                vals.append(v.strftime("%Y-%m-%d"))
            else:
                vals.append(str(v).replace('"', '""'))
        print(",".join(f'"{v}"' for v in vals))

    wb.close()


# ---------------------------------------------------------------------------
# CMD: filter-13f
# ---------------------------------------------------------------------------

def cmd_filter_13f(args):
    """Download and display 13F for a specific fund CIK."""
    cik = args.cik
    print(f"Fetching 13F for CIK {cik}...")
    success = download_13f_for_cik(cik, f"cik-{cik}")
    if success:
        sources = init_sources()
        sources["sec_13f"]["last_download"] = today_str()
        save_sources(sources)


# ---------------------------------------------------------------------------
# CMD: parse-13f
# ---------------------------------------------------------------------------

def cmd_parse_13f(args):
    """Auto-parse ALL downloaded 13F XML files, match CUSIPs, output bulk-update JSON.

    Reads cusip_map.yaml for target CUSIPs, scans all XML in data/holdings/,
    aggregates shares/value across sub-managers (Markel has 8 entries per stock),
    infers fund slug from filename CIK + tracked_funds.yaml.

    Usage: python3 tools/smart_money.py parse-13f | python3 tools/smart_money.py bulk-update
    """
    import xml.etree.ElementTree as ET
    import re

    # Load CUSIP map
    if not os.path.exists(CUSIP_MAP_PATH):
        print("No cusip_map.yaml found.", file=sys.stderr)
        sys.exit(1)
    with open(CUSIP_MAP_PATH) as f:
        cusip_data = yaml.safe_load(f) or {}
    cusip_to_ticker = cusip_data.get("cusip_to_ticker", {})
    if not cusip_to_ticker:
        print("cusip_map.yaml has no cusip_to_ticker entries.", file=sys.stderr)
        sys.exit(1)

    # Load tracked funds for CIK → metadata and slug → metadata mapping
    cik_to_fund = {}
    slug_to_fund = {}
    if os.path.exists(TRACKED_FUNDS_PATH):
        with open(TRACKED_FUNDS_PATH) as f:
            tf = yaml.safe_load(f) or {}
        for fund in tf.get("funds", []):
            cik = str(fund.get("cik", ""))
            if cik:
                cik_to_fund[cik] = fund
            # Build slug from fund name (same logic as download_13f_for_cik)
            name = fund.get("name", "")
            if name:
                slug = name.lower().replace(" ", "-").replace(",", "").replace("'", "")[:30]
                slug_to_fund[slug] = fund

    ops = []
    files_parsed = 0
    files_skipped = 0

    if not os.path.exists(HOLDINGS_DIR):
        print("No holdings directory. Run: python3 tools/smart_money.py download 13f", file=sys.stderr)
        sys.exit(1)

    for fname in sorted(os.listdir(HOLDINGS_DIR)):
        if not fname.endswith(".xml"):
            continue
        fpath = os.path.join(HOLDINGS_DIR, fname)

        # Extract CIK from filename (13f_cik-XXXX_date.xml)
        cik_match = re.search(r'cik-(\d+)', fname)
        file_cik = cik_match.group(1) if cik_match else None

        # Extract filing date from filename
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fname)
        filing_date = date_match.group(1) if date_match else ""

        # Try to parse as XML
        try:
            tree = ET.parse(fpath)
            root = tree.getroot()
        except ET.ParseError:
            files_skipped += 1
            print(f"  SKIP (not valid XML): {fname}", file=sys.stderr)
            continue

        # Detect namespace
        ns = ""
        if "}" in root.tag:
            ns = root.tag.split("}")[0] + "}"

        # Check if this looks like an infotable
        entries = root.findall(f"{ns}infoTable")
        if not entries:
            files_skipped += 1
            print(f"  SKIP (no infoTable entries): {fname}", file=sys.stderr)
            continue

        # Look up fund info: try CIK first, then filename slug
        fund_info = cik_to_fund.get(file_cik, {}) if file_cik else {}
        if not fund_info:
            # Extract slug from filename: 13f_SLUG_DATE.xml
            slug_match = re.match(r'13f_(.+?)_\d{4}-\d{2}-\d{2}\.xml', fname)
            if slug_match:
                file_slug = slug_match.group(1)
                fund_info = slug_to_fund.get(file_slug, {})
        fund_name = fund_info.get("name", f"CIK-{file_cik}" if file_cik else "unknown")
        fund_type = fund_info.get("fund_type", "unknown")
        fund_cik = fund_info.get("cik", file_cik or "")
        # Slug: lowercase, no spaces/commas, max 30 chars
        fund_slug = fund_name.lower().replace(" ", "-").replace(",", "").replace("'", "")[:30]

        # Add fund node op
        ops.append({
            "op": "add_node",
            "type": "fund",
            "id": fund_slug,
            "attrs": {
                "full_name": fund_name,
                "fund_type": fund_type,
                "cik": str(fund_cik),
                "last_filing_date": filing_date,
            }
        })

        # Aggregate holdings by CUSIP (handles multi-manager like Markel)
        holdings = {}  # cusip -> {shares, value}
        for entry in entries:
            cusip = (entry.findtext(f"{ns}cusip") or "").strip()
            if cusip not in cusip_to_ticker:
                continue
            value_str = entry.findtext(f"{ns}value") or "0"
            value = int(value_str)
            shares_el = entry.find(f"{ns}shrsOrPrnAmt")
            shares = int(shares_el.findtext(f"{ns}sshPrnamt") or "0") if shares_el is not None else 0

            if cusip in holdings:
                holdings[cusip]["shares"] += shares
                holdings[cusip]["value"] += value
            else:
                holdings[cusip] = {"shares": shares, "value": value}

        # Generate edge ops for matched holdings
        for cusip, hdata in holdings.items():
            ticker = cusip_to_ticker[cusip]
            # Derive quarter from filing date (13F filed ~45d after quarter end)
            quarter = ""
            if filing_date:
                year = int(filing_date[:4])
                month = int(filing_date[5:7])
                # Feb/Mar filing → Q4 previous year, May filing → Q1, Aug → Q2, Nov → Q3
                if month <= 3:
                    quarter = f"Q4-{year - 1}"
                elif month <= 6:
                    quarter = f"Q1-{year}"
                elif month <= 9:
                    quarter = f"Q2-{year}"
                else:
                    quarter = f"Q3-{year}"

            ops.append({
                "op": "add_edge",
                "from": fund_slug,
                "to": ticker,
                "relation": "holds",
                "attrs": {
                    "shares": hdata["shares"],
                    "value_usd": hdata["value"],
                    "quarter": quarter,
                    "data_source": "13f",
                    "filing_date": filing_date,
                }
            })

        matched = len(holdings)
        total = len(entries)
        files_parsed += 1
        print(f"  PARSED: {fname} → {fund_name} | {matched}/{total} matched our CUSIPs", file=sys.stderr)

    print(f"\nSummary: {files_parsed} parsed, {files_skipped} skipped, {len(ops)} ops generated.", file=sys.stderr)

    # Output JSON to stdout (pipe to bulk-update)
    print(json.dumps(ops, indent=2))


# ---------------------------------------------------------------------------
# CMD: short-interest
# ---------------------------------------------------------------------------

def cmd_short_interest(args):
    """Fetch US short interest from yfinance for all US stocks in the graph.

    Reports shortPercentOfFloat, shortRatio, sharesShort.
    This is market-level data (not fund-attributed), complementing FCA data for UK.
    """
    import yfinance as yf

    G = load_graph()
    stocks = nodes_by_type(G, "stock")
    us_stocks = {t: d for t, d in stocks.items() if d.get("country") == "US"}

    if not us_stocks:
        print("No US stocks in graph. Run sync-portfolio first.")
        return

    print("=== US SHORT INTEREST (yfinance) ===\n")
    print(f"  {'Ticker':12s} {'Name':25s} {'Short%':>8s} {'ShortRatio':>12s} {'SharesShort':>14s} {'Portfolio':>10s}")
    print(f"  {'-'*12} {'-'*25} {'-'*8} {'-'*12} {'-'*14} {'-'*10}")

    for ticker in sorted(us_stocks.keys()):
        data = us_stocks[ticker]
        name = (data.get("name", "") or "")[:25]
        in_pf = "YES" if data.get("in_portfolio") else ""

        try:
            t = yf.Ticker(ticker)
            info = t.info
            short_pct = info.get("shortPercentOfFloat") or 0
            short_ratio = info.get("shortRatio") or 0
            shares_short = info.get("sharesShort") or 0
            # shortPercentOfFloat is already a fraction (0.05 = 5%)
            pct_display = short_pct * 100 if short_pct < 1 else short_pct
            print(f"  {ticker:12s} {name:25s} {pct_display:>7.2f}% {short_ratio:>12.2f} {shares_short:>14,d} {in_pf:>10s}")
        except Exception as e:
            print(f"  {ticker:12s} {name:25s} {'ERROR':>8s}   {str(e)[:40]}")

    print()


# ---------------------------------------------------------------------------
# CMD: ingest-fca
# ---------------------------------------------------------------------------

def cmd_ingest_fca(args):
    """Parse FCA XLSX and apply short positions directly to graph.

    Matches ISINs from identifiers.yaml to our tracked tickers.
    Each row in FCA = one fund's short on one stock (latest file = current state).
    """
    import openpyxl

    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})

    # Build ISIN → ticker map from identifiers
    isin_to_ticker = {}
    for ticker, info in ident.items():
        isin = info.get("isin")
        if isin and ticker_country(ticker) == "UK":
            isin_to_ticker[isin] = ticker

    if not isin_to_ticker:
        print("No UK ISINs in identifiers.yaml. Run 'resolve' first.")
        return

    # Find most recent FCA file
    if args.file:
        fpath = args.file
    else:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
        if not files:
            print("No FCA XLSX files found. Run: python3 tools/smart_money.py download fca")
            return
        fpath = os.path.join(SHORTS_DIR, files[0])

    print(f"Ingesting FCA: {os.path.basename(fpath)}", file=sys.stderr)

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Empty spreadsheet.")
        return

    # Find header row
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = str(row).lower()
        if "position holder" in row_str or "issuer" in row_str:
            header_idx = i
            break

    headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
    holder_col = next((i for i, h in enumerate(headers) if "position holder" in h), 0)
    isin_col = next((i for i, h in enumerate(headers) if "isin" in h), 2)
    pct_col = next((i for i, h in enumerate(headers) if "net short" in h or "position (%)" in h), 3)
    date_col = next((i for i, h in enumerate(headers) if "date" in h), 4)

    G = load_graph()
    fund_aliases = load_fund_aliases()

    # First, remove all existing FCA-sourced short edges (both "fca" and legacy "FCA")
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == "shorts" and str(d.get("data_source", "")).lower() == "fca":
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    total_rows = 0

    # Collect all entries, dedup by (fund_slug, ticker) keeping latest date
    # FCA XLSX contains historical entries — without dedup, same fund-stock pair
    # gets multiple edges, inflating SI% (e.g. DOM.L 634% instead of ~5%)
    active_shorts = {}  # (fund_slug, ticker) → {holder, pct, date, fund_slug, ticker}

    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        total_rows += 1

        isin = str(row[isin_col]).strip() if row[isin_col] else ""
        if isin not in isin_to_ticker:
            continue

        ticker = isin_to_ticker[isin]
        holder = str(row[holder_col]).strip() if row[holder_col] else ""
        pct = row[pct_col] if row[pct_col] else 0
        if isinstance(pct, str):
            try:
                pct = float(pct.replace("%", "").strip())
            except ValueError:
                pct = 0
        pos_date = ""
        if row[date_col]:
            if isinstance(row[date_col], datetime):
                pos_date = row[date_col].strftime("%Y-%m-%d")
            else:
                pos_date = str(row[date_col])[:10]

        # Create fund slug and canonicalize via aliases
        fund_slug = holder.lower().replace(" ", "-").replace(",", "").replace(".", "")[:40]
        fund_slug = canonicalize_fund_slug(fund_slug, fund_aliases)

        key = (fund_slug, ticker)
        # Keep the entry with the latest date
        if key not in active_shorts or pos_date > active_shorts[key]["date"]:
            active_shorts[key] = {
                "holder": holder,
                "fund_slug": fund_slug,
                "ticker": ticker,
                "pct": float(pct),
                "date": pos_date,
            }

    # Filter out positions below 0.50% — FCA requires disclosure at 0.50%+,
    # entries below this threshold are "exiting" notifications (position closed
    # or immaterial). Without this filter, stale sub-threshold entries from
    # years ago inflate SI% counts.
    active_shorts = {k: v for k, v in active_shorts.items() if v["pct"] >= 0.50}

    # Now add deduped entries to graph
    matched = 0
    funds_seen = set()

    for key, data in active_shorts.items():
        fund_slug = data["fund_slug"]
        ticker = data["ticker"]
        funds_seen.add(fund_slug)

        # Ensure fund node exists
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=data["holder"], fund_type="unknown")
        else:
            G.nodes[fund_slug]["full_name"] = data["holder"]

        # Ensure stock node exists
        if ticker not in G.nodes:
            G.add_node(ticker, type="stock", country="UK")

        # Add short edge (one per fund-stock pair)
        G.add_edge(fund_slug, ticker, relation="shorts", pct_shares=data["pct"],
                    date=data["date"], data_source="fca", date_added=today_str())
        matched += 1

    save_graph(G)
    print(f"FCA ingest: {matched} short positions across {len(funds_seen)} funds "
          f"(from {total_rows} total rows, {len(isin_to_ticker)} ISINs tracked)")


# ---------------------------------------------------------------------------
# CMD: ingest-amf
# ---------------------------------------------------------------------------

def cmd_ingest_amf(args):
    """Parse AMF CSV and apply short positions directly to graph.

    AMF CSV: semicolon-delimited. Columns:
      holder; LEI; issuer; ratio(%); ISIN; start_date; pub_start; pub_end
    Active positions: "Date de fin de publication position" is empty.
    Takes latest entry per (holder, ISIN) pair.
    """
    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})

    # Build ISIN → ticker map for FR stocks
    isin_to_ticker = {}
    for ticker, info in ident.items():
        isin = info.get("isin")
        if isin and isin.startswith("FR"):
            isin_to_ticker[isin] = ticker

    if not isin_to_ticker:
        print("No French ISINs in identifiers.yaml. Run 'resolve' first.")
        return

    # Find most recent AMF file
    files = sorted([f for f in os.listdir(SHORTS_DIR)
                    if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
    if not files:
        print("No AMF CSV files found. Run: python3 tools/smart_money.py download amf")
        return

    fpath = os.path.join(SHORTS_DIR, files[0])
    print(f"Ingesting AMF: {os.path.basename(fpath)}", file=sys.stderr)

    # Parse CSV: semicolon-delimited, all fields quoted
    # For each (holder, ISIN): keep only latest active entry
    active_shorts = {}  # (holder, isin) → {ratio, date, ...}
    total_rows = 0

    with open(fpath, "r", encoding="utf-8", errors="replace") as f:
        header_line = f.readline()  # skip header
        for line in f:
            total_rows += 1
            parts = line.strip().split(";")
            if len(parts) < 8:
                continue

            holder = parts[0].strip().strip('"')
            issuer = parts[2].strip().strip('"')
            ratio_str = parts[3].strip().strip('"')
            isin = parts[4].strip().strip('"')
            start_date = parts[5].strip().strip('"')
            pub_end = parts[7].strip().strip('"')

            # Skip non-tracked ISINs
            if isin not in isin_to_ticker:
                continue

            # Skip closed positions (non-empty end date)
            if pub_end:
                continue

            try:
                ratio = float(ratio_str)
            except ValueError:
                continue

            key = (holder, isin)
            # Keep the entry with the latest start_date
            if key not in active_shorts or start_date > active_shorts[key]["date"]:
                active_shorts[key] = {
                    "holder": holder,
                    "isin": isin,
                    "ticker": isin_to_ticker[isin],
                    "ratio": ratio,
                    "date": start_date,
                    "issuer": issuer,
                }

    G = load_graph()
    fund_aliases = load_fund_aliases()

    # Remove existing AMF-sourced short edges
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == "shorts" and d.get("data_source") == "amf":
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    funds_seen = set()
    for key, data in active_shorts.items():
        holder = data["holder"]
        ticker = data["ticker"]
        fund_slug = holder.lower().replace(" ", "-").replace(",", "").replace(".", "")[:40]
        fund_slug = canonicalize_fund_slug(fund_slug, fund_aliases)
        funds_seen.add(fund_slug)

        # Ensure fund node
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=holder, fund_type="unknown")
        else:
            G.nodes[fund_slug]["full_name"] = holder

        # Ensure stock node
        if ticker not in G.nodes:
            G.add_node(ticker, type="stock", country="FR")

        G.add_edge(fund_slug, ticker, relation="shorts", pct_shares=data["ratio"],
                    date=data["date"], data_source="amf", date_added=today_str())

    save_graph(G)

    # Summary by ticker
    by_ticker = defaultdict(list)
    for key, data in active_shorts.items():
        by_ticker[data["ticker"]].append(data)

    print(f"AMF ingest: {len(active_shorts)} active short positions across {len(funds_seen)} funds")
    for ticker in sorted(by_ticker.keys()):
        entries = by_ticker[ticker]
        total_pct = sum(e["ratio"] for e in entries)
        print(f"  {ticker}: {total_pct:.2f}% total ({len(entries)} funds)")
        for e in sorted(entries, key=lambda x: -x["ratio"]):
            print(f"    {e['holder'][:45]:45s} {e['ratio']:.2f}%  ({e['date']})")


# ---------------------------------------------------------------------------
# CMD: ingest-13f
# ---------------------------------------------------------------------------

def cmd_ingest_13f(args):
    """Parse all 13F XMLs and apply holdings directly to graph.

    Uses identifiers.yaml CUSIPs for matching (plus legacy cusip_map.yaml).
    Refactored from parse-13f to apply directly instead of outputting JSON.
    """
    import xml.etree.ElementTree as ET
    import re

    # Build CUSIP → ticker map from identifiers + legacy cusip_map
    cusip_to_ticker = {}

    # Legacy cusip_map.yaml
    if os.path.exists(CUSIP_MAP_PATH):
        with open(CUSIP_MAP_PATH) as f:
            cusip_data = yaml.safe_load(f) or {}
        cusip_to_ticker.update(cusip_data.get("cusip_to_ticker", {}))

    # identifiers.yaml CUSIPs
    identifiers_data = load_identifiers()
    for ticker, info in identifiers_data.get("identifiers", {}).items():
        cusip = info.get("cusip")
        if cusip:
            cusip_to_ticker[cusip] = ticker

    if not cusip_to_ticker:
        print("No CUSIPs available. Run 'resolve' first or check cusip_map.yaml.")
        return

    # Load tracked funds for CIK → metadata
    cik_to_fund = {}
    slug_to_fund = {}
    if os.path.exists(TRACKED_FUNDS_PATH):
        with open(TRACKED_FUNDS_PATH) as f:
            tf = yaml.safe_load(f) or {}
        for fund in tf.get("funds", []):
            cik = str(fund.get("cik", ""))
            if cik:
                cik_to_fund[cik] = fund
            name = fund.get("name", "")
            if name:
                slug = name.lower().replace(" ", "-").replace(",", "").replace("'", "")[:30]
                slug_to_fund[slug] = fund

    G = load_graph()

    # Remove existing 13f-sourced hold edges
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == "holds" and d.get("data_source") == "13f":
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    if not os.path.exists(HOLDINGS_DIR):
        print("No holdings directory. Run: python3 tools/smart_money.py download 13f")
        return

    # Also load identifiers for enhanced CUSIP resolution
    identifiers_data = load_identifiers()

    files_parsed = 0
    total_edges = 0
    # Track unresolved CUSIPs: cusip -> {issuer, funds, total_value}
    unresolved_cusips = defaultdict(lambda: {"issuer": "", "funds": set(), "total_value": 0})
    late_resolved = 0

    for fname in sorted(os.listdir(HOLDINGS_DIR)):
        if not fname.endswith(".xml"):
            continue
        fpath = os.path.join(HOLDINGS_DIR, fname)

        cik_match = re.search(r'cik-(\d+)', fname)
        file_cik = cik_match.group(1) if cik_match else None

        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', fname)
        filing_date = date_match.group(1) if date_match else ""

        try:
            tree = ET.parse(fpath)
            root = tree.getroot()
        except ET.ParseError:
            continue

        ns = ""
        if "}" in root.tag:
            ns = root.tag.split("}")[0] + "}"

        entries = root.findall(f"{ns}infoTable")
        if not entries:
            continue

        # Look up fund info
        fund_info = cik_to_fund.get(file_cik, {}) if file_cik else {}
        if not fund_info:
            slug_match = re.match(r'13f_(.+?)_\d{4}-\d{2}-\d{2}\.xml', fname)
            if slug_match:
                file_slug = slug_match.group(1)
                fund_info = slug_to_fund.get(file_slug, {})

        fund_name = fund_info.get("name", f"CIK-{file_cik}" if file_cik else "unknown")
        fund_type = fund_info.get("fund_type", "unknown")
        fund_cik = fund_info.get("cik", file_cik or "")
        fund_slug = fund_name.lower().replace(" ", "-").replace(",", "").replace("'", "")[:30]

        # Ensure fund node
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=fund_name,
                       fund_type=fund_type, cik=str(fund_cik))
        else:
            G.nodes[fund_slug].update({"full_name": fund_name, "fund_type": fund_type,
                                       "cik": str(fund_cik), "last_filing_date": filing_date})

        # Aggregate holdings by CUSIP
        holdings = {}
        for entry in entries:
            cusip = (entry.findtext(f"{ns}cusip") or "").strip()
            if not cusip:
                continue
            if cusip not in cusip_to_ticker:
                # Try enhanced resolution
                resolved = _cusip_to_ticker_lookup(cusip, identifiers_data)
                if resolved:
                    cusip_to_ticker[cusip] = resolved
                    late_resolved += 1
                else:
                    # Track as unresolved
                    issuer = (entry.findtext(f"{ns}nameOfIssuer") or "").strip()
                    value = int(entry.findtext(f"{ns}value") or "0")
                    rec = unresolved_cusips[cusip]
                    rec["issuer"] = issuer or rec["issuer"]
                    rec["funds"].add(fund_name)
                    rec["total_value"] += value
                    continue
            value = int(entry.findtext(f"{ns}value") or "0")
            shares_el = entry.find(f"{ns}shrsOrPrnAmt")
            shares = int(shares_el.findtext(f"{ns}sshPrnamt") or "0") if shares_el is not None else 0
            if cusip in holdings:
                holdings[cusip]["shares"] += shares
                holdings[cusip]["value"] += value
            else:
                holdings[cusip] = {"shares": shares, "value": value}

        # Add edges
        for cusip, hdata in holdings.items():
            ticker = cusip_to_ticker[cusip]
            quarter = ""
            if filing_date:
                year = int(filing_date[:4])
                month = int(filing_date[5:7])
                if month <= 3:
                    quarter = f"Q4-{year - 1}"
                elif month <= 6:
                    quarter = f"Q1-{year}"
                elif month <= 9:
                    quarter = f"Q2-{year}"
                else:
                    quarter = f"Q3-{year}"

            if ticker not in G.nodes:
                G.add_node(ticker, type="stock", country=ticker_country(ticker))

            G.add_edge(fund_slug, ticker, relation="holds",
                       shares=hdata["shares"], value_usd=hdata["value"],
                       quarter=quarter, data_source="13f", filing_date=filing_date,
                       date_added=today_str())
            total_edges += 1

        files_parsed += 1

    save_graph(G)
    print(f"13F ingest: {files_parsed} filings parsed, {total_edges} hold edges created "
          f"({len(cusip_to_ticker)} CUSIPs tracked)")
    if late_resolved:
        print(f"  Enhanced CUSIP resolution: {late_resolved} additional CUSIPs resolved via ISIN matching")

    # Report unresolved CUSIPs held by 2+ tracked funds (high-value unknowns)
    multi_fund_unresolved = {c: r for c, r in unresolved_cusips.items() if len(r["funds"]) >= 2}
    if multi_fund_unresolved:
        ranked = sorted(multi_fund_unresolved.items(), key=lambda x: -len(x[1]["funds"]))
        print(f"\n  Unresolved 13F CUSIPs held by 2+ tracked funds ({len(ranked)} total):")
        print(f"  {'CUSIP':12s} {'Issuer':35s} {'Funds':>5s} {'Value':>12s}")
        print(f"  {'-'*12} {'-'*35} {'-'*5} {'-'*12}")
        for cusip, rec in ranked[:15]:
            val_str = f"${rec['total_value']:,.0f}" if rec["total_value"] else ""
            print(f"  {cusip:12s} {rec['issuer'][:35]:35s} {len(rec['funds']):>5d} {val_str:>12s}")
        if len(ranked) > 15:
            print(f"  ... and {len(ranked) - 15} more")


def _get_top_universe_tickers(max_count=25, country=None):
    """Load top tickers from quality universe data (by QS), optionally filtered by country."""
    universe_path = os.path.join(PROJECT_ROOT, "state", "quality_universe.yaml")
    alt_path = os.path.join(DATA_DIR, "data", "quality_universe.yaml")
    data = None
    for path in [universe_path, alt_path]:
        if os.path.exists(path):
            try:
                with open(path) as f:
                    data = yaml.safe_load(f) or {}
                break
            except Exception:
                continue
    if not data:
        return []

    stocks = data.get("stocks", [])
    if country:
        stocks = [s for s in stocks if ticker_country(s.get("ticker", "")) == country]

    # Sort by QS descending
    stocks.sort(key=lambda s: s.get("qs", 0), reverse=True)
    return [s["ticker"] for s in stocks[:max_count] if s.get("ticker")]


# ---------------------------------------------------------------------------
# CMD: ingest-insider
# ---------------------------------------------------------------------------

def cmd_ingest_insider(args):
    """Ingest insider transactions from yfinance for stocks in graph.

    Creates person nodes and insider_buy/insider_sell edges.
    Also stores shortPercentOfFloat on stock nodes.

    --universe: also ingest top 25 US universe stocks (by QS or E[CAGR])
    --all-enrolled: ingest for ALL enrolled stocks (respects yfinance budget)
    """
    import yfinance as yf
    import time

    G = load_graph()
    stocks = nodes_by_type(G, "stock")

    if args.tickers:
        target_tickers = args.tickers
    elif getattr(args, "all_enrolled", False):
        # All enrolled stocks — limit to 50 for yfinance budget
        target_tickers = sorted(n for n, d in stocks.items())[:50]
    elif getattr(args, "universe", False):
        # Portfolio + top 25 US universe stocks
        portfolio_tickers = sorted(n for n, d in stocks.items() if d.get("in_portfolio"))
        # Load universe to find top US stocks
        universe_tickers = _get_top_universe_tickers(max_count=25, country="US")
        # Combine, deduplicate
        combined = list(dict.fromkeys(portfolio_tickers + universe_tickers))
        target_tickers = combined[:50]  # yfinance budget
    else:
        # Default: portfolio stocks
        target_tickers = sorted(n for n, d in stocks.items() if d.get("in_portfolio"))

    if not target_tickers:
        print("No tickers to process. Run sync-portfolio first.")
        return

    print(f"=== INGESTING INSIDER DATA ({len(target_tickers)} tickers) ===\n")

    # Remove existing insider edges for target tickers
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") in ("insider_buy", "insider_sell") and v in target_tickers:
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    total_buys = 0
    total_sells = 0

    for ticker in target_tickers:
        try:
            t = yf.Ticker(ticker)

            # Short interest (store on stock node)
            try:
                info = t.info
                short_pct = info.get("shortPercentOfFloat")
                if short_pct is not None:
                    if short_pct < 1:
                        short_pct = short_pct * 100  # Convert to percentage
                    G.nodes[ticker]["short_pct_float"] = round(short_pct, 2)
            except Exception:
                pass

            # Insider transactions
            try:
                insider_df = t.insider_transactions
                if insider_df is not None and not insider_df.empty:
                    # Detect columns (yfinance changes column names across versions)
                    cols = [c.lower() for c in insider_df.columns]
                    insider_df.columns = cols

                    for _, row in insider_df.iterrows():
                        # Try to extract fields
                        name = ""
                        for col in ["insider", "insider trading", "name"]:
                            if col in cols:
                                name = str(row[col]) if row[col] else ""
                                break

                        # Classify transaction — check "text" first (has "Sale at price..."
                        # or "Purchase at price..."), then "transaction", then "type"
                        txn_text = ""
                        for col in ["text", "transaction", "type"]:
                            if col in cols and row[col]:
                                val = str(row[col]).lower().strip()
                                # Skip date-like values (yfinance "Transaction" column is often a date)
                                if len(val) == 10 and val[4] == '-':
                                    continue
                                txn_text += " " + val

                        shares_val = 0
                        for col in ["shares", "shares traded", "#shares"]:
                            if col in cols:
                                try:
                                    shares_val = abs(int(float(row[col]))) if row[col] else 0
                                except (ValueError, TypeError):
                                    pass
                                break

                        value_val = 0
                        for col in ["value", "$ value", "value ($)"]:
                            if col in cols:
                                try:
                                    value_val = abs(float(row[col])) if row[col] else 0
                                except (ValueError, TypeError):
                                    pass
                                break

                        txn_date = ""
                        for col in ["start date", "date", "date reported"]:
                            if col in cols:
                                try:
                                    d = row[col]
                                    if hasattr(d, 'strftime'):
                                        txn_date = d.strftime("%Y-%m-%d")
                                    else:
                                        txn_date = str(d)[:10]
                                except Exception:
                                    pass
                                break

                        if not name or not txn_text:
                            continue

                        # Classify as buy or sell from text description
                        is_buy = any(w in txn_text for w in ["purchase", "buy", "acquisition"])
                        is_sell = any(w in txn_text for w in ["sale", "sell", "dispos"])

                        if not is_buy and not is_sell:
                            continue

                        # Get role/position
                        role = ""
                        for col in ["position", "role", "title"]:
                            if col in cols and row[col]:
                                role = str(row[col]).strip()
                                break

                        # Create person node
                        person_slug = name.lower().replace(" ", "-").replace(",", "").replace(".", "")[:30]
                        if person_slug not in G.nodes:
                            G.add_node(person_slug, type="person", full_name=name, role=role)
                        elif role:
                            G.nodes[person_slug]["role"] = role

                        relation = "insider_buy" if is_buy else "insider_sell"
                        G.add_edge(person_slug, ticker, relation=relation,
                                   shares=shares_val, value=value_val, date=txn_date,
                                   role=role, data_source="yfinance", date_added=today_str())

                        if is_buy:
                            total_buys += 1
                        else:
                            total_sells += 1

                    n_txns = len(insider_df)
                    print(f"  {ticker:12s}  {n_txns} transactions found")
                else:
                    print(f"  {ticker:12s}  no insider data")
            except Exception as e:
                print(f"  {ticker:12s}  insider error: {e}")

        except Exception as e:
            print(f"  {ticker:12s}  ERROR: {e}")

        time.sleep(0.5)  # Rate limit

    save_graph(G)
    print(f"\nInsider ingest: {total_buys} buys, {total_sells} sells across {len(target_tickers)} tickers")


# ---------------------------------------------------------------------------
# CMD: coverage
# ---------------------------------------------------------------------------

def cmd_coverage(args):
    """Show per-stock data coverage gap report."""
    G = load_graph()
    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})

    stocks = nodes_by_type(G, "stock")
    if args.portfolio_only:
        stocks = {n: d for n, d in stocks.items() if d.get("in_portfolio")}

    if not stocks:
        print("No stocks in graph. Run sync-portfolio first.")
        return

    print("=== COVERAGE REPORT ===\n")
    print(f"  {'Ticker':12s} {'ISIN':>5s} {'CUSIP':>6s} {'CIK':>4s} {'13F':>4s} "
          f"{'FCA':>4s} {'AMF':>4s} {'Insider':>8s} {'Short%':>7s} {'Coverage':>9s}")
    print(f"  {'-'*12} {'-'*5} {'-'*6} {'-'*4} {'-'*4} {'-'*4} {'-'*4} {'-'*8} {'-'*7} {'-'*9}")

    fully_covered = 0
    partial = 0
    no_data = 0

    for ticker in sorted(stocks.keys()):
        info = ident.get(ticker, {})
        country = ticker_country(ticker)

        has_isin = "Y" if info.get("isin") else "-"
        has_cusip = "Y" if info.get("cusip") else "-"
        has_cik = "Y" if info.get("cik") else "-"

        # Check edges
        has_13f = "-"
        has_fca = "-"
        has_amf = "-"
        has_insider = "-"
        has_short_pct = "-"

        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            src = d.get("data_source", "")
            if rel == "holds" and src == "13f":
                has_13f = "Y"
            elif rel == "shorts" and src == "fca":
                has_fca = "Y"
            elif rel == "shorts" and src == "amf":
                has_amf = "Y"
            elif rel in ("insider_buy", "insider_sell"):
                has_insider = "Y"

        short_pct = stocks[ticker].get("short_pct_float")
        if short_pct is not None:
            has_short_pct = f"{short_pct:.1f}%"

        # Calculate coverage score
        checks = 0
        possible = 0

        # ISIN always relevant
        if info.get("isin"):
            checks += 1
        possible += 1

        # 13F relevant for US stocks
        if country == "US":
            possible += 1
            if has_13f == "Y":
                checks += 1

        # FCA relevant for UK
        if country == "UK":
            possible += 1
            if has_fca == "Y":
                checks += 1

        # AMF relevant for FR
        if country == "FR":
            possible += 1
            if has_amf == "Y":
                checks += 1

        # Insider always relevant
        possible += 1
        if has_insider == "Y":
            checks += 1

        coverage_pct = (checks / possible * 100) if possible > 0 else 0

        if coverage_pct >= 80:
            fully_covered += 1
        elif coverage_pct > 0:
            partial += 1
        else:
            no_data += 1

        # N/A markers for irrelevant sources
        if country != "US":
            has_13f = "n/a"
            has_cusip = "n/a"
            has_cik = "n/a"
        if country != "UK":
            has_fca = "n/a"
        if country != "FR":
            has_amf = "n/a"

        print(f"  {ticker:12s} {has_isin:>5s} {has_cusip:>6s} {has_cik:>4s} {has_13f:>4s} "
              f"{has_fca:>4s} {has_amf:>4s} {has_insider:>8s} {has_short_pct:>7s} {coverage_pct:>8.0f}%")

    print(f"\n  Summary: {len(stocks)} stocks, {fully_covered} fully covered, "
          f"{partial} partial, {no_data} no data")


# ---------------------------------------------------------------------------
# Auto-expand: discover stocks from regulatory data and auto-enroll
# ---------------------------------------------------------------------------

def _auto_expand_from_regulatory(G, identifiers_data, min_funds_13f=2, min_funds_short=3, max_new=20):
    """Find stocks held/shorted by tracked funds NOT in graph and return tickers to enroll.

    Reuses discovery logic from cmd_discover internally.
    Returns list of (ticker, source, evidence) tuples, max max_new items.
    """
    import xml.etree.ElementTree as ET
    import re

    our_stocks = set(n for n, d in G.nodes(data=True) if d.get("type") == "stock")
    ident = identifiers_data.get("identifiers", {})

    # Known CUSIPs and ISINs
    known_cusips = set()
    known_isins = set()
    for ticker, info in ident.items():
        cusip = info.get("cusip")
        if cusip:
            known_cusips.add(cusip)
        isin = info.get("isin")
        if isin:
            known_isins.add(isin)
    if os.path.exists(CUSIP_MAP_PATH):
        with open(CUSIP_MAP_PATH) as f:
            cm = yaml.safe_load(f) or {}
        known_cusips.update(cm.get("cusip_to_ticker", {}).keys())

    # Load tracked funds for CIK mapping
    cik_to_fund = {}
    if os.path.exists(TRACKED_FUNDS_PATH):
        with open(TRACKED_FUNDS_PATH) as f:
            tf = yaml.safe_load(f) or {}
        for fund in tf.get("funds", []):
            cik = str(fund.get("cik", ""))
            if cik:
                cik_to_fund[cik] = fund

    candidates = []

    # --- 13F: stocks held by min_funds_13f+ tracked funds ---
    if os.path.exists(HOLDINGS_DIR):
        cusip_hits = defaultdict(lambda: {"funds": [], "issuer": "", "cusip": "", "value": 0})
        for fname in os.listdir(HOLDINGS_DIR):
            if not fname.endswith(".xml"):
                continue
            fpath = os.path.join(HOLDINGS_DIR, fname)
            cik_match = re.search(r'cik-(\d+)', fname)
            file_cik = cik_match.group(1) if cik_match else None
            fund_info = cik_to_fund.get(file_cik, {}) if file_cik else {}
            fund_name = fund_info.get("name", f"CIK-{file_cik}" if file_cik else "unknown")

            try:
                tree = ET.parse(fpath)
                root = tree.getroot()
            except ET.ParseError:
                continue

            ns = ""
            if "}" in root.tag:
                ns = root.tag.split("}")[0] + "}"

            for entry in root.findall(f"{ns}infoTable"):
                cusip = (entry.findtext(f"{ns}cusip") or "").strip()
                if not cusip or cusip in known_cusips:
                    continue
                issuer = (entry.findtext(f"{ns}nameOfIssuer") or "").strip()
                value = int(entry.findtext(f"{ns}value") or "0")

                rec = cusip_hits[cusip]
                rec["cusip"] = cusip
                rec["issuer"] = issuer or rec["issuer"]
                if fund_name not in rec["funds"]:
                    rec["funds"].append(fund_name)
                rec["value"] += value

        # Try to resolve CUSIP → ticker using EDGAR company_tickers.json
        for cusip, rec in cusip_hits.items():
            if len(rec["funds"]) < min_funds_13f:
                continue
            # Try to derive ticker from issuer name (simple heuristic)
            ticker_guess = _guess_ticker_from_issuer(rec["issuer"])
            if ticker_guess and ticker_guess not in our_stocks:
                candidates.append((ticker_guess, "13f", f"{len(rec['funds'])} funds, ${rec['value']:,.0f}"))

    # --- FCA: UK stocks shorted by min_funds_short+ funds ---
    fca_files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
    if fca_files:
        try:
            import openpyxl
            fpath = os.path.join(SHORTS_DIR, fca_files[0])
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            header_idx = 0
            for i, row in enumerate(rows):
                if "position holder" in str(row).lower() or "issuer" in str(row).lower():
                    header_idx = i
                    break

            headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
            holder_col = next((i for i, h in enumerate(headers) if "position holder" in h), 0)
            isin_col = next((i for i, h in enumerate(headers) if "isin" in h), 2)
            issuer_col = next((i for i, h in enumerate(headers) if "issuer" in h or "name of share" in h), 1)

            isin_hits = defaultdict(lambda: {"funds": [], "issuer": "", "isin": ""})
            for row in rows[header_idx + 1:]:
                if all(v is None for v in row):
                    continue
                isin = str(row[isin_col]).strip() if row[isin_col] else ""
                if not isin.startswith("GB") or isin in known_isins:
                    continue
                holder = str(row[holder_col]).strip() if row[holder_col] else ""
                issuer = str(row[issuer_col]).strip() if row[issuer_col] else ""
                rec = isin_hits[isin]
                rec["isin"] = isin
                rec["issuer"] = issuer or rec["issuer"]
                if holder not in rec["funds"]:
                    rec["funds"].append(holder)

            unresolved_fca = []
            for isin, rec in isin_hits.items():
                if len(rec["funds"]) >= min_funds_short:
                    ticker_guess = _guess_uk_ticker(rec["issuer"], isin=isin)
                    if ticker_guess and ticker_guess not in our_stocks:
                        candidates.append((ticker_guess, "fca", f"{len(rec['funds'])} short funds, ISIN {isin}"))
                    elif not ticker_guess:
                        unresolved_fca.append((isin, rec["issuer"], len(rec["funds"])))
            if unresolved_fca:
                print(f"  FCA auto-expand: {len(unresolved_fca)} unresolved UK stocks (ISIN not in identifiers):")
                for isin, issuer, nfunds in unresolved_fca[:10]:
                    print(f"    {isin}  {issuer:30s}  ({nfunds} funds)")
        except ImportError:
            pass

    # --- AMF: FR stocks shorted by min_funds_short+ funds ---
    amf_files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
    if amf_files:
        fpath = os.path.join(SHORTS_DIR, amf_files[0])
        isin_hits = defaultdict(lambda: {"funds": [], "issuer": "", "isin": ""})
        try:
            with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                f.readline()  # skip header
                for line in f:
                    parts = line.strip().split(";")
                    if len(parts) < 8:
                        continue
                    holder = parts[0].strip().strip('"')
                    issuer = parts[2].strip().strip('"')
                    isin = parts[4].strip().strip('"')
                    pub_end = parts[7].strip().strip('"')
                    if not isin.startswith("FR") or isin in known_isins or pub_end:
                        continue
                    rec = isin_hits[isin]
                    rec["isin"] = isin
                    rec["issuer"] = issuer or rec["issuer"]
                    if holder not in rec["funds"]:
                        rec["funds"].append(holder)
        except Exception:
            pass

        unresolved_amf = []
        for isin, rec in isin_hits.items():
            if len(rec["funds"]) >= min_funds_short:
                ticker_guess = _guess_fr_ticker(rec["issuer"], isin=isin)
                if ticker_guess and ticker_guess not in our_stocks:
                    candidates.append((ticker_guess, "amf", f"{len(rec['funds'])} short funds, ISIN {isin}"))
                elif not ticker_guess:
                    unresolved_amf.append((isin, rec["issuer"], len(rec["funds"])))
        if unresolved_amf:
            print(f"  AMF auto-expand: {len(unresolved_amf)} unresolved FR stocks (ISIN not in identifiers):")
            for isin, issuer, nfunds in unresolved_amf[:10]:
                print(f"    {isin}  {issuer:30s}  ({nfunds} funds)")

    # Deduplicate and limit
    seen = set()
    unique = []
    for ticker, source, evidence in candidates:
        if ticker not in seen:
            seen.add(ticker)
            unique.append((ticker, source, evidence))
    return unique[:max_new]


def _guess_ticker_from_issuer(issuer_name):
    """Best-effort: guess US ticker from 13F issuer name using EDGAR tickers cache."""
    if not issuer_name:
        return None
    # Try EDGAR company_tickers.json
    tickers_path = os.path.join(DATA_DIR, "data", "company_tickers.json")
    if not os.path.exists(tickers_path):
        return None
    try:
        with open(tickers_path) as f:
            tickers_data = json.load(f)
    except Exception:
        return None

    issuer_upper = issuer_name.upper().replace(",", "").replace(".", "").strip()
    # Exact or prefix match on title
    for entry in tickers_data.values():
        title = entry.get("title", "").upper().replace(",", "").replace(".", "").strip()
        if title == issuer_upper or title.startswith(issuer_upper.split()[0] + " "):
            return entry.get("ticker", "").upper()
    return None


def _cusip_to_ticker_lookup(cusip, identifiers_data=None):
    """Try to resolve CUSIP → ticker through multiple sources:
    1. identifiers.yaml (ISIN[2:11] or cusip field match)
    2. cusip_map.yaml (legacy)
    3. EDGAR company_tickers.json (issuer name → ticker via _guess_ticker_from_issuer)
    Returns ticker string or None.
    """
    if not cusip:
        return None

    # 1. identifiers.yaml — check cusip field and ISIN substring
    if identifiers_data is None:
        identifiers_data = load_identifiers()
    for ticker, info in identifiers_data.get("identifiers", {}).items():
        if info.get("cusip") == cusip:
            return ticker
        isin = info.get("isin", "")
        # US CUSIPs are ISIN chars 2-10 (ISIN = country(2) + CUSIP(9) + check(1))
        if isin and len(isin) >= 11 and isin[2:11] == cusip:
            return ticker

    # 2. cusip_map.yaml (legacy)
    if os.path.exists(CUSIP_MAP_PATH):
        with open(CUSIP_MAP_PATH) as f:
            cm = yaml.safe_load(f) or {}
        ticker = cm.get("cusip_to_ticker", {}).get(cusip)
        if ticker:
            return ticker

    return None


def _isin_to_ticker_reverse(isin):
    """Reverse lookup: find ticker from ISIN via identifiers.yaml cache."""
    if not isin:
        return None
    data = load_identifiers()
    for ticker, info in data.get("identifiers", {}).items():
        if info.get("isin") == isin:
            return ticker
    return None


def _guess_uk_ticker(issuer_name, isin=None):
    """Best-effort: guess .L ticker from FCA issuer name or ISIN reverse lookup."""
    if isin:
        ticker = _isin_to_ticker_reverse(isin)
        if ticker:
            return ticker
    if not issuer_name:
        return None
    return None


def _guess_fr_ticker(issuer_name, isin=None):
    """Best-effort: guess .PA ticker from AMF issuer name or ISIN reverse lookup."""
    if isin:
        ticker = _isin_to_ticker_reverse(isin)
        if ticker:
            return ticker
    if not issuer_name:
        return None
    return None


# ---------------------------------------------------------------------------
# CMD: refresh
# ---------------------------------------------------------------------------

def cmd_refresh(args):
    """One-command full refresh cycle.

    sync-portfolio → resolve → check staleness → download (if stale) →
    ingest-fca → ingest-amf → ingest-consob → ingest-afm-nl → ingest-13f →
    ingest-insider → [expand] → snapshot → coverage

    --expand: auto-discover stocks from regulatory data and enroll new ones.
    --full: implies --expand, force download all sources.
    """
    import time

    expand = getattr(args, "expand", False) or getattr(args, "full", False)
    total_steps = 12 if expand else 10

    print("=" * 60)
    print(f"  SMART MONEY GRAPH — {'EXPANDING ' if expand else ''}REFRESH")
    print("=" * 60)

    step = 0

    # Step 1: sync-portfolio
    step += 1
    print(f"\n[{step}/{total_steps}] Syncing portfolio...")
    cmd_sync_portfolio(argparse.Namespace())

    # Step 2: resolve
    step += 1
    print(f"\n[{step}/{total_steps}] Resolving identifiers...")
    cmd_resolve(argparse.Namespace(tickers=None, force=False, retry_failed=False))

    # Step 2b: harvest ISINs from regulatory files
    step += 1
    print(f"\n[{step}/{total_steps}] Harvesting ISINs from regulatory files...")
    cmd_harvest_isins(argparse.Namespace())

    # Step 3: Check staleness & download if needed
    step += 1
    if not args.skip_download:
        sources = init_sources()
        print(f"\n[{step}/{total_steps}] Checking staleness & downloading if needed...")

        for source_key, download_arg in [
            ("fca_uk_shorts", "fca"),
            ("amf_france_shorts", "amf"),
        ]:
            info = sources.get(source_key, {})
            last = info.get("last_download")
            cadence = info.get("cadence_days", 3)
            days = days_since(last)
            label = staleness_label(days, cadence)

            if args.full or label != "FRESH":
                print(f"  {source_key}: {label} → downloading...")
                cmd_download(argparse.Namespace(source=download_arg))
                time.sleep(1)
            else:
                print(f"  {source_key}: FRESH (skipping)")

        # 13F: only on full refresh (90-day cadence)
        info_13f = sources.get("sec_13f", {})
        days_13f = days_since(info_13f.get("last_download"))
        if args.full or staleness_label(days_13f, 90) != "FRESH":
            print(f"  sec_13f: downloading...")
            cmd_download(argparse.Namespace(source="13f"))
        else:
            print(f"  sec_13f: FRESH (skipping)")
    else:
        print(f"\n[{step}/{total_steps}] Skipping downloads (--skip-download)")

    # Step 5: ingest-fca
    step += 1
    print(f"\n[{step}/{total_steps}] Ingesting FCA UK shorts...")
    fca_files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
    if fca_files:
        cmd_ingest_fca(argparse.Namespace(file=None))
    else:
        print("  No FCA data available.")

    # Step 6: ingest-amf
    step += 1
    print(f"\n[{step}/{total_steps}] Ingesting AMF France shorts...")
    amf_files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
    if amf_files:
        cmd_ingest_amf(argparse.Namespace())
    else:
        print("  No AMF data available.")

    # Step 6b: ingest-consob (if data exists)
    consob_files = sorted([f for f in os.listdir(SHORTS_DIR)
                           if f.startswith("consob_italy_") and f.endswith(".xlsx")], reverse=True)
    if consob_files:
        print(f"  Ingesting CONSOB Italy shorts...")
        cmd_ingest_consob(argparse.Namespace(file=None))

    # Step 6c: ingest-afm-nl (if data exists)
    afm_nl_files = sorted([f for f in os.listdir(SHORTS_DIR)
                           if f.startswith("afm_nl_") and f.endswith(".csv")], reverse=True)
    if afm_nl_files:
        print(f"  Ingesting AFM NL shorts...")
        cmd_ingest_afm_nl(argparse.Namespace(file=None))

    # Step 7: ingest-13f
    step += 1
    print(f"\n[{step}/{total_steps}] Ingesting 13F holdings...")
    if os.path.exists(HOLDINGS_DIR) and any(f.endswith(".xml") for f in os.listdir(HOLDINGS_DIR)):
        cmd_ingest_13f(argparse.Namespace())
    else:
        print("  No 13F data available.")

    # Step 8: ingest-insider
    step += 1
    print(f"\n[{step}/{total_steps}] Ingesting insider data (portfolio + universe)...")
    cmd_ingest_insider(argparse.Namespace(tickers=None, universe=False, all_enrolled=False))

    # Step 9: Auto-expand (if --expand or --full)
    if expand:
        step += 1
        print(f"\n[{step}/{total_steps}] Auto-expanding graph from regulatory data...")
        G = load_graph()
        identifiers_data = load_identifiers()
        new_tickers = _auto_expand_from_regulatory(G, identifiers_data)
        if new_tickers:
            print(f"  Discovered {len(new_tickers)} new candidates:")
            enrolled = 0
            for ticker, source, evidence in new_tickers:
                print(f"    {ticker:12s}  [{source}]  {evidence}")
                # Enroll in graph
                if ticker not in G.nodes:
                    G.add_node(ticker, type="stock", country=ticker_country(ticker),
                               enrolled_date=today_str(), auto_expanded=True)
                    enrolled += 1
            save_graph(G)
            if enrolled:
                print(f"  Enrolled {enrolled} new stocks. Resolving identifiers...")
                cmd_resolve(argparse.Namespace(
                    tickers=[t for t, _, _ in new_tickers],
                    force=False, retry_failed=False))
                # Re-run ingest for newly enrolled stocks to pick up edges
                print(f"  Re-ingesting for newly enrolled stocks...")
                if fca_files:
                    cmd_ingest_fca(argparse.Namespace(file=None))
                if amf_files:
                    cmd_ingest_amf(argparse.Namespace())
        else:
            print("  No new candidates found (graph already covers tracked fund holdings).")

        # Also retry failed resolutions
        step += 1
        print(f"\n[{step}/{total_steps}] Retrying failed identifier resolutions...")
        identifiers_data = load_identifiers()
        failed = identifiers_data.get("failed_resolutions", [])
        if failed:
            print(f"  {len(failed)} failed tickers to retry...")
            cmd_resolve(argparse.Namespace(tickers=None, force=False, retry_failed=True))
        else:
            print("  No failed resolutions to retry.")

    # Step 10: snapshot + coverage
    step += 1
    print(f"\n[{step}/{total_steps}] Snapshot + coverage...")
    cmd_snapshot(argparse.Namespace())
    print()
    cmd_coverage(argparse.Namespace(portfolio_only=True))

    # Final stats
    print()
    cmd_stats(argparse.Namespace())


# ---------------------------------------------------------------------------
# CMD: enroll
# ---------------------------------------------------------------------------

def cmd_enroll(args):
    """Add stocks to graph from universe, pipeline, or explicit list.

    --from-universe: load all tickers from quality_universe.py data
    --from-pipeline: load tickers from pipeline_tracker.yaml
    [TICKER ...]: explicit tickers to add
    """
    G = load_graph()
    identifiers_data = load_identifiers()
    added = 0

    tickers_to_add = []

    if args.from_universe:
        universe_path = os.path.join(DATA_DIR, "data", "quality_universe.yaml")
        # Try the tool's output location
        alt_path = os.path.join(PROJECT_ROOT, "state", "quality_universe.yaml")
        for path in [universe_path, alt_path]:
            if os.path.exists(path):
                with open(path) as f:
                    udata = yaml.safe_load(f) or {}
                # Support both formats: {universe: [...]} and {quality_universe: {companies: [...]}}
                entries = udata.get("universe", [])
                if not entries:
                    qu = udata.get("quality_universe", {})
                    if isinstance(qu, dict):
                        entries = qu.get("companies", [])
                for entry in entries:
                    ticker = entry.get("ticker", "")
                    if ticker:
                        tickers_to_add.append(ticker)
                break
        if not tickers_to_add:
            print("Could not find quality_universe data.")

    if args.from_pipeline:
        pipeline_path = os.path.join(PROJECT_ROOT, "state", "pipeline_tracker.yaml")
        if os.path.exists(pipeline_path):
            with open(pipeline_path) as f:
                pdata = yaml.safe_load(f) or {}
            for section in ["active_research", "watchlist", "pipeline"]:
                for entry in pdata.get(section, []):
                    ticker = entry.get("ticker", "")
                    if ticker:
                        tickers_to_add.append(ticker)

    if args.tickers:
        tickers_to_add.extend(args.tickers)

    if not tickers_to_add:
        print("No tickers to enroll. Use --from-universe, --from-pipeline, or provide tickers.")
        return

    # Deduplicate
    seen = set()
    unique_tickers = []
    for t in tickers_to_add:
        if t not in seen:
            seen.add(t)
            unique_tickers.append(t)

    for ticker in unique_tickers:
        if ticker not in G.nodes:
            country = ticker_country(ticker)
            G.add_node(ticker, type="stock", name="", sector="", country=country,
                       in_portfolio=False, in_universe=True)
            added += 1
        else:
            G.nodes[ticker]["in_universe"] = True

    save_graph(G)
    print(f"Enrolled: {added} new stocks added, {len(unique_tickers)} total processed "
          f"({len(unique_tickers) - added} already in graph)")

    # Auto-resolve new tickers
    if added > 0 and not args.skip_resolve:
        new_tickers = [t for t in unique_tickers if t not in identifiers_data["identifiers"]]
        if new_tickers:
            print(f"\nAuto-resolving {len(new_tickers)} new tickers...")
            cmd_resolve(argparse.Namespace(tickers=new_tickers, force=False))


# ---------------------------------------------------------------------------
# CMD: report
# ---------------------------------------------------------------------------

def cmd_report(args):
    G = load_graph()
    stocks = nodes_by_type(G, "stock")

    if args.portfolio_only:
        stocks = {n: d for n, d in stocks.items() if d.get("in_portfolio")}

    if not stocks:
        print("No stocks in graph. Run: python3 tools/smart_money.py sync-portfolio")
        return

    print("=== SMART MONEY REPORT ===\n")

    for ticker in sorted(stocks.keys()):
        data = stocks[ticker]
        name = data.get("name", "")
        in_pf = " [PORTFOLIO]" if data.get("in_portfolio") else ""
        country = data.get("country", "")

        # Gather edges
        holders = []
        shorts = []
        insider_buys = []
        insider_sells = []

        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            if rel == "holds":
                holders.append((u, d))
            elif rel == "shorts":
                shorts.append((u, d))
            elif rel == "insider_buy":
                insider_buys.append((u, d))
            elif rel == "insider_sell":
                insider_sells.append((u, d))

        print(f"  {ticker} ({name}){in_pf}")
        print(f"    Country: {country}")

        if holders:
            print(f"    Holders ({len(holders)}):")
            for fund_id, ed in sorted(holders, key=lambda x: -(x[1].get("value_usd", 0) or 0))[:5]:
                fund_name = G.nodes[fund_id].get("full_name", fund_id) if fund_id in G.nodes else fund_id
                pct = ed.get("pct_portfolio", "")
                change = ed.get("change", "")
                quarter = ed.get("quarter", "")
                pct_str = f" ({pct}% of pf)" if pct else ""
                change_str = f" [{change}]" if change else ""
                q_str = f" Q:{quarter}" if quarter else ""
                print(f"      {fund_name}{pct_str}{change_str}{q_str}")
        else:
            print("    Holders: none tracked")

        if shorts:
            total_short = sum(s[1].get("pct_shares", 0) or 0 for s in shorts)
            print(f"    Shorts ({len(shorts)}, total {total_short:.2f}%):")
            for fund_id, ed in sorted(shorts, key=lambda x: -(x[1].get("pct_shares", 0) or 0)):
                fund_name = G.nodes[fund_id].get("full_name", fund_id) if fund_id in G.nodes else fund_id
                pct = ed.get("pct_shares", 0)
                dt = ed.get("date", "")
                print(f"      {fund_name}: {pct:.2f}% ({dt})")
        else:
            print("    Shorts: none reported")

        # Insider activity
        if insider_buys or insider_sells:
            buy_total = sum(b[1].get("value", 0) or 0 for b in insider_buys)
            sell_total = sum(s[1].get("value", 0) or 0 for s in insider_sells)
            net = buy_total - sell_total
            net_label = "NET BUY" if net > 0 else "NET SELL"
            print(f"    Insiders: {len(insider_buys)} buys (${buy_total:,.0f}), {len(insider_sells)} sells (${sell_total:,.0f}) → {net_label} ${abs(net):,.0f}")
        else:
            print("    Insiders: no data")

        print()


# ---------------------------------------------------------------------------
# CMD: stock-profile
# ---------------------------------------------------------------------------

def cmd_stock_profile(args):
    G = load_graph()
    ticker = args.ticker

    if ticker not in G.nodes:
        print(f"Ticker {ticker} not in graph. Run sync-portfolio or add-node first.")
        return

    data = G.nodes[ticker]
    print(f"=== STOCK PROFILE: {ticker} ===\n")
    print(f"  Name:         {data.get('name', 'N/A')}")
    print(f"  Country:      {data.get('country', 'N/A')}")
    print(f"  Sector:       {data.get('sector', 'N/A')}")
    print(f"  In Portfolio: {data.get('in_portfolio', False)}")
    print(f"  Conviction:   {data.get('conviction', 'N/A')}")

    # Holders
    holders = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds"]
    print(f"\n  Fund Holders ({len(holders)}):")
    if holders:
        for fund_id, ed in sorted(holders, key=lambda x: -(x[1].get("value_usd", 0) or 0)):
            fund_name = G.nodes[fund_id].get("full_name", fund_id) if fund_id in G.nodes else fund_id
            fund_type = G.nodes[fund_id].get("fund_type", "") if fund_id in G.nodes else ""
            shares = ed.get("shares", "")
            val = ed.get("value_usd", "")
            pct = ed.get("pct_portfolio", "")
            change = ed.get("change", "")
            q = ed.get("quarter", "")
            parts = [fund_name]
            if fund_type:
                parts[0] += f" ({fund_type})"
            if shares:
                parts.append(f"shares={shares:,}" if isinstance(shares, int) else f"shares={shares}")
            if val:
                parts.append(f"${val:,.0f}" if isinstance(val, (int, float)) else f"${val}")
            if pct:
                parts.append(f"{pct}% of portfolio")
            if change:
                parts.append(f"[{change}]")
            if q:
                parts.append(f"Q:{q}")
            print(f"    {' | '.join(parts)}")
    else:
        print("    No tracked holders.")

    # Shorts
    shorts = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "shorts"]
    total_short = sum(s[1].get("pct_shares", 0) or 0 for s in shorts)
    print(f"\n  Short Positions ({len(shorts)}, total {total_short:.2f}%):")
    if shorts:
        for fund_id, ed in sorted(shorts, key=lambda x: -(x[1].get("pct_shares", 0) or 0)):
            fund_name = G.nodes[fund_id].get("full_name", fund_id) if fund_id in G.nodes else fund_id
            pct = ed.get("pct_shares", 0)
            dt = ed.get("date", "")
            src = ed.get("data_source", "")
            print(f"    {fund_name}: {pct:.2f}% ({dt}) [{src}]")
    else:
        print("    No short positions reported.")

    # Insiders
    insider_buys = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "insider_buy"]
    insider_sells = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "insider_sell"]
    print(f"\n  Insider Activity:")
    if insider_buys:
        print(f"    Buys ({len(insider_buys)}):")
        for person_id, ed in sorted(insider_buys, key=lambda x: x[1].get("date", ""), reverse=True)[:10]:
            name = G.nodes[person_id].get("full_name", person_id) if person_id in G.nodes else person_id
            role = ed.get("role", "")
            val = ed.get("value", 0)
            dt = ed.get("date", "")
            print(f"      {name} ({role}): ${val:,.0f} on {dt}")
    if insider_sells:
        print(f"    Sells ({len(insider_sells)}):")
        for person_id, ed in sorted(insider_sells, key=lambda x: x[1].get("date", ""), reverse=True)[:10]:
            name = G.nodes[person_id].get("full_name", person_id) if person_id in G.nodes else person_id
            role = ed.get("role", "")
            val = ed.get("value", 0)
            dt = ed.get("date", "")
            print(f"      {name} ({role}): ${val:,.0f} on {dt}")
    if not insider_buys and not insider_sells:
        print("    No insider data.")

    # Crowding
    median_holders = _median_holder_count(G)
    crowding = len(holders) / median_holders if median_holders > 0 else 0
    print(f"\n  Crowding Score: {crowding:.2f}x median ({len(holders)} holders, median={median_holders:.0f})")

    # Co-holdings (which other stocks share the most holders with this one)
    co_holdings = _co_holdings(G, ticker)
    if co_holdings:
        print(f"\n  Co-held with (top 5 by shared holders):")
        for other, count in co_holdings[:5]:
            other_name = G.nodes[other].get("name", other)
            print(f"    {other} ({other_name}): {count} shared holders")

    # Contrarian score
    if shorts and holders:
        quality_holders = sum(1 for u, d in holders
                              if G.nodes.get(u, {}).get("fund_type") in ("value", "activist", "quality"))
        if total_short > 2 and quality_holders >= 2:
            print(f"\n  CONTRARIAN SIGNAL: High short ({total_short:.1f}%) + {quality_holders} quality holders")

    print()


def _median_holder_count(G):
    """Get median holder count across all stock nodes."""
    stocks = nodes_by_type(G, "stock")
    counts = []
    for ticker in stocks:
        holders = sum(1 for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds")
        counts.append(holders)
    if not counts:
        return 1
    counts.sort()
    mid = len(counts) // 2
    return counts[mid] if len(counts) % 2 else (counts[mid - 1] + counts[mid]) / 2


def _co_holdings(G, ticker, min_shared=1):
    """Find stocks that share holders with the given ticker."""
    my_holders = set()
    for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
        if d.get("relation") == "holds":
            my_holders.add(u)

    if not my_holders:
        return []

    co_counts = defaultdict(int)
    for holder in my_holders:
        for u, v, k, d in G.out_edges(holder, data=True, keys=True):
            if d.get("relation") == "holds" and v != ticker:
                co_counts[v] += 1

    return sorted(co_counts.items(), key=lambda x: -x[1])


# ---------------------------------------------------------------------------
# CMD: who-holds
# ---------------------------------------------------------------------------

def cmd_who_holds(args):
    G = load_graph()
    ticker = args.ticker
    if ticker not in G.nodes:
        print(f"Ticker {ticker} not in graph.")
        return

    holders = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds"]
    print(f"=== WHO HOLDS {ticker} ({len(holders)} tracked funds) ===\n")
    for fund_id, ed in sorted(holders, key=lambda x: -(x[1].get("value_usd", 0) or 0)):
        fund_name = G.nodes[fund_id].get("full_name", fund_id) if fund_id in G.nodes else fund_id
        fund_type = G.nodes[fund_id].get("fund_type", "") if fund_id in G.nodes else ""
        change = ed.get("change", "")
        q = ed.get("quarter", "")
        val = ed.get("value_usd", "")
        val_str = f"${val:,.0f}" if isinstance(val, (int, float)) else ""
        print(f"  {fund_name:30s}  type={fund_type:10s}  {val_str:>12s}  {change:6s}  Q:{q}")
    print()


# ---------------------------------------------------------------------------
# CMD: crowding
# ---------------------------------------------------------------------------

def cmd_crowding(args):
    G = load_graph()
    stocks = nodes_by_type(G, "stock")
    top_n = args.top or 20
    median = _median_holder_count(G)

    scores = []
    for ticker, data in stocks.items():
        holders = sum(1 for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds")
        crowding = holders / median if median > 0 else 0
        in_pf = data.get("in_portfolio", False)
        scores.append((ticker, data.get("name", ""), holders, crowding, in_pf))

    scores.sort(key=lambda x: -x[3])
    print(f"=== CROWDING ANALYSIS (top {top_n}, median holders={median:.0f}) ===\n")
    print(f"  {'Ticker':12s} {'Name':30s} {'Holders':>8s} {'Crowding':>10s} {'Portfolio':>10s}")
    print(f"  {'-'*12} {'-'*30} {'-'*8} {'-'*10} {'-'*10}")
    for ticker, name, holders, crowding, in_pf in scores[:top_n]:
        pf_marker = "YES" if in_pf else ""
        print(f"  {ticker:12s} {name[:30]:30s} {holders:>8d} {crowding:>10.2f}x {pf_marker:>10s}")
    print()


# ---------------------------------------------------------------------------
# CMD: alerts
# ---------------------------------------------------------------------------

def cmd_alerts(args):
    """Compare current graph vs most recent snapshot and report changes."""
    G = load_graph()

    # Find most recent snapshot (not today's)
    snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
    prev_snap = None
    for s in reversed(snaps):
        snap_date = s[6:16]
        if snap_date != today_str():
            prev_snap = s
            break

    if not prev_snap:
        print("No previous snapshot for comparison. Run 'snapshot' first, then again after updates.")
        return

    with open(os.path.join(SNAPSHOTS_DIR, prev_snap)) as f:
        prev_data = json.load(f)
    G_prev = nx.node_link_graph(prev_data, directed=True, multigraph=True)

    print(f"=== ALERTS (vs {prev_snap}) ===\n")
    portfolio_tickers = {n for n, d in G.nodes(data=True) if d.get("type") == "stock" and d.get("in_portfolio")}

    alerts = []

    for ticker in portfolio_tickers:
        # Short changes
        curr_shorts = {}
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "shorts":
                curr_shorts[u] = d.get("pct_shares", 0) or 0

        prev_shorts = {}
        if ticker in G_prev.nodes:
            for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True):
                if d.get("relation") == "shorts":
                    prev_shorts[u] = d.get("pct_shares", 0) or 0

        curr_total = sum(curr_shorts.values())
        prev_total = sum(prev_shorts.values())
        delta = curr_total - prev_total

        if delta > 0.3:
            alerts.append(("SHORT_INCREASE", ticker, f"short {prev_total:.2f}% → {curr_total:.2f}% (+{delta:.2f}pp)"))
        elif delta < -0.3:
            alerts.append(("SHORT_DECREASE", ticker, f"short {prev_total:.2f}% → {curr_total:.2f}% ({delta:.2f}pp) [bullish]"))

        # Holder changes
        curr_holders = set()
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "holds":
                curr_holders.add(u)

        prev_holders = set()
        if ticker in G_prev.nodes:
            for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True):
                if d.get("relation") == "holds":
                    prev_holders.add(u)

        new_holders = curr_holders - prev_holders
        exit_holders = prev_holders - curr_holders

        for h in new_holders:
            fund_name = G.nodes[h].get("full_name", h) if h in G.nodes else h
            alerts.append(("NEW_HOLDER", ticker, f"{fund_name} entered position"))
        for h in exit_holders:
            fund_name = G_prev.nodes[h].get("full_name", h) if h in G_prev.nodes else h
            alerts.append(("EXIT_HOLDER", ticker, f"{fund_name} exited position"))

        # Insider cluster buys (3+ in 30 days)
        recent_buys = []
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "insider_buy":
                buy_date = d.get("date", "")
                days = days_since(buy_date)
                if days is not None and days <= 30:
                    recent_buys.append((u, d))

        if len(recent_buys) >= 3:
            total_val = sum(b[1].get("value", 0) or 0 for b in recent_buys)
            alerts.append(("INSIDER_CLUSTER_BUY", ticker, f"{len(recent_buys)} insiders bought in 30d (${total_val:,.0f} total)"))

        # Convergence signal
        if delta < -0.3 and len(recent_buys) >= 2:
            alerts.append(("CONVERGENCE", ticker, "insider buy + short decrease = strong bullish signal"))

    if alerts:
        for alert_type, ticker, msg in sorted(alerts, key=lambda x: x[0]):
            marker = "!!" if alert_type in ("CONVERGENCE", "INSIDER_CLUSTER_BUY") else "  "
            print(f"  {marker} [{alert_type}] {ticker}: {msg}")
    else:
        print("  No alerts. Graph unchanged vs previous snapshot.")

    print()


# ---------------------------------------------------------------------------
# CMD: metrics
# ---------------------------------------------------------------------------

def cmd_metrics(args):
    G = load_graph()
    if G.number_of_nodes() == 0:
        print("Empty graph. Nothing to analyze.")
        return

    print("=== GRAPH METRICS ===\n")

    # Convert to simple DiGraph for PageRank
    G_simple = nx.DiGraph()
    for u, v, d in G.edges(data=True):
        if G_simple.has_edge(u, v):
            G_simple[u][v]["weight"] = G_simple[u][v].get("weight", 1) + 1
        else:
            G_simple.add_edge(u, v, weight=1)
    # Add isolated nodes
    for n in G.nodes:
        if n not in G_simple:
            G_simple.add_node(n)

    # PageRank
    try:
        pr = nx.pagerank(G_simple, weight="weight")
        funds = {n: d for n, d in G.nodes(data=True) if d.get("type") == "fund"}
        fund_pr = {n: pr.get(n, 0) for n in funds}
        top_funds = sorted(fund_pr.items(), key=lambda x: -x[1])[:10]

        print("  PageRank — Top Funds (centrality):")
        for fund_id, score in top_funds:
            name = G.nodes[fund_id].get("full_name", fund_id)
            print(f"    {name:35s}  PR={score:.4f}")
    except Exception as e:
        print(f"  PageRank failed: {e}")

    # Degree centrality for stocks
    stocks = {n: d for n, d in G.nodes(data=True) if d.get("type") == "stock"}
    stock_degree = {}
    for ticker in stocks:
        in_deg = sum(1 for u, v, k, d in G.in_edges(ticker, data=True, keys=True)
                     if d.get("relation") in ("holds", "shorts", "insider_buy"))
        stock_degree[ticker] = in_deg

    top_stocks = sorted(stock_degree.items(), key=lambda x: -x[1])[:10]
    print("\n  Most Connected Stocks (in-degree):")
    for ticker, deg in top_stocks:
        name = stocks[ticker].get("name", "")
        in_pf = " [PF]" if stocks[ticker].get("in_portfolio") else ""
        print(f"    {ticker:12s} {name[:25]:25s}  connections={deg}{in_pf}")

    print()


# ---------------------------------------------------------------------------
# CMD: communities
# ---------------------------------------------------------------------------

def cmd_communities(args):
    G = load_graph()
    if G.number_of_nodes() < 3:
        print("Too few nodes for community detection.")
        return

    import community as community_louvain

    # Build undirected version for Louvain
    G_undirected = nx.Graph()
    for u, v, d in G.edges(data=True):
        if d.get("relation") == "holds":
            if G_undirected.has_edge(u, v):
                G_undirected[u][v]["weight"] += 1
            else:
                G_undirected.add_edge(u, v, weight=1)

    if G_undirected.number_of_edges() == 0:
        print("No 'holds' edges for community detection. Need fund holder data.")
        return

    partition = community_louvain.best_partition(G_undirected, weight="weight")

    # Group by community
    communities = defaultdict(list)
    for node, comm_id in partition.items():
        node_type = G.nodes[node].get("type", "unknown") if node in G.nodes else "unknown"
        communities[comm_id].append((node, node_type))

    print(f"=== COMMUNITIES (Louvain, {len(communities)} detected) ===\n")
    for comm_id in sorted(communities.keys()):
        members = communities[comm_id]
        stocks_in = [(n, t) for n, t in members if t == "stock"]
        funds_in = [(n, t) for n, t in members if t == "fund"]
        if not stocks_in and not funds_in:
            continue

        print(f"  Community {comm_id}:")
        if stocks_in:
            stock_names = []
            for n, _ in stocks_in:
                name = G.nodes[n].get("name", n) if n in G.nodes else n
                pf = " *" if G.nodes.get(n, {}).get("in_portfolio") else ""
                stock_names.append(f"{n}{pf}")
            print(f"    Stocks: {', '.join(stock_names)}")
        if funds_in:
            fund_names = [G.nodes[n].get("full_name", n) if n in G.nodes else n for n, _ in funds_in]
            print(f"    Funds:  {', '.join(fund_names[:5])}")
        print()


# ---------------------------------------------------------------------------
# CMD: visualize
# ---------------------------------------------------------------------------

def cmd_visualize(args):
    G = load_graph()
    if G.number_of_nodes() == 0:
        print("Empty graph. Nothing to visualize.")
        return

    from pyvis.network import Network

    output = args.output or os.path.join(DATA_DIR, "smart_money_graph.html")

    # Filter to portfolio + standing orders if requested
    if args.portfolio_only:
        focus_tickers = {n for n, d in G.nodes(data=True)
                         if d.get("type") == "stock" and (d.get("in_portfolio") or d.get("has_standing_order"))}
        # Get ego graph: focus stocks + their neighbors
        keep_nodes = set(focus_tickers)
        for t in focus_tickers:
            for u, v, k, d in G.in_edges(t, data=True, keys=True):
                keep_nodes.add(u)
            for u, v, k, d in G.out_edges(t, data=True, keys=True):
                keep_nodes.add(v)
        G_viz = G.subgraph(keep_nodes).copy()
    else:
        G_viz = G

    # --hide-islands: remove stock nodes with zero in-degree AND zero out-degree
    if args.hide_islands:
        island_nodes = [n for n, d in G_viz.nodes(data=True)
                        if d.get("type") == "stock"
                        and G_viz.in_degree(n) == 0
                        and G_viz.out_degree(n) == 0]
        if island_nodes:
            G_viz = G_viz.copy() if not isinstance(G_viz, nx.MultiDiGraph) else G_viz
            # Ensure we have a mutable copy
            if not hasattr(G_viz, '_adj') or G_viz is G:
                G_viz = G_viz.copy()
            for n in island_nodes:
                G_viz.remove_node(n)
            print(f"Hidden {len(island_nodes)} island nodes: {', '.join(sorted(island_nodes))}")

    net = Network(height="800px", width="100%", directed=True, cdn_resources="in_line",
                  bgcolor="#1a1a2e", font_color="white")
    net.barnes_hut(gravity=-3000, central_gravity=0.3, spring_length=150)

    # Add nodes with styling
    for node, data in G_viz.nodes(data=True):
        node_type = data.get("type", "unknown")
        label = node
        title_parts = [f"<b>{node}</b>"]

        if node_type == "stock":
            name = data.get("name", "")
            in_pf = data.get("in_portfolio", False)
            has_so = data.get("has_standing_order", False)
            so_category = data.get("so_category", "")
            country = data.get("country", "")
            # Count connections
            n_holders = sum(1 for _, _, _, d in G_viz.in_edges(node, data=True, keys=True) if d.get("relation") == "holds")
            n_shorts = sum(1 for _, _, _, d in G_viz.in_edges(node, data=True, keys=True) if d.get("relation") == "shorts")

            # Color logic: portfolio=red/gold, SO ACTIVE=cyan/gold, SO GATED=cyan/grey, plain=blue
            if in_pf:
                color = "#e94560"
                border_color = "#ffd700"
                border_width = 3
            elif has_so and so_category == "ACTIVE":
                color = "#00bcd4"
                border_color = "#ffd700"
                border_width = 3
            elif has_so:
                color = "#00bcd4"
                border_color = "#607d8b"
                border_width = 2
            else:
                color = "#4a90d9"
                border_color = "#4a90d9"
                border_width = 1

            size = 20 + n_holders * 3
            # SO-only stocks slightly smaller
            if not in_pf and has_so:
                size = max(15, size - 3)
            label = node if len(node) <= 8 else node[:8]
            title_parts.extend([name, f"Country: {country}", f"Holders: {n_holders}", f"Shorts: {n_shorts}"])
            if in_pf:
                title_parts.append("IN PORTFOLIO")
                conviction = data.get("conviction", "")
                if conviction:
                    title_parts.append(f"Conviction: {conviction}")
            if has_so:
                so_action = data.get("so_action", "")
                so_trigger = data.get("so_trigger", "")
                so_tier = data.get("so_tier", "")
                so_fv = data.get("so_fair_value", "")
                title_parts.append(f"STANDING ORDER: {so_action} {so_category}")
                if so_trigger:
                    title_parts.append(f"Trigger: {so_trigger}")
                if so_fv:
                    title_parts.append(f"Fair Value: {so_fv}")
                if so_tier:
                    title_parts.append(f"Tier: {so_tier}")

            net.add_node(node, label=label, color={"background": color, "border": border_color},
                         borderWidth=border_width, size=size,
                         title="<br>".join(title_parts), shape="dot")

        elif node_type == "fund":
            fund_type = data.get("fund_type", "unknown")
            full_name = data.get("full_name", node)
            colors = {"value": "#2ecc71", "activist": "#e67e22", "quality": "#27ae60",
                      "quant": "#95a5a6", "index": "#7f8c8d"}
            color = colors.get(fund_type, "#3498db")
            n_holdings = sum(1 for _, _, _, d in G_viz.out_edges(node, data=True, keys=True) if d.get("relation") == "holds")
            size = 15 + n_holdings * 2
            label = full_name[:15] if len(full_name) > 15 else full_name
            title_parts.extend([full_name, f"Type: {fund_type}", f"Holdings tracked: {n_holdings}"])
            net.add_node(node, label=label, color=color, size=size,
                         title="<br>".join(title_parts), shape="diamond")

        elif node_type == "person":
            full_name = data.get("full_name", node)
            role = data.get("role", "")
            # Check for recent insider buys
            has_recent_buy = any(d.get("relation") == "insider_buy" and days_since(d.get("date")) is not None and days_since(d.get("date")) <= 90
                                 for _, _, _, d in G_viz.out_edges(node, data=True, keys=True))
            color = "#e74c3c" if has_recent_buy else "#bdc3c7"
            size = 10
            label = full_name[:12] if len(full_name) > 12 else full_name
            title_parts.extend([full_name, f"Role: {role}"])
            net.add_node(node, label=label, color=color, size=size,
                         title="<br>".join(title_parts), shape="triangle")

    # Add edges with styling
    for u, v, k, d in G_viz.edges(data=True, keys=True):
        relation = d.get("relation", "")
        edge_colors = {"holds": "#2ecc71", "shorts": "#e74c3c",
                       "insider_buy": "#f1c40f", "insider_sell": "#e67e22",
                       "manages": "#95a5a6"}
        color = edge_colors.get(relation, "#7f8c8d")
        width = 1
        if relation == "holds":
            val = d.get("value_usd", 0) or 0
            width = max(1, min(5, val / 10_000_000)) if isinstance(val, (int, float)) else 1
        elif relation == "shorts":
            pct = d.get("pct_shares", 0) or 0
            width = max(1, min(5, pct * 2))

        title = f"{relation}"
        if d.get("quarter"):
            title += f" (Q:{d['quarter']})"
        if d.get("pct_shares"):
            title += f" {d['pct_shares']:.2f}%"

        net.add_edge(u, v, color=color, width=width, title=title, arrows="to")

    net.save_graph(output)

    # Inject floating legend into saved HTML
    legend_html = """
<div id="sm-legend" style="
    position:fixed; bottom:16px; left:16px; z-index:9999;
    background:rgba(26,26,46,0.92); border:1px solid #444; border-radius:8px;
    padding:14px 18px; color:#ddd; font-family:sans-serif; font-size:12px;
    max-width:260px; line-height:1.6; box-shadow:0 2px 12px rgba(0,0,0,0.5);
">
<div style="font-weight:bold;font-size:13px;margin-bottom:8px;color:#fff;">Legend</div>

<div style="margin-bottom:6px;font-weight:600;color:#aaa;">Node Shapes</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><circle cx="7" cy="7" r="6" fill="#4a90d9"/></svg>
  <span>Stock</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><polygon points="7,1 13,7 7,13 1,7" fill="#2ecc71"/></svg>
  <span>Fund</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
  <svg width="14" height="14"><polygon points="7,1 13,13 1,13" fill="#bdc3c7"/></svg>
  <span>Person</span>
</div>

<div style="margin-bottom:6px;font-weight:600;color:#aaa;">Stock Colors</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><circle cx="7" cy="7" r="5" fill="#e94560" stroke="#ffd700" stroke-width="2"/></svg>
  <span>Portfolio</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><circle cx="7" cy="7" r="5" fill="#00bcd4" stroke="#ffd700" stroke-width="2"/></svg>
  <span>SO Active</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><circle cx="7" cy="7" r="5" fill="#00bcd4" stroke="#607d8b" stroke-width="2"/></svg>
  <span>SO Gated</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
  <svg width="14" height="14"><circle cx="7" cy="7" r="5" fill="#4a90d9" stroke="#4a90d9" stroke-width="1"/></svg>
  <span>Other</span>
</div>

<div style="margin-bottom:6px;font-weight:600;color:#aaa;">Fund Colors</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="14" height="14"><polygon points="7,1 13,7 7,13 1,7" fill="#2ecc71"/></svg>
  <span>Value</span>
  <svg width="14" height="14" style="margin-left:6px;"><polygon points="7,1 13,7 7,13 1,7" fill="#e67e22"/></svg>
  <span>Activist</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
  <svg width="14" height="14"><polygon points="7,1 13,7 7,13 1,7" fill="#27ae60"/></svg>
  <span>Quality</span>
  <svg width="14" height="14" style="margin-left:6px;"><polygon points="7,1 13,7 7,13 1,7" fill="#95a5a6"/></svg>
  <span>Quant</span>
</div>

<div style="margin-bottom:6px;font-weight:600;color:#aaa;">Edge Colors</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:3px;">
  <svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#2ecc71" stroke-width="2"/></svg>
  <span>Holds</span>
  <svg width="20" height="10" style="margin-left:6px;"><line x1="0" y1="5" x2="20" y2="5" stroke="#e74c3c" stroke-width="2"/></svg>
  <span>Shorts</span>
</div>
<div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
  <svg width="20" height="10"><line x1="0" y1="5" x2="20" y2="5" stroke="#f1c40f" stroke-width="2"/></svg>
  <span>Insider buy</span>
  <svg width="20" height="10" style="margin-left:6px;"><line x1="0" y1="5" x2="20" y2="5" stroke="#e67e22" stroke-width="2"/></svg>
  <span>Insider sell</span>
</div>

<div style="color:#777;font-size:10px;margin-top:4px;">Click node to highlight. Hover for details.</div>
</div>
"""
    with open(output, "r") as f:
        html = f.read()
    html = html.replace("</body>", legend_html + "</body>")
    with open(output, "w") as f:
        f.write(html)

    print(f"Visualization saved: {output}")
    print(f"Open in browser to view interactively.")


# ---------------------------------------------------------------------------
# CMD: signals (WS2 — Signal Engine)
# ---------------------------------------------------------------------------

def cmd_signals(args):
    """Detect actionable signals from graph edge patterns.

    Signals:
      INSIDER_CLUSTER_BUY  — 3+ insiders buy within 60 days
      CONVERGENCE          — 2+ quality/value funds hold same stock
      SHORT_SQUEEZE_RISK   — SI >10% + quality fund holds + insider buying
      SMART_EXIT           — Quality fund exited position (was holding, now gone)
      SHORT_ESCALATION     — Total SI > 8% with 5+ funds short
      QUIET_ACCUMULATION   — New fund position + low SI + insider buy
      HERD_WARNING         — 5+ funds hold same stock (crowding >3x median)
      FOUNDER_CONVICTION   — CEO/founder buying >$100K in 90 days
    """
    G = load_graph()
    stocks = nodes_by_type(G, "stock")
    portfolio_tickers = {n for n, d in stocks.items() if d.get("in_portfolio")}

    if args.ticker:
        target_stocks = {args.ticker: stocks.get(args.ticker, {})} if args.ticker in stocks else {}
    elif args.portfolio_only:
        target_stocks = {n: d for n, d in stocks.items() if d.get("in_portfolio")}
    else:
        target_stocks = stocks

    if not target_stocks:
        print("No stocks to analyze. Run sync-portfolio first.")
        return

    # Load previous snapshot for SMART_EXIT detection
    G_prev = None
    if os.path.exists(SNAPSHOTS_DIR):
        snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
        for s in reversed(snaps):
            snap_date = s[6:16]
            if snap_date != today_str():
                try:
                    with open(os.path.join(SNAPSHOTS_DIR, s)) as f:
                        G_prev = nx.node_link_graph(json.load(f), directed=True, multigraph=True)
                except Exception:
                    pass
                break

    median = _median_holder_count(G)
    signals = []

    for ticker, sdata in target_stocks.items():
        # Gather edges
        holders = []
        shorts = []
        insider_buys = []

        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            if rel == "holds":
                holders.append((u, d))
            elif rel == "shorts":
                shorts.append((u, d))
            elif rel == "insider_buy":
                insider_buys.append((u, d))

        total_si = sum(s[1].get("pct_shares", 0) or 0 for s in shorts)
        n_short_funds = len(shorts)
        in_pf = ticker in portfolio_tickers

        # Quality/value fund holders
        quality_holders = []
        for fund_id, ed in holders:
            ft = G.nodes.get(fund_id, {}).get("fund_type", "")
            if ft in ("value", "quality", "activist"):
                quality_holders.append((fund_id, ed))

        # Recent insider buys (within 60 days)
        recent_buys_60d = []
        recent_buys_90d = []
        for person_id, ed in insider_buys:
            buy_date = ed.get("date", "")
            days = days_since(buy_date)
            if days is not None:
                if days <= 60:
                    recent_buys_60d.append((person_id, ed))
                if days <= 90:
                    recent_buys_90d.append((person_id, ed))

        # --- Signal: INSIDER_CLUSTER_BUY ---
        if len(recent_buys_60d) >= 3:
            total_val = sum(b[1].get("value", 0) or 0 for b in recent_buys_60d)
            signals.append(("INSIDER_CLUSTER_BUY", ticker, "STRONG BULL",
                            f"{len(recent_buys_60d)} insiders bought in 60d (${total_val:,.0f})"))

        # --- Signal: CONVERGENCE ---
        if len(quality_holders) >= 2:
            names = [G.nodes.get(h[0], {}).get("full_name", h[0])[:20] for h in quality_holders[:3]]
            signals.append(("CONVERGENCE", ticker, "BULL",
                            f"{len(quality_holders)} quality/value funds: {', '.join(names)}"))

        # --- Signal: SHORT_SQUEEZE_RISK ---
        if total_si > 10 and quality_holders and recent_buys_90d:
            signals.append(("SHORT_SQUEEZE_RISK", ticker, "CONTROVERSY",
                            f"SI {total_si:.1f}% + {len(quality_holders)} quality holders + insider buying"))

        # --- Signal: SMART_EXIT ---
        if G_prev and ticker in G_prev.nodes:
            prev_holders = {u for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True)
                           if d.get("relation") == "holds"}
            curr_holders = {h[0] for h in holders}
            exited = prev_holders - curr_holders
            for ex_fund in exited:
                ft = G_prev.nodes.get(ex_fund, {}).get("fund_type", "")
                if ft in ("value", "quality", "activist"):
                    name = G_prev.nodes.get(ex_fund, {}).get("full_name", ex_fund)
                    signals.append(("SMART_EXIT", ticker, "BEAR WARNING",
                                    f"{name} ({ft}) exited position"))

        # --- Signal: SHORT_ESCALATION ---
        if total_si > 8 and n_short_funds >= 5:
            signals.append(("SHORT_ESCALATION", ticker, "BEAR",
                            f"SI {total_si:.1f}% ({n_short_funds} funds)"))

        # --- Signal: QUIET_ACCUMULATION ---
        if holders and total_si < 3 and recent_buys_90d and not in_pf:
            new_holder = False
            if G_prev and ticker in G_prev.nodes:
                prev_h = {u for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True)
                          if d.get("relation") == "holds"}
                curr_h = {h[0] for h in holders}
                new_holder = bool(curr_h - prev_h)
            if new_holder or not G_prev:
                signals.append(("QUIET_ACCUMULATION", ticker, "STEALTH BULL",
                                f"New fund + low SI ({total_si:.1f}%) + insider buying"))

        # --- Signal: HERD_WARNING ---
        crowding = len(holders) / median if median > 0 else 0
        if len(holders) >= 5 and crowding > 3.0:
            signals.append(("HERD_WARNING", ticker, "RISK",
                            f"{len(holders)} funds hold ({crowding:.1f}x median)"))

        # --- Signal: FOUNDER_CONVICTION ---
        for person_id, ed in recent_buys_90d:
            role = (ed.get("role", "") or "").lower()
            value = ed.get("value", 0) or 0
            if value >= 100000 and any(w in role for w in ["ceo", "chief executive", "founder", "chairman", "director"]):
                name = G.nodes.get(person_id, {}).get("full_name", person_id)
                signals.append(("FOUNDER_CONVICTION", ticker, "CONVICTION",
                                f"{name} ({ed.get('role', '')}) ${value:,.0f}"))
                break  # One per ticker

    # Output
    print(f"=== SMART MONEY SIGNALS ({len(signals)} detected) ===\n")
    if not signals:
        print("  No signals detected. Graph may need more data (run refresh).")
        print()
        return

    # Sort by signal type for grouping
    signals.sort(key=lambda x: (x[0], x[1]))

    # Column formatting
    pf_tickers = portfolio_tickers
    for sig_type, ticker, weight, detail in signals:
        pf_marker = " [PF]" if ticker in pf_tickers else ""
        print(f"  {ticker:12s}  {sig_type:24s}  {weight:14s}  {detail}{pf_marker}")

    # Summary
    by_type = defaultdict(int)
    for s in signals:
        by_type[s[0]] += 1
    print(f"\n  Summary: {', '.join(f'{t}={c}' for t, c in sorted(by_type.items()))}")
    print()


# ---------------------------------------------------------------------------
# CMD: discover (WS6 — Discovery Engine)
# ---------------------------------------------------------------------------

def cmd_discover(args):
    """Find stocks held/shorted by tracked funds that are NOT in our graph.

    Scans all 13F XML files and regulatory data to find stocks we're missing.
    Anti-popularity-bias: these are stocks quality investors actually own.
    """
    import xml.etree.ElementTree as ET
    import re

    G = load_graph()
    our_stocks = set(n for n, d in G.nodes(data=True) if d.get("type") == "stock")

    min_funds = args.min_funds or 2
    source = args.source or "all"

    # Load tracked funds for CIK → metadata
    cik_to_fund = {}
    if os.path.exists(TRACKED_FUNDS_PATH):
        with open(TRACKED_FUNDS_PATH) as f:
            tf = yaml.safe_load(f) or {}
        for fund in tf.get("funds", []):
            cik = str(fund.get("cik", ""))
            if cik:
                cik_to_fund[cik] = fund

    discoveries_13f = defaultdict(lambda: {"funds": [], "total_value": 0, "cusip": ""})
    discoveries_fca = defaultdict(lambda: {"funds": [], "total_pct": 0, "isin": "", "issuer": ""})
    discoveries_amf = defaultdict(lambda: {"funds": [], "total_pct": 0, "isin": "", "issuer": ""})

    # --- 13F Discovery ---
    if source in ("13f", "all") and os.path.exists(HOLDINGS_DIR):
        # Build reverse CUSIP→ticker map for known stocks
        identifiers_data = load_identifiers()
        known_cusips = set()
        for ticker, info in identifiers_data.get("identifiers", {}).items():
            cusip = info.get("cusip")
            if cusip:
                known_cusips.add(cusip)
        # Also from cusip_map
        if os.path.exists(CUSIP_MAP_PATH):
            with open(CUSIP_MAP_PATH) as f:
                cm = yaml.safe_load(f) or {}
            known_cusips.update(cm.get("cusip_to_ticker", {}).keys())

        # Scan all 13F XMLs for unknown CUSIPs
        for fname in sorted(os.listdir(HOLDINGS_DIR)):
            if not fname.endswith(".xml"):
                continue
            fpath = os.path.join(HOLDINGS_DIR, fname)

            cik_match = re.search(r'cik-(\d+)', fname)
            file_cik = cik_match.group(1) if cik_match else None
            fund_info = cik_to_fund.get(file_cik, {}) if file_cik else {}
            fund_name = fund_info.get("name", f"CIK-{file_cik}" if file_cik else "unknown")

            try:
                tree = ET.parse(fpath)
                root = tree.getroot()
            except ET.ParseError:
                continue

            ns = ""
            if "}" in root.tag:
                ns = root.tag.split("}")[0] + "}"

            entries = root.findall(f"{ns}infoTable")
            if not entries:
                continue

            for entry in entries:
                cusip = (entry.findtext(f"{ns}cusip") or "").strip()
                if not cusip or cusip in known_cusips:
                    continue
                name_of_issuer = (entry.findtext(f"{ns}nameOfIssuer") or "").strip()
                value = int(entry.findtext(f"{ns}value") or "0")
                shares_el = entry.find(f"{ns}shrsOrPrnAmt")
                shares = int(shares_el.findtext(f"{ns}sshPrnamt") or "0") if shares_el is not None else 0

                rec = discoveries_13f[cusip]
                rec["cusip"] = cusip
                rec["issuer"] = name_of_issuer or rec["issuer"]
                if fund_name not in rec["funds"]:
                    rec["funds"].append(fund_name)
                rec["total_value"] += value

        # Filter by min_funds
        discoveries_13f = {k: v for k, v in discoveries_13f.items() if len(v["funds"]) >= min_funds}

    # --- FCA Discovery (UK heavily shorted stocks not in our graph) ---
    if source in ("fca", "all"):
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("fca_uk_") and f.endswith(".xlsx")], reverse=True)
        if files:
            try:
                import openpyxl
                fpath = os.path.join(SHORTS_DIR, files[0])
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

                header_idx = 0
                for i, row in enumerate(rows):
                    if "position holder" in str(row).lower() or "issuer" in str(row).lower():
                        header_idx = i
                        break

                headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
                holder_col = next((i for i, h in enumerate(headers) if "position holder" in h), 0)
                isin_col = next((i for i, h in enumerate(headers) if "isin" in h), 2)
                issuer_col = next((i for i, h in enumerate(headers) if "issuer" in h or "name of share" in h), 1)
                pct_col = next((i for i, h in enumerate(headers) if "net short" in h or "position (%)" in h), 3)

                # Build set of ISINs we already track
                identifiers_data = load_identifiers()
                known_isins = set()
                for ticker, info in identifiers_data.get("identifiers", {}).items():
                    isin = info.get("isin")
                    if isin:
                        known_isins.add(isin)

                # Dedup by (holder, isin) keeping latest date — FCA XLSX contains
                # full history, without dedup SI% inflates 10-100x (same bug as ingest-fca)
                date_col = next((i for i, h in enumerate(headers) if "date" in h or "position date" in h), len(headers) - 1)
                active_shorts_disc = {}  # (holder_lower, isin) → {holder, isin, issuer, pct, date}

                for row in rows[header_idx + 1:]:
                    if all(v is None for v in row):
                        continue
                    isin = str(row[isin_col]).strip() if row[isin_col] else ""
                    if not isin.startswith("GB") or isin in known_isins:
                        continue
                    holder = str(row[holder_col]).strip() if row[holder_col] else ""
                    issuer = str(row[issuer_col]).strip() if row[issuer_col] else ""
                    pct = 0
                    if row[pct_col]:
                        try:
                            pct = float(str(row[pct_col]).replace("%", "").strip())
                        except ValueError:
                            pct = 0
                    pos_date = ""
                    if date_col < len(row) and row[date_col]:
                        from datetime import datetime as _dt
                        if isinstance(row[date_col], _dt):
                            pos_date = row[date_col].strftime("%Y-%m-%d")
                        else:
                            pos_date = str(row[date_col])[:10]

                    key = (holder.lower().strip(), isin)
                    if key not in active_shorts_disc or pos_date > active_shorts_disc[key].get("date", ""):
                        active_shorts_disc[key] = {
                            "holder": holder, "isin": isin,
                            "issuer": issuer, "pct": pct, "date": pos_date,
                        }

                # Filter <0.50% (below-threshold exit notifications)
                active_shorts_disc = {k: v for k, v in active_shorts_disc.items() if v["pct"] >= 0.50}

                # Build discoveries from deduped entries
                for key, data in active_shorts_disc.items():
                    isin = data["isin"]
                    rec = discoveries_fca[isin]
                    rec["isin"] = isin
                    rec["issuer"] = data["issuer"] or rec["issuer"]
                    if data["holder"] not in rec["funds"]:
                        rec["funds"].append(data["holder"])
                    rec["total_pct"] += data["pct"]

            except ImportError:
                pass

        discoveries_fca = {k: v for k, v in discoveries_fca.items() if len(v["funds"]) >= min_funds}

    # --- AMF Discovery (FR heavily shorted stocks not in our graph) ---
    if source in ("amf", "all"):
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("amf_france_") and f.endswith(".csv")], reverse=True)
        if files:
            fpath = os.path.join(SHORTS_DIR, files[0])
            identifiers_data = load_identifiers()
            known_isins = set()
            for ticker, info in identifiers_data.get("identifiers", {}).items():
                isin = info.get("isin")
                if isin:
                    known_isins.add(isin)

            try:
                # Dedup by (holder_lower, isin) keeping latest entry — same pattern as FCA
                amf_active = {}  # (holder_lower, isin) → {holder, isin, issuer, ratio, date}
                with open(fpath, "r", encoding="utf-8", errors="replace") as f:
                    f.readline()  # skip header
                    for line in f:
                        parts = line.strip().split(";")
                        if len(parts) < 8:
                            continue
                        holder = parts[0].strip().strip('"')
                        issuer = parts[2].strip().strip('"')
                        ratio_str = parts[3].strip().strip('"')
                        isin = parts[4].strip().strip('"')
                        pub_date = parts[6].strip().strip('"') if len(parts) > 6 else ""
                        pub_end = parts[7].strip().strip('"')

                        if not isin.startswith("FR") or isin in known_isins or pub_end:
                            continue

                        try:
                            ratio = float(ratio_str)
                        except ValueError:
                            continue

                        key = (holder.lower().strip(), isin)
                        if key not in amf_active or pub_date > amf_active[key].get("date", ""):
                            amf_active[key] = {
                                "holder": holder, "isin": isin,
                                "issuer": issuer, "ratio": ratio, "date": pub_date,
                            }

                # Build discoveries from deduped entries
                for key, data in amf_active.items():
                    isin = data["isin"]
                    rec = discoveries_amf[isin]
                    rec["isin"] = isin
                    rec["issuer"] = data["issuer"] or rec["issuer"]
                    if data["holder"] not in rec["funds"]:
                        rec["funds"].append(data["holder"])
                    rec["total_pct"] += data["ratio"]
            except Exception:
                pass

        discoveries_amf = {k: v for k, v in discoveries_amf.items() if len(v["funds"]) >= min_funds}

    # --- Output ---
    print(f"=== DISCOVERY ENGINE (min {min_funds} funds) ===\n")

    if discoveries_13f:
        items = sorted(discoveries_13f.values(), key=lambda x: -len(x["funds"]))
        print(f"  STOCKS HELD BY {min_funds}+ TRACKED FUNDS (not in graph): {len(items)}")
        for item in items[:15]:
            val_str = f"${item['total_value']:,.0f}" if item['total_value'] else ""
            print(f"    CUSIP {item['cusip']}  {item['issuer'][:30]:30s}  {len(item['funds'])} funds  {val_str}")
            if len(item["funds"]) <= 4:
                print(f"      Funds: {', '.join(item['funds'])}")
    else:
        print(f"  13F Discovery: No stocks held by {min_funds}+ funds outside our graph.")

    if discoveries_fca:
        items = sorted(discoveries_fca.values(), key=lambda x: -x["total_pct"])
        print(f"\n  HEAVILY SHORTED UK STOCKS (not tracked): {len(items)}")
        for item in items[:15]:
            print(f"    ISIN {item['isin']}  {item['issuer'][:30]:30s}  {len(item['funds'])} funds  SI {item['total_pct']:.1f}%")
    else:
        if source in ("fca", "all"):
            print(f"\n  FCA Discovery: No UK stocks shorted by {min_funds}+ funds outside our graph.")

    if discoveries_amf:
        items = sorted(discoveries_amf.values(), key=lambda x: -x["total_pct"])
        print(f"\n  HEAVILY SHORTED FR STOCKS (not tracked): {len(items)}")
        for item in items[:15]:
            print(f"    ISIN {item['isin']}  {item['issuer'][:30]:30s}  {len(item['funds'])} funds  SI {item['total_pct']:.1f}%")
    else:
        if source in ("amf", "all"):
            print(f"\n  AMF Discovery: No FR stocks shorted by {min_funds}+ funds outside our graph.")

    total = len(discoveries_13f) + len(discoveries_fca) + len(discoveries_amf)
    print(f"\n  Total discoveries: {total}")
    print()


# ---------------------------------------------------------------------------
# Helper: append fund to tracked_funds.yaml
# ---------------------------------------------------------------------------

def _append_tracked_fund(name, cik, fund_type="discovered", notes=""):
    """Append fund to tracked_funds.yaml if not already present. Returns True if added."""
    if not os.path.exists(TRACKED_FUNDS_PATH):
        data = {"funds": []}
    else:
        with open(TRACKED_FUNDS_PATH) as f:
            data = yaml.safe_load(f) or {"funds": []}

    funds = data.get("funds", [])

    # Check for duplicates by CIK or name
    cik_str = str(cik) if cik else ""
    for existing in funds:
        if cik_str and str(existing.get("cik", "")) == cik_str:
            return False
        if existing.get("name", "").lower() == name.lower():
            return False

    new_fund = {"name": name, "fund_type": fund_type}
    if cik_str:
        new_fund["cik"] = cik_str
    if notes:
        new_fund["notes"] = notes

    funds.append(new_fund)
    data["funds"] = funds

    with open(TRACKED_FUNDS_PATH, "w") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True)

    return True


# ---------------------------------------------------------------------------
# CMD: discover-funds (v4.0 — Fund Discovery)
# ---------------------------------------------------------------------------

def cmd_discover_funds(args):
    """Analyze graph to discover fund entities worth tracking.

    Reverse-engineers the graph: which fund nodes appear on 3+ of our stocks?
    Cross-references FCA/AMF short data with 13F holding data.
    Outputs ranked list of candidate funds with YAML snippet for tracked_funds.yaml.
    """
    G = load_graph()
    fund_nodes = nodes_by_type(G, "fund")
    stock_nodes = nodes_by_type(G, "stock")

    # Load currently tracked funds
    tracked_ciks = set()
    tracked_names = set()
    if os.path.exists(TRACKED_FUNDS_PATH):
        with open(TRACKED_FUNDS_PATH) as f:
            tf = yaml.safe_load(f) or {}
        for fund in tf.get("funds", []):
            cik = str(fund.get("cik", ""))
            if cik:
                tracked_ciks.add(cik)
            name = fund.get("name", "").lower()
            if name:
                tracked_names.add(name)

    # Analyze: for each fund node, count how many of our stocks it's connected to
    fund_coverage = {}
    for fund_slug, fund_data in fund_nodes.items():
        # Check if already tracked
        full_name = fund_data.get("full_name", fund_slug)
        if full_name.lower() in tracked_names:
            continue

        stocks_connected = set()
        short_stocks = []
        hold_stocks = []

        for _, target, _, edge_data in G.out_edges(fund_slug, data=True, keys=True):
            if target in stock_nodes:
                stocks_connected.add(target)
                rel = edge_data.get("relation", "")
                if rel == "shorts":
                    short_stocks.append((target, edge_data.get("pct_shares", 0)))
                elif rel == "holds":
                    hold_stocks.append((target, edge_data.get("value", 0)))

        if len(stocks_connected) < 3:
            continue

        # Classify fund type
        if short_stocks and hold_stocks:
            fund_type = "multi-strategy"
        elif short_stocks and not hold_stocks:
            fund_type = "hedge-fund"
        elif hold_stocks and not short_stocks:
            fund_type = "long-only"
        else:
            fund_type = "unknown"

        # Check for CIK in fund data
        cik = fund_data.get("cik", "")

        fund_coverage[fund_slug] = {
            "full_name": full_name,
            "stocks": len(stocks_connected),
            "short_stocks": short_stocks,
            "hold_stocks": hold_stocks,
            "fund_type": fund_type,
            "cik": cik,
            "data_sources": set(),
        }

        # Track data sources
        for _, _, _, edge_data in G.out_edges(fund_slug, data=True, keys=True):
            ds = edge_data.get("data_source", "")
            if ds:
                fund_coverage[fund_slug]["data_sources"].add(ds)

    # Sort by stock count descending
    ranked = sorted(fund_coverage.items(), key=lambda x: -x[1]["stocks"])

    min_stocks = getattr(args, "min_stocks", 3) or 3
    ranked = [(slug, info) for slug, info in ranked if info["stocks"] >= min_stocks]

    print(f"=== FUND DISCOVERY (min {min_stocks} stocks connected) ===\n")

    if not ranked:
        print("  No untracked funds found with sufficient stock connections.")
        return

    print(f"  Found {len(ranked)} candidate funds:\n")
    print(f"  {'Fund':40s} {'Stocks':>7s} {'Type':15s} {'Sources':20s} {'CIK':>10s}")
    print(f"  {'-'*40} {'-'*7} {'-'*15} {'-'*20} {'-'*10}")

    for slug, info in ranked[:25]:
        name = info["full_name"][:40]
        sources_str = ", ".join(sorted(info["data_sources"]))
        cik_str = info["cik"] or "-"
        print(f"  {name:40s} {info['stocks']:>7d} {info['fund_type']:15s} {sources_str:20s} {cik_str:>10s}")

        # Show connected stocks
        shorts = [f"{t}({p:.1f}%)" for t, p in info["short_stocks"]]
        holds = [t for t, _ in info["hold_stocks"]]
        if shorts:
            print(f"    Shorts: {', '.join(shorts[:5])}")
        if holds:
            print(f"    Holds:  {', '.join(holds[:5])}")

    # Auto-enroll or show YAML snippet
    auto_enroll = getattr(args, "auto_enroll", False)

    if auto_enroll:
        # Filter to funds with CIK (required for 13F download)
        enrollable = [(slug, info) for slug, info in ranked if info["cik"]]
        if not enrollable:
            print(f"\n  No enrollable funds (all lack CIK). Use YAML snippet for regulatory-only funds.")
        else:
            max_enroll = 5
            enrolled_ciks = []
            print(f"\n  --- AUTO-ENROLLING (max {max_enroll}, CIK required) ---")
            for slug, info in enrollable[:max_enroll]:
                notes = (f"Auto-discovered {today_str()}. "
                         f"{info['stocks']} stocks, sources: {', '.join(sorted(info['data_sources']))}")
                added = _append_tracked_fund(
                    name=info["full_name"],
                    cik=info["cik"],
                    fund_type=info["fund_type"],
                    notes=notes,
                )
                if added:
                    print(f"  + ENROLLED: {info['full_name']} (CIK {info['cik']}, {info['fund_type']})")
                    enrolled_ciks.append(info["cik"])
                else:
                    print(f"  = SKIPPED (already tracked): {info['full_name']}")

            # Download 13F for newly enrolled CIKs and ingest
            if enrolled_ciks:
                import time
                print(f"\n  Downloading 13F for {len(enrolled_ciks)} new funds...")
                for cik in enrolled_ciks:
                    print(f"    Fetching 13F for CIK {cik}...")
                    download_13f_for_cik(str(cik), f"CIK-{cik}")
                    time.sleep(0.5)

                print(f"  Ingesting 13F holdings for new funds...")
                cmd_ingest_13f(argparse.Namespace())
                print(f"\n  Auto-enroll complete. {len(enrolled_ciks)} funds added + ingested.")
            else:
                print(f"\n  No new funds enrolled (all already tracked).")

        # Also show remaining non-CIK funds
        no_cik = [(slug, info) for slug, info in ranked if not info["cik"]]
        if no_cik:
            print(f"\n  Regulatory-only funds (no CIK, cannot auto-enroll):")
            for slug, info in no_cik[:5]:
                print(f"    {info['full_name']} — {info['stocks']} stocks, {info['fund_type']}")
    else:
        # Show YAML snippet (original behavior)
        print(f"\n  --- YAML snippet for tracked_funds.yaml ---")
        print(f"  # Add these to the 'funds' list in {TRACKED_FUNDS_PATH}")
        for slug, info in ranked[:10]:
            if info["cik"]:
                print(f"  - name: \"{info['full_name']}\"")
                print(f"    cik: \"{info['cik']}\"")
                print(f"    type: {info['fund_type']}")
                print(f"    # Connected to {info['stocks']} stocks, sources: {', '.join(sorted(info['data_sources']))}")
            else:
                print(f"  # {info['full_name']} — {info['stocks']} stocks, {info['fund_type']} (no CIK — regulatory data only)")
        print(f"\n  TIP: Use --auto-enroll to automatically add funds with CIK and download their 13F.")

    print()


# ---------------------------------------------------------------------------
# CMD: ingest-live (WS4 — Live Intelligence Ingest)
# ---------------------------------------------------------------------------

def cmd_ingest_live(args):
    """Ingest live intelligence discovered during sessions.

    Creates/updates nodes and edges from ad-hoc intelligence:
    - holder: Fund took a position (e.g., "Elliott took 5% stake in LULU")
    - short: Short position info (from news/filings)
    - insider: Insider transaction (from earnings, news)
    - mention: Fund letter mention or commentary

    Tagged with source="live_session" for lower confidence than regulatory data.
    """
    G = load_graph()

    live_type = args.type
    fund = args.fund
    ticker = args.ticker
    data_str = args.data or "{}"

    try:
        data = json.loads(data_str)
    except json.JSONDecodeError:
        print(f"Error: invalid JSON data: {data_str}")
        sys.exit(1)

    fund_slug = fund.lower().replace(" ", "-").replace(",", "").replace(".", "")[:40] if fund else ""

    # Ensure stock node
    if ticker not in G.nodes:
        G.add_node(ticker, type="stock", country=ticker_country(ticker))

    if live_type == "holder":
        if not fund_slug:
            print("Error: --fund required for holder type")
            sys.exit(1)
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=fund, fund_type="unknown")

        action = data.get("action", "hold")
        if action == "exit":
            # Remove existing hold edges
            keys_to_remove = []
            if G.has_node(fund_slug):
                for k, d in G[fund_slug].get(ticker, {}).items():
                    if d.get("relation") == "holds":
                        keys_to_remove.append(k)
                for k in keys_to_remove:
                    G.remove_edge(fund_slug, ticker, key=k)
            print(f"Removed hold edge: {fund} -> {ticker} (exit)")
        else:
            pct = data.get("pct", 0)
            G.add_edge(fund_slug, ticker, relation="holds",
                       pct_portfolio=pct, source="live_session",
                       session_date=today_str(), date_added=today_str(),
                       note=data.get("source", ""))
            print(f"Added hold edge: {fund} -> {ticker} ({pct}%)")

    elif live_type == "short":
        if not fund_slug:
            print("Error: --fund required for short type")
            sys.exit(1)
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=fund, fund_type="unknown")

        pct = data.get("pct", 0)
        G.add_edge(fund_slug, ticker, relation="shorts",
                   pct_shares=float(pct), source="live_session",
                   session_date=today_str(), date_added=today_str(),
                   data_source="live", note=data.get("source", ""))
        print(f"Added short edge: {fund} -> {ticker} ({pct}%)")

    elif live_type == "insider":
        name = data.get("name", "unknown")
        person_slug = name.lower().replace(" ", "-").replace(",", "").replace(".", "")[:30]
        action = data.get("action", "buy")
        value = data.get("value", 0)
        role = data.get("role", "")

        if person_slug not in G.nodes:
            G.add_node(person_slug, type="person", full_name=name, role=role)

        relation = "insider_buy" if "buy" in action.lower() else "insider_sell"
        G.add_edge(person_slug, ticker, relation=relation,
                   value=value, date=data.get("date", today_str()),
                   role=role, source="live_session",
                   data_source="live", date_added=today_str())
        print(f"Added {relation} edge: {name} -> {ticker} (${value:,.0f})")

    elif live_type == "mention":
        if not fund_slug:
            print("Error: --fund required for mention type")
            sys.exit(1)
        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=fund, fund_type="unknown")

        G.add_edge(fund_slug, ticker, relation="mentions",
                   sentiment=data.get("sentiment", "neutral"),
                   context=data.get("context", ""),
                   source="live_session",
                   session_date=today_str(), date_added=today_str())
        print(f"Added mention edge: {fund} -> {ticker}")

    else:
        print(f"Unknown type: {live_type}. Use: holder, short, insider, mention")
        sys.exit(1)

    save_graph(G)


# ---------------------------------------------------------------------------
# CMD: capture (v4.1 — Quick-Capture Session Intelligence)
# ---------------------------------------------------------------------------

import re as _re_module

CAPTURE_PATTERNS = [
    # "Elliott holds 5.2% LULU" / "Elliott hold 5.2% LULU"
    (_re_module.compile(r"^(.+?)\s+holds?\s+([\d.]+)%\s+(\S+)$", _re_module.IGNORECASE), "holder_hold"),
    # "Citadel shorts 2.1% EDEN.PA" / "Citadel short 2.1% EDEN.PA"
    (_re_module.compile(r"^(.+?)\s+shorts?\s+([\d.]+)%\s+(\S+)$", _re_module.IGNORECASE), "short"),
    # "Baupost exited GL" / "Baupost exit GL"
    (_re_module.compile(r"^(.+?)\s+exit(?:ed|s)?\s+(\S+)$", _re_module.IGNORECASE), "holder_exit"),
    # "CEO Narayen bought 2.5M ADBE" / "CFO Smith bought $500K LULU"
    (_re_module.compile(r"^(\w+)\s+(.+?)\s+bought?\s+\$?([\d.]+[MmKk]?)\s+(\S+)$", _re_module.IGNORECASE), "insider_buy"),
    # "Fundsmith bullish ADBE" / "Einhorn bearish TSLA"
    (_re_module.compile(r"^(.+?)\s+(bullish|bearish|neutral)\s+(\S+)$", _re_module.IGNORECASE), "mention"),
]


def _parse_value_shorthand(val_str):
    """Parse value like '2.5M', '500K', '2500000' → float."""
    val_str = val_str.strip().lstrip("$")
    multiplier = 1
    if val_str.upper().endswith("M"):
        multiplier = 1_000_000
        val_str = val_str[:-1]
    elif val_str.upper().endswith("K"):
        multiplier = 1_000
        val_str = val_str[:-1]
    try:
        return float(val_str) * multiplier
    except ValueError:
        return 0


def cmd_capture(args):
    """Quick-capture session intelligence with natural-ish syntax.

    Parses common patterns and delegates to ingest-live internally.
    Examples:
        capture Elliott holds 5.2% LULU
        capture Citadel shorts 2.1% EDEN.PA
        capture Baupost exited GL
        capture CEO Narayen bought 2.5M ADBE
        capture Fundsmith bullish ADBE
    """
    text = " ".join(args.text).strip()
    if not text:
        print("Error: provide text to capture.")
        print("Examples:")
        print('  capture Elliott holds 5.2% LULU')
        print('  capture Citadel shorts 2.1% EDEN.PA')
        print('  capture Baupost exited GL')
        print('  capture CEO Narayen bought 2.5M ADBE')
        print('  capture Fundsmith bullish ADBE')
        return

    matched = False
    for pattern, ptype in CAPTURE_PATTERNS:
        m = pattern.match(text)
        if not m:
            continue

        if ptype == "holder_hold":
            fund, pct, ticker = m.group(1).strip(), m.group(2), m.group(3).upper()
            ns = argparse.Namespace(
                type="holder", fund=fund, ticker=ticker,
                data=json.dumps({"action": "hold", "pct": float(pct), "source": "capture"})
            )
            print(f"  → ingest-live: {fund} holds {pct}% {ticker}")
            cmd_ingest_live(ns)
            matched = True
            break

        elif ptype == "short":
            fund, pct, ticker = m.group(1).strip(), m.group(2), m.group(3).upper()
            ns = argparse.Namespace(
                type="short", fund=fund, ticker=ticker,
                data=json.dumps({"pct": float(pct), "source": "capture"})
            )
            print(f"  → ingest-live: {fund} shorts {pct}% {ticker}")
            cmd_ingest_live(ns)
            matched = True
            break

        elif ptype == "holder_exit":
            fund, ticker = m.group(1).strip(), m.group(2).upper()
            ns = argparse.Namespace(
                type="holder", fund=fund, ticker=ticker,
                data=json.dumps({"action": "exit", "source": "capture"})
            )
            print(f"  → ingest-live: {fund} exited {ticker}")
            cmd_ingest_live(ns)
            matched = True
            break

        elif ptype == "insider_buy":
            role, name, val_str, ticker = m.group(1), m.group(2).strip(), m.group(3), m.group(4).upper()
            value = _parse_value_shorthand(val_str)
            ns = argparse.Namespace(
                type="insider", fund=None, ticker=ticker,
                data=json.dumps({
                    "name": name, "action": "buy", "value": value,
                    "role": role, "date": today_str(), "source": "capture"
                })
            )
            print(f"  → ingest-live: {role} {name} bought ${value:,.0f} {ticker}")
            cmd_ingest_live(ns)
            matched = True
            break

        elif ptype == "mention":
            fund, sentiment, ticker = m.group(1).strip(), m.group(2).lower(), m.group(3).upper()
            ns = argparse.Namespace(
                type="mention", fund=fund, ticker=ticker,
                data=json.dumps({"sentiment": sentiment, "source": "capture"})
            )
            print(f"  → ingest-live: {fund} {sentiment} {ticker}")
            cmd_ingest_live(ns)
            matched = True
            break

    if not matched:
        print(f"Could not parse: \"{text}\"")
        print("Supported patterns:")
        print("  {fund} holds {pct}% {ticker}")
        print("  {fund} shorts {pct}% {ticker}")
        print("  {fund} exited {ticker}")
        print("  {role} {name} bought {$value} {ticker}")
        print("  {fund} bullish|bearish|neutral {ticker}")
        print("\nFor complex inputs, use ingest-live directly.")


# ---------------------------------------------------------------------------
# CMD: sector-overlay (WS5 — Sector View Integration)
# ---------------------------------------------------------------------------

def cmd_sector_overlay(args):
    """Generate institutional positioning overlay for a sector.

    Aggregates holders, shorts, insider activity, and signals for all
    stocks in a given sector. Outputs markdown ready to paste into sector views.
    """
    G = load_graph()
    target_sector = args.sector.lower().replace("-", " ").replace("_", " ")

    # Load quality universe for sector mapping
    universe_path = os.path.join(PROJECT_ROOT, "state", "quality_universe.yaml")
    sector_map = {}  # ticker -> sector
    if os.path.exists(universe_path):
        with open(universe_path) as f:
            udata = yaml.safe_load(f) or {}
        for entry in udata.get("quality_universe", {}).get("companies", []):
            ticker = entry.get("ticker", "")
            sector = entry.get("sector", "")
            if ticker and sector:
                sector_map[ticker] = sector

    # Find matching stocks
    stocks = nodes_by_type(G, "stock")
    matching = {}
    for ticker, data in stocks.items():
        stock_sector = (sector_map.get(ticker, "") or data.get("sector", "")).lower().replace("-", " ").replace("_", " ")
        if target_sector in stock_sector or stock_sector in target_sector:
            matching[ticker] = data

    if not matching:
        print(f"No stocks found for sector '{args.sector}' in graph.")
        print(f"  Available sectors: {', '.join(sorted(set(sector_map.values())))[:200]}")
        return

    # Aggregate data per stock
    rows = []
    total_si = 0
    total_holder_count = 0
    total_insider_buys = 0
    sector_quality_funds = set()

    for ticker in sorted(matching.keys()):
        holders = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds"]
        shorts = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "shorts"]
        insider_buys = [(u, d) for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "insider_buy"]

        si = sum(s[1].get("pct_shares", 0) or 0 for s in shorts)
        total_si += si
        total_holder_count += len(holders)
        total_insider_buys += len(insider_buys)

        for h_id, _ in holders:
            ft = G.nodes.get(h_id, {}).get("fund_type", "")
            if ft in ("value", "quality", "activist"):
                sector_quality_funds.add(h_id)

        # Simple signal detection for this stock
        signal = ""
        if si > 8 and len(shorts) >= 5:
            signal = "SHORT_ESCALATION"
        elif len(holders) >= 2 and any(G.nodes.get(h[0], {}).get("fund_type") in ("value", "quality") for h in holders):
            signal = "CONVERGENCE"
        elif len([b for b in insider_buys if days_since(b[1].get("date")) is not None and days_since(b[1].get("date")) <= 60]) >= 3:
            signal = "INSIDER_CLUSTER"

        in_pf = matching[ticker].get("in_portfolio", False)
        rows.append({
            "ticker": ticker,
            "si": si,
            "n_holders": len(holders),
            "n_insiders": len(insider_buys),
            "signal": signal,
            "in_portfolio": in_pf,
        })

    n_stocks = len(matching)
    avg_si = total_si / n_stocks if n_stocks > 0 else 0

    # Determine net sentiment
    if avg_si > 5 and len(sector_quality_funds) >= 2:
        sentiment = "MIXED (high shorts + quality accumulation = controversy)"
    elif avg_si > 5:
        sentiment = "BEARISH (elevated short interest)"
    elif len(sector_quality_funds) >= 3:
        sentiment = "BULLISH (quality fund accumulation)"
    else:
        sentiment = "NEUTRAL (limited data)"

    # Output markdown
    print(f"## Institutional Positioning — {args.sector} (auto-generated {today_str()})\n")
    print(f"**Sector avg short interest:** {avg_si:.1f}%")
    print(f"**Quality fund exposure:** {len(sector_quality_funds)} funds across {n_stocks} stocks")
    print(f"**Net sentiment:** {sentiment}\n")

    print(f"| Stock | SI% | Holders | Insiders | Signal | Portfolio |")
    print(f"|-------|-----|---------|----------|--------|-----------|")
    for r in sorted(rows, key=lambda x: -x["si"]):
        pf = "YES" if r["in_portfolio"] else ""
        print(f"| {r['ticker']} | {r['si']:.1f}% | {r['n_holders']} | {r['n_insiders']} | {r['signal']} | {pf} |")

    print()


# ---------------------------------------------------------------------------
# CMD: ingest-consob (WS1 — CONSOB Italy)
# ---------------------------------------------------------------------------

def cmd_ingest_consob(args):
    """Parse CONSOB XLSX and apply short positions to graph for Italian stocks.

    CONSOB publishes net short positions in XLSX format similar to FCA.
    Download from: https://www.consob.it/web/area-pubblica/pnc
    Save to tools/smart_money/data/shorts/consob_italy_YYYY-MM-DD.xlsx
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required. Install: pip install openpyxl")
        return

    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})

    # Build ISIN → ticker map for IT stocks
    isin_to_ticker = {}
    for ticker, info in ident.items():
        isin = info.get("isin")
        if isin and (isin.startswith("IT") or ticker.endswith(".MI")):
            isin_to_ticker[isin] = ticker

    if not isin_to_ticker:
        print("No Italian ISINs in identifiers.yaml. Run 'resolve' for .MI tickers first.")
        return

    # Find most recent CONSOB file
    if args.file:
        fpath = args.file
    else:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("consob_italy_") and f.endswith(".xlsx")], reverse=True)
        if not files:
            print("No CONSOB XLSX files found.")
            print("  Download from: https://www.consob.it/web/area-pubblica/pnc")
            print("  Save as: tools/smart_money/data/shorts/consob_italy_YYYY-MM-DD.xlsx")
            return
        fpath = os.path.join(SHORTS_DIR, files[0])

    print(f"Ingesting CONSOB: {os.path.basename(fpath)}", file=sys.stderr)

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Empty spreadsheet.")
        return

    # Find header row (CONSOB: "Soggetto", "Emittente", "ISIN", "Posizione Netta Corta (%)")
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = str(row).lower()
        if any(w in row_str for w in ["soggetto", "position holder", "emittente", "issuer"]):
            header_idx = i
            break

    headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
    holder_col = next((i for i, h in enumerate(headers) if any(w in h for w in ["soggetto", "position holder", "holder"])), 0)
    isin_col = next((i for i, h in enumerate(headers) if "isin" in h), None)
    pct_col = next((i for i, h in enumerate(headers) if any(w in h for w in ["posizione netta", "net short", "position (%)"])), None)
    date_col = next((i for i, h in enumerate(headers) if "data" in h or "date" in h), None)

    if isin_col is None:
        print("Could not find ISIN column in CONSOB file. Check format.")
        return

    G = load_graph()
    fund_aliases = load_fund_aliases()

    # Remove existing CONSOB-sourced edges
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == "shorts" and d.get("data_source") == "consob":
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    matched = 0
    total_rows = 0
    funds_seen = set()

    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        total_rows += 1

        isin = str(row[isin_col]).strip() if row[isin_col] else ""
        if isin not in isin_to_ticker:
            continue

        ticker = isin_to_ticker[isin]
        holder = str(row[holder_col]).strip() if row[holder_col] else ""
        pct = 0
        if pct_col is not None and row[pct_col]:
            try:
                pct = float(str(row[pct_col]).replace("%", "").replace(",", ".").strip())
            except ValueError:
                pct = 0

        pos_date = ""
        if date_col is not None and row[date_col]:
            if isinstance(row[date_col], datetime):
                pos_date = row[date_col].strftime("%Y-%m-%d")
            else:
                pos_date = str(row[date_col])[:10]

        fund_slug = holder.lower().replace(" ", "-").replace(",", "").replace(".", "")[:40]
        fund_slug = canonicalize_fund_slug(fund_slug, fund_aliases)
        funds_seen.add(fund_slug)

        if fund_slug not in G.nodes:
            G.add_node(fund_slug, type="fund", full_name=holder, fund_type="unknown")
        if ticker not in G.nodes:
            G.add_node(ticker, type="stock", country="IT")

        G.add_edge(fund_slug, ticker, relation="shorts", pct_shares=float(pct),
                   date=pos_date, data_source="consob", date_added=today_str())
        matched += 1

    save_graph(G)
    print(f"CONSOB ingest: {matched} short positions across {len(funds_seen)} funds "
          f"(from {total_rows} rows, {len(isin_to_ticker)} ISINs tracked)")


# ---------------------------------------------------------------------------
# CMD: ingest-afm-nl (WS1 — AFM Netherlands)
# ---------------------------------------------------------------------------

def cmd_ingest_afm_nl(args):
    """Parse AFM Netherlands CSV and apply short positions to graph for Dutch stocks.

    AFM publishes net short positions in CSV format.
    Download from: https://www.afm.nl/en/sector/registers/meldingenregisters/netto-shortposities-actueel
    Save to tools/smart_money/data/shorts/afm_nl_YYYY-MM-DD.csv
    """
    identifiers_data = load_identifiers()
    ident = identifiers_data.get("identifiers", {})

    # Build ISIN → ticker map for NL stocks
    isin_to_ticker = {}
    for ticker, info in ident.items():
        isin = info.get("isin")
        if isin and (isin.startswith("NL") or ticker.endswith(".AS")):
            isin_to_ticker[isin] = ticker

    if not isin_to_ticker:
        print("No Dutch ISINs in identifiers.yaml. Run 'resolve' for .AS tickers first.")
        return

    # Find most recent AFM NL file
    if args.file:
        fpath = args.file
    else:
        files = sorted([f for f in os.listdir(SHORTS_DIR)
                        if f.startswith("afm_nl_") and f.endswith(".csv")], reverse=True)
        if not files:
            print("No AFM NL CSV files found.")
            print("  Download from: https://www.afm.nl/en/sector/registers/meldingenregisters/netto-shortposities-actueel")
            print("  Save as: tools/smart_money/data/shorts/afm_nl_YYYY-MM-DD.csv")
            return
        fpath = os.path.join(SHORTS_DIR, files[0])

    print(f"Ingesting AFM NL: {os.path.basename(fpath)}", file=sys.stderr)

    # Parse CSV — AFM NL format: semicolon or comma delimited
    # Try to detect format from header
    G = load_graph()
    fund_aliases = load_fund_aliases()

    # Remove existing AFM-NL-sourced edges
    edges_to_remove = []
    for u, v, k, d in G.edges(data=True, keys=True):
        if d.get("relation") == "shorts" and d.get("data_source") == "afm_nl":
            edges_to_remove.append((u, v, k))
    for u, v, k in edges_to_remove:
        G.remove_edge(u, v, key=k)

    matched = 0
    total_rows = 0
    funds_seen = set()

    with open(fpath, "r", encoding="utf-8", errors="replace") as f:
        header_line = f.readline()
        # Detect delimiter
        delimiter = ";" if ";" in header_line else ","
        header_lower = header_line.lower()

        for line in f:
            total_rows += 1
            parts = line.strip().split(delimiter)
            if len(parts) < 4:
                continue

            # AFM NL typical columns: Position holder, Issuer, ISIN, Net short position (%), Date
            holder = parts[0].strip().strip('"')
            issuer = parts[1].strip().strip('"') if len(parts) > 1 else ""
            isin = ""
            pct = 0
            pos_date = ""

            # Find ISIN column (starts with NL or other country code, length 12)
            for part in parts:
                clean = part.strip().strip('"')
                if len(clean) == 12 and clean[:2].isalpha() and clean[2:].replace(" ", "").isalnum():
                    isin = clean
                    break

            if isin not in isin_to_ticker:
                continue

            ticker = isin_to_ticker[isin]

            # Find percentage (contains . and is small number)
            for part in parts:
                clean = part.strip().strip('"').replace(",", ".")
                try:
                    val = float(clean)
                    if 0 < val < 100:
                        pct = val
                        break
                except ValueError:
                    pass

            fund_slug = holder.lower().replace(" ", "-").replace(",", "").replace(".", "")[:40]
            fund_slug = canonicalize_fund_slug(fund_slug, fund_aliases)
            funds_seen.add(fund_slug)

            if fund_slug not in G.nodes:
                G.add_node(fund_slug, type="fund", full_name=holder, fund_type="unknown")
            if ticker not in G.nodes:
                G.add_node(ticker, type="stock", country="NL")

            G.add_edge(fund_slug, ticker, relation="shorts", pct_shares=float(pct),
                       date=pos_date, data_source="afm_nl", date_added=today_str())
            matched += 1

    save_graph(G)
    print(f"AFM NL ingest: {matched} short positions across {len(funds_seen)} funds "
          f"(from {total_rows} rows, {len(isin_to_ticker)} ISINs tracked)")



# ---------------------------------------------------------------------------
# CMD: weekly-report
# ---------------------------------------------------------------------------

REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports", "smart_money")
UNIVERSE_PATH = os.path.join(PROJECT_ROOT, "state", "quality_universe.yaml")


def _get_universe_tickers():
    """Load ticker set from quality_universe.yaml."""
    if not os.path.exists(UNIVERSE_PATH):
        return set()
    with open(UNIVERSE_PATH) as f:
        data = yaml.safe_load(f) or {}
    return {c["ticker"] for c in data.get("quality_universe", {}).get("companies", []) if c.get("ticker")}


def _gather_signals(G, portfolio_tickers):
    """Reuse signal detection logic from cmd_signals. Returns list of (type, ticker, weight, detail)."""
    stocks = nodes_by_type(G, "stock")
    # Load previous snapshot
    G_prev = None
    if os.path.exists(SNAPSHOTS_DIR):
        snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
        for s in reversed(snaps):
            snap_date = s[6:16]
            if snap_date != today_str():
                try:
                    with open(os.path.join(SNAPSHOTS_DIR, s)) as f2:
                        G_prev = nx.node_link_graph(json.load(f2), directed=True, multigraph=True)
                except Exception:
                    pass
                break

    median = _median_holder_count(G)
    signals = []

    for ticker, sdata in stocks.items():
        holders = []
        shorts = []
        insider_buys = []

        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            if rel == "holds":
                holders.append((u, d))
            elif rel == "shorts":
                shorts.append((u, d))
            elif rel == "insider_buy":
                insider_buys.append((u, d))

        total_si = sum(s[1].get("pct_shares", 0) or 0 for s in shorts)
        n_short_funds = len(shorts)
        in_pf = ticker in portfolio_tickers

        quality_holders = []
        for fund_id, ed in holders:
            ft = G.nodes.get(fund_id, {}).get("fund_type", "")
            if ft in ("value", "quality", "activist"):
                quality_holders.append((fund_id, ed))

        recent_buys_60d = []
        recent_buys_90d = []
        for person_id, ed in insider_buys:
            buy_date = ed.get("date", "")
            days = days_since(buy_date)
            if days is not None:
                if days <= 60:
                    recent_buys_60d.append((person_id, ed))
                if days <= 90:
                    recent_buys_90d.append((person_id, ed))

        if len(recent_buys_60d) >= 3:
            total_val = sum(b[1].get("value", 0) or 0 for b in recent_buys_60d)
            signals.append(("INSIDER_CLUSTER_BUY", ticker, "STRONG BULL",
                            f"{len(recent_buys_60d)} insiders bought in 60d (${total_val:,.0f})"))

        if len(quality_holders) >= 2:
            names = [G.nodes.get(h[0], {}).get("full_name", h[0])[:20] for h in quality_holders[:3]]
            signals.append(("CONVERGENCE", ticker, "BULL",
                            f"{len(quality_holders)} quality/value funds: {', '.join(names)}"))

        if total_si > 10 and quality_holders and recent_buys_90d:
            signals.append(("SHORT_SQUEEZE_RISK", ticker, "CONTROVERSY",
                            f"SI {total_si:.1f}% + {len(quality_holders)} quality holders + insider buying"))

        if G_prev and ticker in G_prev.nodes:
            prev_holders = {u for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True)
                           if d.get("relation") == "holds"}
            curr_holders = {h[0] for h in holders}
            exited = prev_holders - curr_holders
            for ex_fund in exited:
                ft = G_prev.nodes.get(ex_fund, {}).get("fund_type", "")
                if ft in ("value", "quality", "activist"):
                    name = G_prev.nodes.get(ex_fund, {}).get("full_name", ex_fund)
                    signals.append(("SMART_EXIT", ticker, "BEAR WARNING",
                                    f"{name} ({ft}) exited position"))

        if total_si > 8 and n_short_funds >= 5:
            signals.append(("SHORT_ESCALATION", ticker, "BEAR",
                            f"SI {total_si:.1f}% ({n_short_funds} funds)"))

        if holders and total_si < 3 and recent_buys_90d and not in_pf:
            new_holder = False
            if G_prev and ticker in G_prev.nodes:
                prev_h = {u for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True)
                          if d.get("relation") == "holds"}
                curr_h = {h[0] for h in holders}
                new_holder = bool(curr_h - prev_h)
            if new_holder or not G_prev:
                signals.append(("QUIET_ACCUMULATION", ticker, "STEALTH BULL",
                                f"New fund + low SI ({total_si:.1f}%) + insider buying"))

        crowding = len(holders) / median if median > 0 else 0
        if len(holders) >= 5 and crowding > 3.0:
            signals.append(("HERD_WARNING", ticker, "RISK",
                            f"{len(holders)} funds hold ({crowding:.1f}x median)"))

        for person_id, ed in recent_buys_90d:
            role = (ed.get("role", "") or "").lower()
            value = ed.get("value", 0) or 0
            if value >= 100000 and any(w in role for w in ["ceo", "chief executive", "founder", "chairman", "director"]):
                name = G.nodes.get(person_id, {}).get("full_name", person_id)
                signals.append(("FOUNDER_CONVICTION", ticker, "CONVICTION",
                                f"{name} ({ed.get('role', '')}) ${value:,.0f}"))
                break

    signals.sort(key=lambda x: (x[0], x[1]))
    return signals


def _gather_alerts(G, portfolio_tickers):
    """Reuse alert detection logic from cmd_alerts. Returns list of (type, ticker, msg) and snapshot name."""
    snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
    prev_snap = None
    for s in reversed(snaps):
        snap_date = s[6:16]
        if snap_date != today_str():
            prev_snap = s
            break

    if not prev_snap:
        return [], None

    with open(os.path.join(SNAPSHOTS_DIR, prev_snap)) as f:
        prev_data = json.load(f)
    G_prev = nx.node_link_graph(prev_data, directed=True, multigraph=True)

    alerts = []
    for ticker in portfolio_tickers:
        curr_shorts = {}
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "shorts":
                curr_shorts[u] = d.get("pct_shares", 0) or 0

        prev_shorts = {}
        if ticker in G_prev.nodes:
            for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True):
                if d.get("relation") == "shorts":
                    prev_shorts[u] = d.get("pct_shares", 0) or 0

        curr_total = sum(curr_shorts.values())
        prev_total = sum(prev_shorts.values())
        delta = curr_total - prev_total

        if delta > 0.3:
            alerts.append(("SHORT_INCREASE", ticker, f"short {prev_total:.2f}% -> {curr_total:.2f}% (+{delta:.2f}pp)"))
        elif delta < -0.3:
            alerts.append(("SHORT_DECREASE", ticker, f"short {prev_total:.2f}% -> {curr_total:.2f}% ({delta:.2f}pp) [bullish]"))

        curr_holders = set()
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "holds":
                curr_holders.add(u)

        prev_holders = set()
        if ticker in G_prev.nodes:
            for u, v, k, d in G_prev.in_edges(ticker, data=True, keys=True):
                if d.get("relation") == "holds":
                    prev_holders.add(u)

        new_holders = curr_holders - prev_holders
        exit_holders = prev_holders - curr_holders

        for h in new_holders:
            fund_name = G.nodes[h].get("full_name", h) if h in G.nodes else h
            alerts.append(("NEW_HOLDER", ticker, f"{fund_name} entered position"))
        for h in exit_holders:
            fund_name = G_prev.nodes[h].get("full_name", h) if h in G_prev.nodes else h
            alerts.append(("EXIT_HOLDER", ticker, f"{fund_name} exited position"))

        recent_buys = []
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "insider_buy":
                buy_date = d.get("date", "")
                days_ago = days_since(buy_date)
                if days_ago is not None and days_ago <= 30:
                    recent_buys.append((u, d))

        if len(recent_buys) >= 3:
            total_val = sum(b[1].get("value", 0) or 0 for b in recent_buys)
            alerts.append(("INSIDER_CLUSTER_BUY", ticker, f"{len(recent_buys)} insiders bought in 30d (${total_val:,.0f} total)"))

        if delta < -0.3 and len(recent_buys) >= 2:
            alerts.append(("CONVERGENCE", ticker, "insider buy + short decrease = strong bullish signal"))

    alerts.sort(key=lambda x: x[0])
    return alerts, prev_snap


def _gather_crowding(G, top_n=15):
    """Reuse crowding logic. Returns list of (ticker, name, holders, crowding, in_pf) and median."""
    stocks = nodes_by_type(G, "stock")
    median = _median_holder_count(G)
    scores = []
    for ticker, data in stocks.items():
        holders = sum(1 for u, v, k, d in G.in_edges(ticker, data=True, keys=True) if d.get("relation") == "holds")
        crowding = holders / median if median > 0 else 0
        in_pf = data.get("in_portfolio", False)
        scores.append((ticker, data.get("name", ""), holders, crowding, in_pf))
    scores.sort(key=lambda x: -x[3])
    return scores[:top_n], median


def _gather_discovery_untracked(G, universe_tickers, portfolio_tickers, min_funds=3):
    """Find stocks in graph held by 3+ quality funds NOT in our universe or portfolio."""
    stocks = nodes_by_type(G, "stock")
    known = universe_tickers | portfolio_tickers
    discoveries = []

    for ticker, sdata in stocks.items():
        if ticker in known:
            continue
        holders = []
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            if d.get("relation") == "holds":
                fund_type = G.nodes.get(u, {}).get("fund_type", "")
                if fund_type in ("value", "quality", "activist"):
                    fund_name = G.nodes.get(u, {}).get("full_name", u)
                    holders.append(fund_name)
        if len(holders) >= min_funds:
            name = sdata.get("name", "")
            discoveries.append((ticker, name, holders))

    discoveries.sort(key=lambda x: -len(x[2]))
    return discoveries


def _gather_insider_summary(G):
    """Summarize insider activity across graph. Returns (total_buys, total_sells, buy_value, sell_value, clusters)."""
    stocks = nodes_by_type(G, "stock")
    total_buys = 0
    total_sells = 0
    buy_value = 0
    sell_value = 0
    clusters = []  # (ticker, count, value)

    for ticker in stocks:
        buys_90d = []
        sells_90d = []
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            dt = d.get("date", "")
            days_ago = days_since(dt)
            if days_ago is not None and days_ago <= 90:
                val = d.get("value", 0) or 0
                if rel == "insider_buy":
                    buys_90d.append(val)
                elif rel == "insider_sell":
                    sells_90d.append(val)

        total_buys += len(buys_90d)
        total_sells += len(sells_90d)
        buy_value += sum(buys_90d)
        sell_value += sum(sells_90d)

        if len(buys_90d) >= 3:
            clusters.append((ticker, len(buys_90d), sum(buys_90d)))

    clusters.sort(key=lambda x: -x[2])
    return total_buys, total_sells, buy_value, sell_value, clusters


def _gather_changes_vs_snapshot(G):
    """Compare current graph vs last snapshot. Returns dict of changes."""
    snaps = sorted(f for f in os.listdir(SNAPSHOTS_DIR) if f.startswith("graph_") and f.endswith(".json"))
    prev_snap = None
    for s in reversed(snaps):
        snap_date = s[6:16]
        if snap_date != today_str():
            prev_snap = s
            break

    if not prev_snap:
        return None

    with open(os.path.join(SNAPSHOTS_DIR, prev_snap)) as f:
        G_prev = nx.node_link_graph(json.load(f), directed=True, multigraph=True)

    snap_date = prev_snap[6:16]

    # Node changes
    curr_stocks = set(n for n, d in G.nodes(data=True) if d.get("type") == "stock")
    prev_stocks = set(n for n, d in G_prev.nodes(data=True) if d.get("type") == "stock")
    curr_funds = set(n for n, d in G.nodes(data=True) if d.get("type") == "fund")
    prev_funds = set(n for n, d in G_prev.nodes(data=True) if d.get("type") == "fund")

    new_stocks = curr_stocks - prev_stocks
    removed_stocks = prev_stocks - curr_stocks
    new_funds = curr_funds - prev_funds

    # Edge count changes
    curr_edges = G.number_of_edges()
    prev_edges = G_prev.number_of_edges()

    # Relation-type edge counts
    curr_holds = len([1 for u, v, k, d in G.edges(data=True, keys=True) if d.get("relation") == "holds"])
    prev_holds = len([1 for u, v, k, d in G_prev.edges(data=True, keys=True) if d.get("relation") == "holds"])
    curr_shorts_e = len([1 for u, v, k, d in G.edges(data=True, keys=True) if d.get("relation") == "shorts"])
    prev_shorts_e = len([1 for u, v, k, d in G_prev.edges(data=True, keys=True) if d.get("relation") == "shorts"])
    curr_ins = len([1 for u, v, k, d in G.edges(data=True, keys=True) if d.get("relation") in ("insider_buy", "insider_sell")])
    prev_ins = len([1 for u, v, k, d in G_prev.edges(data=True, keys=True) if d.get("relation") in ("insider_buy", "insider_sell")])

    return {
        "snapshot_date": snap_date,
        "new_stocks": sorted(new_stocks),
        "removed_stocks": sorted(removed_stocks),
        "new_funds": sorted(new_funds),
        "edge_delta": curr_edges - prev_edges,
        "holds_delta": curr_holds - prev_holds,
        "shorts_delta": curr_shorts_e - prev_shorts_e,
        "insider_delta": curr_ins - prev_ins,
        "curr_nodes": G.number_of_nodes(),
        "curr_edges": curr_edges,
    }


def cmd_weekly_report(args):
    """Generate a comprehensive markdown report and save to reports/smart_money/YYYY-MM-DD.md."""
    os.makedirs(REPORTS_DIR, exist_ok=True)

    G = load_graph()
    if not G.number_of_nodes():
        print("Graph is empty. Run sync-portfolio and refresh first.")
        return

    portfolio_tickers = set(get_portfolio_tickers())
    universe_tickers = _get_universe_tickers()
    stocks = nodes_by_type(G, "stock")
    funds = nodes_by_type(G, "fund")

    # Gather all data
    signals = _gather_signals(G, portfolio_tickers)
    alerts, prev_snap = _gather_alerts(G, portfolio_tickers)
    crowding_data, median_holders = _gather_crowding(G, top_n=15)
    discoveries = _gather_discovery_untracked(G, universe_tickers, portfolio_tickers, min_funds=3)
    total_buys, total_sells, buy_value, sell_value, insider_clusters = _gather_insider_summary(G)
    changes = _gather_changes_vs_snapshot(G)

    # Separate portfolio signals
    pf_signals = [(t, tk, w, d) for t, tk, w, d in signals if tk in portfolio_tickers]
    pf_alerts = [(t, tk, m) for t, tk, m in alerts]

    report_date = today_str()
    lines = []
    lines.append(f"# Smart Money Weekly Report - {report_date}\n")
    lines.append(f"Graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges | "
                 f"{len(stocks)} stocks, {len(funds)} funds | "
                 f"Portfolio: {len(portfolio_tickers)} positions | "
                 f"Universe: {len(universe_tickers)} tracked\n")

    # --- Executive Summary ---
    lines.append("## Executive Summary\n")
    summary_bullets = []
    if signals:
        by_type = defaultdict(int)
        for s in signals:
            by_type[s[0]] += 1
        summary_bullets.append(f"- **{len(signals)} signals detected**: "
                               + ", ".join(f"{t} ({c})" for t, c in sorted(by_type.items())))
    else:
        summary_bullets.append("- No actionable signals detected (graph may need refresh)")

    if alerts:
        summary_bullets.append(f"- **{len(alerts)} portfolio alerts** vs last snapshot")
    else:
        summary_bullets.append("- No portfolio alerts vs last snapshot")

    bull_signals = [s for s in signals if s[2] in ("BULL", "STRONG BULL", "CONVICTION", "STEALTH BULL")]
    bear_signals = [s for s in signals if s[2] in ("BEAR", "BEAR WARNING")]
    if bull_signals:
        tickers = sorted(set(s[1] for s in bull_signals))
        summary_bullets.append(f"- **Bullish signals** on: {', '.join(tickers[:8])}")
    if bear_signals:
        tickers = sorted(set(s[1] for s in bear_signals))
        summary_bullets.append(f"- **Bearish signals** on: {', '.join(tickers[:8])}")

    if discoveries:
        summary_bullets.append(f"- **{len(discoveries)} untracked opportunities** held by 3+ quality funds")

    if insider_clusters:
        cluster_tickers = [c[0] for c in insider_clusters[:5]]
        summary_bullets.append(f"- **Insider cluster buys** (90d): {', '.join(cluster_tickers)}")

    for b in summary_bullets[:5]:
        lines.append(b)
    lines.append("")

    # --- Changes vs Last Report ---
    lines.append("## Changes vs Last Report\n")
    if changes:
        lines.append(f"Comparing against snapshot from **{changes['snapshot_date']}**.\n")
        lines.append(f"| Metric | Change |")
        lines.append(f"|--------|--------|")
        if changes["new_stocks"]:
            lines.append(f"| New stocks enrolled | {', '.join(changes['new_stocks'][:10])} ({len(changes['new_stocks'])} total) |")
        if changes["removed_stocks"]:
            lines.append(f"| Stocks removed | {', '.join(changes['removed_stocks'][:10])} |")
        if changes["new_funds"]:
            lines.append(f"| New funds tracked | {len(changes['new_funds'])} |")
        d = changes
        lines.append(f"| Edges (total) | {d['edge_delta']:+d} (now {d['curr_edges']}) |")
        lines.append(f"| Holdings edges | {d['holds_delta']:+d} |")
        lines.append(f"| Short edges | {d['shorts_delta']:+d} |")
        lines.append(f"| Insider edges | {d['insider_delta']:+d} |")
    else:
        lines.append("No previous snapshot available for comparison. Run `snapshot` to establish a baseline.")
    lines.append("")

    # --- Active Signals ---
    lines.append("## Active Signals\n")
    if signals:
        lines.append(f"| Ticker | Signal | Weight | Detail | Portfolio |")
        lines.append(f"|--------|--------|--------|--------|-----------|")
        for sig_type, ticker, weight, detail in signals:
            pf_marker = "[PF]" if ticker in portfolio_tickers else ""
            lines.append(f"| {ticker} | {sig_type} | {weight} | {detail} | {pf_marker} |")
    else:
        lines.append("No signals detected. Graph may need more data (run `refresh`).")
    lines.append("")

    # --- Alerts ---
    lines.append("## Alerts\n")
    if prev_snap:
        lines.append(f"Changes for portfolio positions vs **{prev_snap}**.\n")
    if pf_alerts:
        lines.append(f"| Type | Ticker | Detail |")
        lines.append(f"|------|--------|--------|")
        for alert_type, ticker, msg in pf_alerts:
            priority = "!!" if alert_type in ("CONVERGENCE", "INSIDER_CLUSTER_BUY") else ""
            lines.append(f"| {priority}{alert_type} | {ticker} | {msg} |")
    else:
        lines.append("No portfolio alerts. Graph unchanged vs previous snapshot.")
    lines.append("")

    # --- Crowding Risk ---
    lines.append("## Crowding Risk\n")
    lines.append(f"Top 15 most-held stocks (median holders: {median_holders:.0f}).\n")
    lines.append(f"| Ticker | Name | Holders | Crowding | Portfolio |")
    lines.append(f"|--------|------|---------|----------|-----------|")
    for ticker, name, holders, crowding, in_pf in crowding_data:
        pf_marker = "[PF]" if in_pf else ""
        lines.append(f"| {ticker} | {name[:30]} | {holders} | {crowding:.2f}x | {pf_marker} |")
    lines.append("")

    # --- Discovery ---
    lines.append("## Discovery -- Untracked Opportunities\n")
    lines.append("Stocks held by 3+ quality/value/activist funds NOT in our universe or portfolio.\n")
    if discoveries:
        lines.append(f"| Ticker | Name | Quality Funds | Fund Names |")
        lines.append(f"|--------|------|---------------|------------|")
        for ticker, name, holder_names in discoveries[:20]:
            funds_str = ", ".join(h[:25] for h in holder_names[:4])
            if len(holder_names) > 4:
                funds_str += f" +{len(holder_names)-4} more"
            lines.append(f"| {ticker} | {name[:30]} | {len(holder_names)} | {funds_str} |")
    else:
        lines.append("No untracked stocks held by 3+ quality funds found.")
    lines.append("")

    # --- Insider Activity Summary ---
    lines.append("## Insider Activity Summary\n")
    lines.append(f"**Last 90 days**: {total_buys} buys (${buy_value:,.0f}) vs {total_sells} sells (${sell_value:,.0f})\n")
    net = buy_value - sell_value
    net_label = "NET BUY" if net > 0 else "NET SELL"
    lines.append(f"**Net**: {net_label} ${abs(net):,.0f}\n")
    if insider_clusters:
        lines.append("### Cluster Buys (3+ insiders in 90 days)\n")
        lines.append(f"| Ticker | Insiders | Total Value |")
        lines.append(f"|--------|----------|-------------|")
        for ticker, count, value in insider_clusters:
            pf_marker = " [PF]" if ticker in portfolio_tickers else ""
            lines.append(f"| {ticker}{pf_marker} | {count} | ${value:,.0f} |")
    else:
        lines.append("No insider cluster buys detected (90d window).")
    lines.append("")

    # --- Portfolio Overlay ---
    lines.append("## Portfolio Overlay\n")
    lines.append("Signals and positioning for stocks we own.\n")
    if pf_signals:
        lines.append(f"| Ticker | Signal | Weight | Detail |")
        lines.append(f"|--------|--------|--------|--------|")
        for sig_type, ticker, weight, detail in pf_signals:
            lines.append(f"| {ticker} | {sig_type} | {weight} | {detail} |")
    else:
        lines.append("No active signals for portfolio positions.")

    # Portfolio positions with their holder/short summary
    lines.append("")
    lines.append("### Position Summary\n")
    lines.append(f"| Ticker | Holders | Short Funds | Total SI% | Insider Buys (90d) | Insider Sells (90d) |")
    lines.append(f"|--------|---------|-------------|-----------|--------------------|--------------------|")
    for ticker in sorted(portfolio_tickers):
        if ticker not in G.nodes:
            continue
        n_holders = 0
        n_short_funds = 0
        total_si = 0
        n_insider_buys = 0
        n_insider_sells = 0
        for u, v, k, d in G.in_edges(ticker, data=True, keys=True):
            rel = d.get("relation", "")
            if rel == "holds":
                n_holders += 1
            elif rel == "shorts":
                n_short_funds += 1
                total_si += d.get("pct_shares", 0) or 0
            elif rel == "insider_buy":
                days_ago = days_since(d.get("date", ""))
                if days_ago is not None and days_ago <= 90:
                    n_insider_buys += 1
            elif rel == "insider_sell":
                days_ago = days_since(d.get("date", ""))
                if days_ago is not None and days_ago <= 90:
                    n_insider_sells += 1
        si_str = f"{total_si:.2f}%" if total_si > 0 else "-"
        lines.append(f"| {ticker} | {n_holders} | {n_short_funds} | {si_str} | {n_insider_buys} | {n_insider_sells} |")

    lines.append("")
    lines.append(f"---\n*Generated {report_date} by smart_money.py weekly-report*\n")

    # Write report
    report_path = os.path.join(REPORTS_DIR, f"{report_date}.md")
    with open(report_path, "w") as f:
        f.write("\n".join(lines))

    # Print summary to stdout
    print(f"=== WEEKLY REPORT GENERATED ===\n")
    print(f"  File: {report_path}")
    print(f"  Date: {report_date}")
    print(f"  Graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")
    print(f"  Signals: {len(signals)}")
    print(f"  Alerts: {len(alerts)}")
    print(f"  Discoveries: {len(discoveries)}")
    print(f"  Insider clusters: {len(insider_clusters)}")
    if pf_signals:
        print(f"  Portfolio signals: {len(pf_signals)}")
        for sig_type, ticker, weight, detail in pf_signals[:5]:
            print(f"    [{weight}] {ticker}: {sig_type} - {detail}")
    print()


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------

def main():
    ensure_dirs()

    parser = argparse.ArgumentParser(description="Smart Money Graph — Institutional overlay")
    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # stale
    subparsers.add_parser("stale", help="Show data source staleness")

    # stats
    subparsers.add_parser("stats", help="Graph statistics")

    # sync-portfolio
    subparsers.add_parser("sync-portfolio", help="Sync stock nodes from portfolio")

    # add-node
    p_an = subparsers.add_parser("add-node", help="Add or update a node")
    p_an.add_argument("node_type", choices=["stock", "fund", "person"])
    p_an.add_argument("node_id")
    p_an.add_argument("--attr", nargs="*", help="Attributes as K=V pairs")

    # add-edge
    p_ae = subparsers.add_parser("add-edge", help="Add or update an edge")
    p_ae.add_argument("from_node")
    p_ae.add_argument("to_node")
    p_ae.add_argument("relation", choices=["holds", "shorts", "insider_buy", "insider_sell", "manages"])
    p_ae.add_argument("--attr", nargs="*", help="Attributes as K=V pairs")

    # bulk-update
    subparsers.add_parser("bulk-update", help="Apply batch operations from JSON stdin")

    # snapshot
    subparsers.add_parser("snapshot", help="Save graph snapshot")

    # gc
    subparsers.add_parser("gc", help="Clean up old snapshots and raw files")

    # download
    p_dl = subparsers.add_parser("download", help="Download raw data")
    p_dl.add_argument("source", choices=["fca", "amf", "consob", "afm-nl", "shorts", "13f", "form4", "all"])

    # parse-fca
    p_pf = subparsers.add_parser("parse-fca", help="Parse FCA XLSX to CSV stdout")
    p_pf.add_argument("--file", help="Specific XLSX file path")

    # filter-13f
    p_f13 = subparsers.add_parser("filter-13f", help="Download 13F for specific fund CIK")
    p_f13.add_argument("cik", help="CIK number")

    # parse-13f
    subparsers.add_parser("parse-13f", help="Auto-parse all 13F XMLs, match CUSIPs, output bulk-update JSON")

    # short-interest
    subparsers.add_parser("short-interest", help="Fetch US short interest from yfinance")

    # resolve
    p_res = subparsers.add_parser("resolve", help="Resolve ISIN/CUSIP/CIK for tickers")
    p_res.add_argument("tickers", nargs="*", help="Tickers to resolve (default: all in graph)")
    p_res.add_argument("--force", action="store_true", help="Force re-resolve even if cached")
    p_res.add_argument("--retry-failed", action="store_true", help="Retry only previously failed resolutions (uses OpenFIGI fallback)")
    p_res.add_argument("--purge-invalid", action="store_true", help="Purge ISINs that fail country-prefix validation")

    # harvest-isins
    subparsers.add_parser("harvest-isins", help="Harvest ISINs from FCA/AMF regulatory files")

    # ingest-fca
    p_ifca = subparsers.add_parser("ingest-fca", help="Parse FCA XLSX and apply shorts to graph")
    p_ifca.add_argument("--file", help="Specific XLSX file path")

    # ingest-amf
    subparsers.add_parser("ingest-amf", help="Parse AMF CSV and apply shorts to graph")

    # ingest-13f
    subparsers.add_parser("ingest-13f", help="Parse 13F XMLs and apply holdings to graph")

    # ingest-insider
    p_ii = subparsers.add_parser("ingest-insider", help="Ingest insider transactions from yfinance")
    p_ii.add_argument("tickers", nargs="*", help="Tickers (default: portfolio stocks)")
    p_ii.add_argument("--universe", action="store_true", help="Include top 25 US universe stocks")
    p_ii.add_argument("--all-enrolled", action="store_true", help="All enrolled stocks (max 50)")

    # coverage
    p_cov = subparsers.add_parser("coverage", help="Per-stock data coverage gap report")
    p_cov.add_argument("--portfolio-only", action="store_true")

    # refresh
    p_ref = subparsers.add_parser("refresh", help="One-command full refresh cycle")
    p_ref.add_argument("--full", action="store_true", help="Force download all sources (implies --expand)")
    p_ref.add_argument("--expand", action="store_true", help="Auto-discover and enroll stocks from regulatory data")
    p_ref.add_argument("--skip-download", action="store_true", help="Skip download step")

    # enroll
    p_enr = subparsers.add_parser("enroll", help="Add stocks from universe/pipeline/explicit list")
    p_enr.add_argument("tickers", nargs="*", help="Tickers to enroll")
    p_enr.add_argument("--from-universe", action="store_true")
    p_enr.add_argument("--from-pipeline", action="store_true")
    p_enr.add_argument("--skip-resolve", action="store_true")

    # report
    p_rp = subparsers.add_parser("report", help="Full smart money overlay report")
    p_rp.add_argument("--portfolio-only", action="store_true")

    # stock-profile
    p_sp = subparsers.add_parser("stock-profile", help="Detailed profile for a ticker")
    p_sp.add_argument("ticker")

    # who-holds
    p_wh = subparsers.add_parser("who-holds", help="Which funds hold a ticker")
    p_wh.add_argument("ticker")

    # crowding
    p_cr = subparsers.add_parser("crowding", help="Crowding analysis")
    p_cr.add_argument("--top", type=int, default=20)

    # alerts
    subparsers.add_parser("alerts", help="Changes vs previous snapshot")

    # metrics
    subparsers.add_parser("metrics", help="Graph analytics (PageRank, centrality)")

    # communities
    subparsers.add_parser("communities", help="Community detection (Louvain)")

    # dedup-funds
    subparsers.add_parser("dedup-funds", help="Merge duplicate fund nodes using fund_aliases.yaml")

    # visualize
    p_viz = subparsers.add_parser("visualize", help="Generate interactive HTML visualization")
    p_viz.add_argument("--portfolio-only", action="store_true")
    p_viz.add_argument("--hide-islands", action="store_true", help="Hide stock nodes with zero connections")
    p_viz.add_argument("--output", help="Output HTML file path")

    # --- v3.0 OSINT Engine commands ---

    # signals (WS2)
    p_sig = subparsers.add_parser("signals", help="Detect actionable signals from graph edge patterns")
    p_sig.add_argument("--ticker", help="Analyze specific ticker only")
    p_sig.add_argument("--portfolio-only", action="store_true", help="Only portfolio stocks")

    # discover (WS6)
    p_disc = subparsers.add_parser("discover", help="Find stocks held/shorted by tracked funds NOT in our graph")
    p_disc.add_argument("--source", choices=["13f", "fca", "amf", "all"], default="all")
    p_disc.add_argument("--min-funds", type=int, default=2, help="Minimum funds to flag (default: 2)")

    # ingest-live (WS4)
    p_live = subparsers.add_parser("ingest-live", help="Ingest live intelligence from sessions")
    p_live.add_argument("--type", required=True, choices=["holder", "short", "insider", "mention"])
    p_live.add_argument("--fund", help="Fund name")
    p_live.add_argument("--ticker", required=True, help="Stock ticker")
    p_live.add_argument("--data", help="JSON data string")

    # sector-overlay (WS5)
    p_so = subparsers.add_parser("sector-overlay", help="Institutional positioning overlay for a sector")
    p_so.add_argument("sector", help="Sector name (e.g., insurance, technology)")

    # discover-funds (v4.0)
    p_df = subparsers.add_parser("discover-funds", help="Discover untracked funds connected to 3+ of our stocks")
    p_df.add_argument("--min-stocks", type=int, default=3, help="Min stock connections to surface (default: 3)")
    p_df.add_argument("--auto-enroll", action="store_true", help="Auto-enroll funds with CIK and download 13F (max 5)")

    # capture (v4.1 — Quick-Capture)
    p_cap = subparsers.add_parser("capture", help="Quick-capture session intelligence with natural syntax")
    p_cap.add_argument("text", nargs="+", help="Natural text, e.g. 'Elliott holds 5.2%% LULU'")

    # ingest-consob (WS1)
    p_icon = subparsers.add_parser("ingest-consob", help="Parse CONSOB XLSX and apply shorts to graph (Italy)")
    p_icon.add_argument("--file", help="Specific XLSX file path")

    # ingest-afm-nl (WS1)
    p_iafm = subparsers.add_parser("ingest-afm-nl", help="Parse AFM NL CSV and apply shorts to graph (Netherlands)")
    p_iafm.add_argument("--file", help="Specific CSV file path")

    # weekly-report
    subparsers.add_parser("weekly-report", help="Generate comprehensive markdown weekly report")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    commands = {
        "stale": cmd_stale,
        "stats": cmd_stats,
        "sync-portfolio": cmd_sync_portfolio,
        "add-node": cmd_add_node,
        "add-edge": cmd_add_edge,
        "bulk-update": cmd_bulk_update,
        "snapshot": cmd_snapshot,
        "gc": cmd_gc,
        "download": cmd_download,
        "parse-fca": cmd_parse_fca,
        "filter-13f": cmd_filter_13f,
        "parse-13f": cmd_parse_13f,
        "short-interest": cmd_short_interest,
        "resolve": cmd_resolve,
        "harvest-isins": cmd_harvest_isins,
        "ingest-fca": cmd_ingest_fca,
        "ingest-amf": cmd_ingest_amf,
        "ingest-13f": cmd_ingest_13f,
        "ingest-insider": cmd_ingest_insider,
        "coverage": cmd_coverage,
        "refresh": cmd_refresh,
        "enroll": cmd_enroll,
        "report": cmd_report,
        "stock-profile": cmd_stock_profile,
        "who-holds": cmd_who_holds,
        "crowding": cmd_crowding,
        "alerts": cmd_alerts,
        "metrics": cmd_metrics,
        "communities": cmd_communities,
        "dedup-funds": cmd_dedup_funds,
        "visualize": cmd_visualize,
        "signals": cmd_signals,
        "discover": cmd_discover,
        "discover-funds": cmd_discover_funds,
        "ingest-live": cmd_ingest_live,
        "capture": cmd_capture,
        "sector-overlay": cmd_sector_overlay,
        "ingest-consob": cmd_ingest_consob,
        "ingest-afm-nl": cmd_ingest_afm_nl,
        "weekly-report": cmd_weekly_report,
    }

    cmd_func = commands.get(args.command)
    if cmd_func:
        cmd_func(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
